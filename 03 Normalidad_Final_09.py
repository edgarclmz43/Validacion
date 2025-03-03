"""
===========================================================
Aplicación de Análisis de Normalidad y Detección de Outliers
===========================================================

Propósito:
    Esta aplicación permite analizar la normalidad y detectar outliers
    en archivos Excel. Se evalúan pruebas estadísticas (Shapiro-Wilk,
    Anderson-Darling y Grubbs) y se aplican transformaciones (Log, Box-Cox,
    Yeo-Johnson) en caso de que los datos no sigan una distribución normal.
    Los resultados, incluyendo gráficos y conclusiones, se guardan en una
    nueva hoja denominada "Normalidad" dentro de cada archivo Excel analizado.

Instrucciones de uso:
    1. Indique la ruta de la carpeta que contiene los archivos Excel, ya sea
       escribiéndola en el campo correspondiente o seleccionándola a través del
       botón "Seleccionar Carpeta".
    2. Una vez seleccionada la carpeta, haga clic en "Procesar Datos" para iniciar
       el análisis.
    3. Los resultados se guardarán en la hoja "Normalidad" de cada archivo Excel.

Nota: No se eliminan los outliers durante el análisis principal.
"""

import os
from io import BytesIO
from tkinter import Tk, filedialog, Label, Entry, Button, Frame, END, messagebox

import numpy as np
import pandas as pd
import scipy.stats as stats
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image
from sklearn.preprocessing import PowerTransformer

# -------------------------
# FUNCIÓN AUXILIAR PARA p-valor en AD
# -------------------------
def ad_test_pvalue(stat, n):
    """
    Aproxima el p-valor para la prueba de Anderson-Darling para normalidad.
    Se utiliza la transformación A2* = stat*(1 + 0.75/n + 2.25/n**2) y fórmulas aproximadas.
    Referencia: Stephens (1974) y otras implementaciones comunes.
    """
    A2_star = stat * (1 + 0.75 / n + 2.25 / n**2)
    if A2_star < 0.2:
        p = 1 - np.exp(-13.436 + 101.14 * A2_star - 223.73 * A2_star**2)
    elif A2_star < 0.34:
        p = 1 - np.exp(-8.318 + 42.796 * A2_star - 59.938 * A2_star**2)
    elif A2_star < 0.6:
        p = np.exp(0.9177 - 4.279 * A2_star - 1.38 * A2_star**2)
    else:
        p = np.exp(1.2937 - 5.709 * A2_star + 0.0186 * A2_star**2)
    return p

# -------------------------
# FUNCIONES DE APOYO
# -------------------------
def seleccionar_carpeta():
    """Abre una ventana para seleccionar una carpeta y retorna la ruta."""
    root = Tk()
    root.withdraw()
    return filedialog.askdirectory(title="Seleccionar carpeta con archivos de Excel")

def eliminar_outliers_iqr(datos):
    """
    Elimina outliers usando el método del rango intercuartílico (IQR).
    (Función auxiliar; en esta variante no se elimina para el análisis principal.)
    """
    Q1 = datos.quantile(0.25)
    Q3 = datos.quantile(0.75)
    IQR = Q3 - Q1
    limite_inferior = Q1 - 1.5 * IQR
    limite_superior = Q3 + 1.5 * IQR
    outliers = datos[(datos < limite_inferior) | (datos > limite_superior)]
    datos_sin_outliers = datos[~datos.isin(outliers)]
    return datos_sin_outliers, len(outliers)

# -------------------------
# PRUEBAS DE NORMALIDAD (Separadas)
# -------------------------
def pruebas_de_normalidad(datos):
    """
    Evalúa la normalidad de 'datos' separando la prueba de Shapiro-Wilk y la de Anderson-Darling.
    Se retorna un diccionario con dos sub-diccionarios ("Shapiro" y "Anderson") que incluyen:
      - Estadístico, p-valor, hipótesis y conclusión.
    Además se retorna una conclusión global.
    """
    resultados = {}
    alpha = 0.05
    n = len(datos)

    # Shapiro–Wilk
    shapiro_stat, shapiro_p = stats.shapiro(datos)
    shapiro_result = {
        "Estadístico": shapiro_stat,
        "p-valor": shapiro_p,
        "Hipótesis": "H0: Los datos siguen una distribución normal; H1: No siguen una distribución normal.",
        "Conclusión": "No se rechaza H0 (normal)" if shapiro_p > alpha else "Se rechaza H0 (no normal)"
    }

    # Anderson–Darling
    ad_result = stats.anderson(datos, dist='norm')
    ad_stat = ad_result.statistic
    sig_levels = list(ad_result.significance_level)
    idx = sig_levels.index(5) if 5 in sig_levels else np.argmin(np.abs(np.array(sig_levels) - 5))
    critical_value = ad_result.critical_values[idx]
    ad_p = ad_test_pvalue(ad_stat, n)
    anderson_result = {
        "Estadístico": ad_stat,
        "Valor Crítico (alfa=5%)": critical_value,
        "p-valor": ad_p,
        "Hipótesis": "H0: Los datos siguen una distribución normal; H1: No siguen una distribución normal.",
        "Conclusión": "No se rechaza H0 (normal)" if ad_stat < critical_value else "Se rechaza H0 (no normal)"
    }

    resultados["Shapiro"] = shapiro_result
    resultados["Anderson"] = anderson_result
    if (shapiro_p > alpha) and (ad_stat < critical_value):
        resultados["Conclusión Global"] = ("Las Muestras parecen pertenecer a una Distribución Normal. "
                                            "Se recomienda utilizar estadísticos paramétricos.")
    else:
        resultados["Conclusión Global"] = ("Las Muestras parecen No pertenecer a una Distribución normal. "
                                            "Se recomienda utilizar estadísticos no paramétricos.")

    return resultados

def aplicar_transformaciones(datos):
    """
    Intenta transformaciones (Log, Box-Cox, Yeo-Johnson) sobre 'datos' y evalúa la normalidad con las pruebas anteriores.
    Retorna un diccionario con los resultados de cada transformación.
    """
    resultados_transformaciones = {}
    datos_mod = datos.copy()

    # Ajuste mínimo si hay poca variabilidad
    if datos_mod.nunique() <= 1:
        if len(datos_mod) > 1:
            datos_mod.iloc[0] += 1e-6
            datos_mod.iloc[1] -= 1e-6
        else:
            datos_mod.iloc[0] += 1e-6

    todos_positivos = (datos_mod > 0).all()

    # Transformación Logarítmica
    if todos_positivos:
        try:
            log_data = np.log(datos_mod)
            res_log = pruebas_de_normalidad(log_data)
            if datos.nunique() <= 1:
                res_log['Nota'] = "Se aplicó un ajuste mínimo para romper la constancia."
            resultados_transformaciones['Log'] = res_log
        except Exception as e:
            resultados_transformaciones['Log'] = {'Conclusión Global': f'Error en Log: {e}'}
    else:
        resultados_transformaciones['Log'] = {'Conclusión Global': 'No aplicable (datos <= 0)'}

    # Transformación Box-Cox
    if todos_positivos:
        try:
            bc_data, lambda_opt = stats.boxcox(datos_mod)
            res_bc = pruebas_de_normalidad(bc_data)
            res_bc['Lambda'] = lambda_opt
            if datos.nunique() <= 1:
                res_bc['Nota'] = "Se aplicó un ajuste mínimo para romper la constancia."
            resultados_transformaciones['Box-Cox'] = res_bc
        except Exception as e:
            resultados_transformaciones['Box-Cox'] = {'Conclusión Global': f'Error en Box-Cox: {e}'}
    else:
        resultados_transformaciones['Box-Cox'] = {'Conclusión Global': 'No aplicable (datos <= 0)'}

    # Transformación Yeo-Johnson
    try:
        pt = PowerTransformer(method='yeo-johnson')
        yj_data = pt.fit_transform(datos_mod.values.reshape(-1, 1)).flatten()
        res_yj = pruebas_de_normalidad(yj_data)
        res_yj['Lambda'] = pt.lambdas_[0]
        if datos.nunique() <= 1:
            res_yj['Nota'] = "Se aplicó un ajuste mínimo para romper la constancia."
        resultados_transformaciones['Yeo-Johnson'] = res_yj
    except Exception as e:
        resultados_transformaciones['Yeo-Johnson'] = {'Conclusión Global': f'Error en Yeo-Johnson: {e}'}

    return resultados_transformaciones

# -------------------------
# PRUEBAS DE GRUBBS
# -------------------------
def grubbs_test_unico(datos, alpha=0.05):
    """
    Prueba de Grubbs para datos únicos.
    Calcula:
      - G = |X_extremo - X̄| / s (para el valor más extremo)
      - G_crit según la fórmula basada en la distribución t.
      - p-valor aproximado usando la transformación a t.
    Retorna un diccionario con la información y un mensaje evaluativo.
    """
    datos = np.array(datos)
    n = len(datos)
    if n < 3:
        return {"Error": "La prueba de Grubbs requiere al menos 3 datos."}

    X_mean = np.mean(datos)
    s = np.std(datos, ddof=1)
    diff = np.abs(datos - X_mean)
    idx_extremo = np.argmax(diff)
    outlier = datos[idx_extremo]
    G = diff[idx_extremo] / s

    t_crit = stats.t.ppf(1 - alpha / (2 * n), df=n - 2)
    G_crit = ((n - 1) / np.sqrt(n)) * np.sqrt(t_crit**2 / (n - 2 + t_crit**2))

    t_val = ((n - 1) * G) / np.sqrt(n * (n - 2 + G**2))
    p_val = n * (1 - stats.t.cdf(t_val, df=n - 2))

    if G > G_crit:
        mensaje = (
            "Prueba de Grubbs para datos únicos:\n"
            f"Según la prueba, el valor {outlier:.6f} es sospechoso de ser un outlier significativo a un nivel de confianza del 95%.\n"
            f"p-valor: {p_val:.2e}\n"
            f"Como G = {G:.4f} > G_crit = {G_crit:.4f}, se rechaza H0.\n"
            "Conclusión: La muestra tiene un posible dato atípico en un extremo."
        )
    else:
        mensaje = (
            "Prueba de Grubbs para datos únicos:\n"
            f"No se detecta evidencia suficiente para considerar el valor {outlier:.6f} como outlier (G = {G:.4f} ≤ G_crit = {G_crit:.4f})."
        )

    return {
        "X̄": X_mean,
        "s": s,
        "G": G,
        "G_crit": G_crit,
        "p-valor": p_val,
        "outlier": outlier,
        "Hipótesis": "H0: Ningún valor es atípico (datos siguen distribución normal) vs H1: El valor extremo es atípico.",
        "mensaje": mensaje
    }

def grubbs_test_extremos(datos, alpha=0.05):
    """
    Prueba de Grubbs para datos en cada extremo.
    Calcula:
      - G_low = (X̄ - X_min) / s y G_high = (X_max - X_mean) / s
      - Para cada extremo, se calcula un p-valor aproximado.
    Se devuelven siempre "outlier_low" = X_min y "outlier_high" = X_max,
    de modo que en el reporte Excel se muestre cuál es el valor mínimo y el valor máximo,
    independientemente de que sean o no considerados outliers.
    """
    datos = np.array(datos)
    n = len(datos)
    if n < 3:
        return {"Error": "La prueba de Grubbs requiere al menos 3 datos."}

    X_mean = np.mean(datos)
    s = np.std(datos, ddof=1)
    X_min = np.min(datos)
    X_max = np.max(datos)

    G_low = (X_mean - X_min) / s
    G_high = (X_max - X_mean) / s

    t_crit = stats.t.ppf(1 - alpha / (2 * n), df=n - 2)
    G_crit = ((n - 1) / np.sqrt(n)) * np.sqrt(t_crit**2 / (n - 2 + t_crit**2))

    t_val_low = ((n - 1) * G_low) / np.sqrt(n * (n - 2 + G_low**2)) if G_low != 0 else 0
    p_val_low = n * (1 - stats.t.cdf(t_val_low, df=n - 2)) if G_low > 0 else 1

    t_val_high = ((n - 1) * G_high) / np.sqrt(n * (n - 2 + G_high**2)) if G_high != 0 else 0
    p_val_high = n * (1 - stats.t.cdf(t_val_high, df=n - 2)) if G_high > 0 else 1

    is_outlier_low = (G_low > G_crit)
    is_outlier_high = (G_high > G_crit)

    if is_outlier_low and is_outlier_high:
        mensaje = (
            "Prueba de Grubbs para datos en cada extremo:\n"
            f"Según la prueba, los valores {X_min:.6f} (extremo inferior) y {X_max:.6f} (extremo superior) "
            f"son sospechosos de ser outliers a un nivel de confianza del 95%.\n"
            f"p-valor (extremo bajo): {p_val_low:.2e}, p-valor (extremo alto): {p_val_high:.2e}\n"
            f"Como G_low = {G_low:.4f} y G_high = {G_high:.4f} > G_crit = {G_crit:.4f}, se rechaza H0.\n"
            "Conclusión: La muestra tiene posibles outliers en ambos extremos."
        )
    elif is_outlier_low:
        mensaje = (
            "Prueba de Grubbs para datos en cada extremo:\n"
            f"El valor {X_min:.6f} (extremo inferior) es sospechoso de ser outlier a un nivel de confianza del 95%, "
            f"pero no el extremo superior.\n"
            f"p-valor (extremo bajo): {p_val_low:.2e}, p-valor (extremo alto): {p_val_high:.2e}\n"
            f"Como G_low = {G_low:.4f} > G_crit = {G_crit:.4f}, se rechaza H0 para el extremo inferior.\n"
            "Conclusión: Se detecta un posible outlier en el extremo inferior."
        )
    elif is_outlier_high:
        mensaje = (
            "Prueba de Grubbs para datos en cada extremo:\n"
            f"El valor {X_max:.6f} (extremo superior) es sospechoso de ser outlier a un nivel de confianza del 95%, "
            f"pero no el extremo inferior.\n"
            f"p-valor (extremo bajo): {p_val_low:.2e}, p-valor (extremo alto): {p_val_high:.2e}\n"
            f"Como G_high = {G_high:.4f} > G_crit = {G_crit:.4f}, se rechaza H0 para el extremo superior.\n"
            "Conclusión: Se detecta un posible outlier en el extremo superior."
        )
    else:
        mensaje = (
            "Prueba de Grubbs para datos en cada extremo:\n"
            f"No se detecta evidencia suficiente para considerar los extremos como outliers.\n"
            f"(G_low = {G_low:.4f}, G_high = {G_high:.4f} ≤ G_crit = {G_crit:.4f})"
        )

    return {
        "G_low": G_low,
        "p-valor_low": p_val_low,
        "outlier_low": X_min,
        "G_high": G_high,
        "p-valor_high": p_val_high,
        "outlier_high": X_max,
        "G_crit": G_crit,
        "mensaje": mensaje
    }

# -------------------------
# ANÁLISIS DE NORMALIDAD Y OUTLIERS (SIN ELIMINAR LOS OUTLIERS)
# -------------------------
def analizar_normalidad_y_outliers(datos, nivel):
    """
    Realiza el análisis de normalidad y outliers para 'datos' sin eliminar los outliers.
    Se evalúa la normalidad usando la muestra original, se intentan transformaciones si es necesario,
    y se aplican las pruebas de Grubbs para detectar posibles outliers (tanto un único dato como en ambos extremos).
    Se generan gráficas y se retornan en 'Gráficas'.
    """
    resultados = {}
    resultados['Número de Outliers (IQR)'] = "No se eliminaron outliers"

    # Pruebas de normalidad
    res_normales = pruebas_de_normalidad(datos)
    resultados['Pruebas Normalidad'] = res_normales

    # Aplicar transformaciones si la conclusión global indica "No normal"
    if "No normal" in res_normales["Conclusión Global"]:
        resultados['Transformaciones'] = aplicar_transformaciones(datos)
    else:
        resultados['Transformaciones'] = {}

    # Pruebas de Grubbs
    grubbs_unico = grubbs_test_unico(datos)
    grubbs_extremos = grubbs_test_extremos(datos)
    resultados["Pruebas Outliers Grubbs"] = {
        "Grubbs Unico": grubbs_unico,
        "Grubbs Extremos": grubbs_extremos
    }

    # Generar gráficas usando subplots
    fig, axs = plt.subplots(2, 3, figsize=(18, 10))

    # Fila 1: Análisis de la muestra original
    sns.histplot(datos, kde=True, bins=10, color='skyblue', ax=axs[0, 0])
    axs[0, 0].set_title(f'Histograma - {nivel} (Original)')
    axs[0, 0].set_xlabel('Valores')
    axs[0, 0].set_ylabel('Frecuencia')

    stats.probplot(datos, dist="norm", plot=axs[0, 1])
    axs[0, 1].set_title(f'Q-Q - {nivel} (Original)')

    sns.boxplot(x=datos, color='lightgreen', ax=axs[0, 2])
    axs[0, 2].set_title(f'Boxplot - {nivel} (Original)')

    # Fila 2: Repetición sin eliminación (mismos gráficos)
    sns.histplot(datos, kde=True, bins=10, color='skyblue', ax=axs[1, 0])
    axs[1, 0].set_title(f'Histograma - {nivel} (Sin eliminación)')
    axs[1, 0].set_xlabel('Valores')
    axs[1, 0].set_ylabel('Frecuencia')

    stats.probplot(datos, dist="norm", plot=axs[1, 1])
    axs[1, 1].set_title(f'Q-Q - {nivel} (Sin eliminación)')

    sns.boxplot(x=datos, color='lightgreen', ax=axs[1, 2])
    axs[1, 2].set_title(f'Boxplot - {nivel} (Sin eliminación)')

    plt.tight_layout()
    imgdata = BytesIO()
    fig.savefig(imgdata, format='png', bbox_inches='tight')
    plt.close(fig)
    imgdata.seek(0)
    resultados['Gráficas'] = imgdata

    return resultados

# -------------------------
# FUNCIÓN PRINCIPAL
# -------------------------
def main(folder=None):
    """
    Función principal para analizar normalidad y outliers (sin eliminar datos) en archivos Excel.
    Se guardan los resultados en una hoja "Normalidad" que incluye:
      - Resultados de las pruebas de Shapiro-Wilk y Anderson-Darling (separadas).
      - Resultados de las pruebas de Grubbs (con p-valores, hipótesis y conclusiones, y mensajes con títulos).
      - Un resumen final con la recomendación de utilizar estadísticos paramétricos (si es normal)
        o no paramétricos (si no es normal).

    Si se proporciona el parámetro 'folder', se utiliza esa carpeta; de lo contrario,
    se solicitará al usuario seleccionar la carpeta.
    """
    if folder is None:
        folder = seleccionar_carpeta()
    if not folder:
        print("No se seleccionó ninguna carpeta. Terminando ejecución.")
        return

    extensiones_validas = ('.xlsx', '.xls', '.xlsm')
    archivos_excel = [os.path.join(folder, f) for f in os.listdir(folder)
                      if f.lower().endswith(extensiones_validas)]
    
    if not archivos_excel:
        print("No se encontraron archivos Excel en la carpeta seleccionada.")
        return

    hoja_original = 'sr ysR(Método) ISO 5725'

    for archivo_excel in archivos_excel:
        print(f"Procesando archivo: {archivo_excel}")
        resumen_conclusiones = []  # Reinicializar resumen para cada archivo
        try:
            df = pd.read_excel(archivo_excel, sheet_name=hoja_original, usecols="G:I", skiprows=9, nrows=40)
        except Exception as e:
            print(f"Error al leer la hoja '{hoja_original}' en '{archivo_excel}': {e}")
            continue

        df.columns = ['Nivel 1', 'Nivel 2', 'Nivel 3']
        resultados_niveles = {}

        for nivel in df.columns:
            datos = df[nivel].dropna()
            resultados = analizar_normalidad_y_outliers(datos, nivel)
            resultados_niveles[nivel] = resultados

            res_norm = resultados.get("Pruebas Normalidad", {})
            concl_global = res_norm.get("Conclusión Global", "")
            if "Normal" in concl_global:
                recomendacion = "Estadísticos paramétricos."
            else:
                recomendacion = "Estadísticos no paramétricos."
            resumen_conclusiones.append({
                'Nivel': nivel,
                'Conclusión': concl_global,
                'Recomendación Estadístico': recomendacion
            })

        try:
            wb = load_workbook(archivo_excel)
        except Exception as e:
            print(f"Error al cargar el archivo '{archivo_excel}': {e}")
            continue

        # Eliminar hojas antiguas si existen
        hoja_a_eliminar = 'Analisis total de Normalidad'
        if hoja_a_eliminar in wb.sheetnames:
            del wb[hoja_a_eliminar]
        if 'Normalidad' in wb.sheetnames:
            ws = wb['Normalidad']
            wb.remove(ws)
        ws = wb.create_sheet(title='Normalidad', index=2)

        # Estilos
        titulo_font = Font(bold=True, size=16)
        subtitulo_font = Font(bold=True, size=14)
        encabezado_font = Font(bold=True, size=12)
        encabezado_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        current_row = 1

        ws.insert_rows(1, 3)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        titulo = ws.cell(row=2, column=1, value="Estudio de Normalidad")
        titulo.font = Font(bold=True, size=20, color="1F4E78")
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 30
        current_row = 4

        # Escribir resultados para cada nivel
        for nivel, resultados in resultados_niveles.items():
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            cell = ws.cell(row=current_row, column=1, value=f'Análisis de Normalidad y Outliers - {nivel}')
            cell.font = titulo_font
            cell.alignment = left_alignment
            current_row += 1

            ws.cell(row=current_row, column=1, value='Prueba').font = encabezado_font
            ws.cell(row=current_row, column=1).fill = encabezado_fill
            ws.cell(row=current_row, column=1).alignment = center_alignment
            ws.cell(row=current_row, column=2, value='Resultado').font = encabezado_font
            ws.cell(row=current_row, column=2).fill = encabezado_fill
            ws.cell(row=current_row, column=2).alignment = center_alignment
            current_row += 1

            ws.cell(row=current_row, column=1, value='Número de Outliers (IQR)').alignment = left_alignment
            ws.cell(row=current_row, column=2, value="No se eliminaron outliers").alignment = left_alignment
            current_row += 1

            # Resultados Shapiro-Wilk
            shapiro_res = resultados['Pruebas Normalidad'].get("Shapiro", {})
            ws.cell(row=current_row, column=1, value="Shapiro-Wilk Estadístico").alignment = left_alignment
            ws.cell(row=current_row, column=2, value=shapiro_res.get("Estadístico")).alignment = left_alignment
            current_row += 1
            ws.cell(row=current_row, column=1, value="Shapiro-Wilk p-valor").alignment = left_alignment
            ws.cell(row=current_row, column=2, value=shapiro_res.get("p-valor")).alignment = left_alignment
            current_row += 1
            ws.cell(row=current_row, column=1, value="Hipótesis (Shapiro)").alignment = left_alignment
            ws.cell(row=current_row, column=2, value=shapiro_res.get("Hipótesis")).alignment = left_alignment
            current_row += 1
            ws.cell(row=current_row, column=1, value="Conclusión (Shapiro)").alignment = left_alignment
            ws.cell(row=current_row, column=2, value=shapiro_res.get("Conclusión")).alignment = left_alignment
            current_row += 1

            # Resultados Anderson-Darling
            anderson_res = resultados['Pruebas Normalidad'].get("Anderson", {})
            ws.cell(row=current_row, column=1, value="Anderson-Darling Estadístico").alignment = left_alignment
            ws.cell(row=current_row, column=2, value=anderson_res.get("Estadístico")).alignment = left_alignment
            current_row += 1
            ws.cell(row=current_row, column=1, value="Valor Crítico (alfa=5%)").alignment = left_alignment
            ws.cell(row=current_row, column=2, value=anderson_res.get("Valor Crítico (alfa=5%)")).alignment = left_alignment
            current_row += 1
            ws.cell(row=current_row, column=1, value="Anderson p-valor").alignment = left_alignment
            ws.cell(row=current_row, column=2, value=anderson_res.get("p-valor")).alignment = left_alignment
            current_row += 1
            ws.cell(row=current_row, column=1, value="Hipótesis (Anderson)").alignment = left_alignment
            ws.cell(row=current_row, column=2, value=anderson_res.get("Hipótesis")).alignment = left_alignment
            current_row += 1
            ws.cell(row=current_row, column=1, value="Conclusión (Anderson)").alignment = left_alignment
            ws.cell(row=current_row, column=2, value=anderson_res.get("Conclusión")).alignment = left_alignment
            current_row += 1

            ws.cell(row=current_row, column=1, value='Conclusión Global').alignment = left_alignment
            ws.cell(row=current_row, column=2, value=resultados['Pruebas Normalidad'].get("Conclusión Global")).alignment = left_alignment
            current_row += 1

            # Transformaciones (si existen)
            if resultados['Transformaciones']:
                ws.cell(row=current_row, column=1, value="Transformaciones:").font = subtitulo_font
                current_row += 1
                for nombre_tf, res_tf in resultados['Transformaciones'].items():
                    ws.cell(row=current_row, column=1, value=f"- {nombre_tf}").font = encabezado_font
                    current_row += 1
                    tf_filas = [
                        ('Shapiro-Wilk Estadístico', res_tf.get('Shapiro-Wilk Estadístico')),
                        ('Shapiro-Wilk p-valor', res_tf.get('Shapiro-Wilk p-valor')),
                        ('Anderson-Darling Estadístico', res_tf.get("Anderson-Darling Estadístico")),
                        ('AD Conclusión', res_tf.get("AD Conclusión")),
                        ('Conclusión Global', res_tf.get("Conclusión Global"))
                    ]
                    if 'Lambda' in res_tf:
                        tf_filas.append(('Lambda', res_tf['Lambda']))
                    if 'Nota' in res_tf:
                        tf_filas.append(('Nota', res_tf['Nota']))
                    for etiqueta, valor in tf_filas:
                        ws.cell(row=current_row, column=1, value=etiqueta).alignment = left_alignment
                        ws.cell(row=current_row, column=2, value=valor).alignment = left_alignment
                        current_row += 1
                    current_row += 1

            # Grubbs: dato único
            ws.cell(row=current_row, column=1, value="Prueba de Grubbs para datos únicos").font = subtitulo_font
            current_row += 1
            grubbs_unico = resultados["Pruebas Outliers Grubbs"].get("Grubbs Unico")
            if grubbs_unico is not None:
                ws.cell(row=current_row, column=1, value="G (dato único)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_unico.get("G")).alignment = left_alignment
                current_row += 1
                ws.cell(row=current_row, column=1, value="G crítico (dato único)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_unico.get("G_crit")).alignment = left_alignment
                current_row += 1
                ws.cell(row=current_row, column=1, value="p-valor (dato único)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_unico.get("p-valor")).alignment = left_alignment
                current_row += 1
                ws.cell(row=current_row, column=1, value="Valor atípico (dato único)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_unico.get("outlier")).alignment = left_alignment
                current_row += 1
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+2, end_column=8)
                cell_msg = ws.cell(row=current_row, column=1)
                cell_msg.value = grubbs_unico.get("mensaje")
                cell_msg.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                current_row += 3

            # Grubbs: extremos
            ws.cell(row=current_row, column=1, value="Prueba de Grubbs para datos en cada extremo").font = subtitulo_font
            current_row += 1
            grubbs_extremos = resultados["Pruebas Outliers Grubbs"].get("Grubbs Extremos")
            if grubbs_extremos is not None:
                ws.cell(row=current_row, column=1, value="G (extremo inferior)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_extremos.get("G_low")).alignment = left_alignment
                current_row += 1
                ws.cell(row=current_row, column=1, value="p-valor (extremo inferior)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_extremos.get("p-valor_low")).alignment = left_alignment
                current_row += 1
                ws.cell(row=current_row, column=1, value="G (extremo superior)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_extremos.get("G_high")).alignment = left_alignment
                current_row += 1
                ws.cell(row=current_row, column=1, value="p-valor (extremo superior)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_extremos.get("p-valor_high")).alignment = left_alignment
                current_row += 1
                ws.cell(row=current_row, column=1, value="G crítico (extremos)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_extremos.get("G_crit")).alignment = left_alignment
                current_row += 1
                ws.cell(row=current_row, column=1, value="Valor atípico (extremo inferior)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_extremos.get("outlier_low")).alignment = left_alignment
                current_row += 1
                ws.cell(row=current_row, column=1, value="Valor atípico (extremo superior)").alignment = left_alignment
                ws.cell(row=current_row, column=2, value=grubbs_extremos.get("outlier_high")).alignment = left_alignment
                current_row += 1
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+2, end_column=8)
                cell_msg = ws.cell(row=current_row, column=1)
                cell_msg.value = grubbs_extremos.get("mensaje")
                cell_msg.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                current_row += 3

            # Agregar Gráficas
            imgdata = resultados['Gráficas']
            img = Image.open(imgdata)
            img_byte_arr = BytesIO()
            img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            img_openpyxl = OpenPyxlImage(img_byte_arr)
            scale_factor = 0.5
            img_openpyxl.width *= scale_factor
            img_openpyxl.height *= scale_factor
            img_position = f'I{current_row - 10}'
            ws.add_image(img_openpyxl, img_position)
            current_row += 2

            # Datos originales
            ws.cell(row=current_row, column=1, value='Datos Originales').font = subtitulo_font
            current_row += 1
            ws.cell(row=current_row, column=1, value='Fila').font = encabezado_font
            ws.cell(row=current_row, column=1).fill = encabezado_fill
            ws.cell(row=current_row, column=1).alignment = center_alignment
            ws.cell(row=current_row, column=2, value=nivel).font = encabezado_font
            ws.cell(row=current_row, column=2).fill = encabezado_fill
            ws.cell(row=current_row, column=2).alignment = center_alignment
            current_row += 1

            for i in range(len(df[nivel])):
                valor = df[nivel].iloc[i]
                ws.cell(row=current_row, column=1, value=(i + 11)).alignment = left_alignment
                ws.cell(row=current_row, column=2, value=valor).alignment = left_alignment
                current_row += 1

            current_row += 2

        # Resumen final
        ws.cell(row=current_row, column=1, value='Resumen Final').font = subtitulo_font
        current_row += 1
        ws.cell(row=current_row, column=1, value='Nivel').font = encabezado_font
        ws.cell(row=current_row, column=1).fill = encabezado_fill
        ws.cell(row=current_row, column=1).alignment = center_alignment
        ws.cell(row=current_row, column=2, value='Conclusión').font = encabezado_font
        ws.cell(row=current_row, column=2).fill = encabezado_fill
        ws.cell(row=current_row, column=2).alignment = center_alignment
        ws.cell(row=current_row, column=4, value='Recomendación Estadístico').font = encabezado_font
        ws.cell(row=current_row, column=4).fill = encabezado_fill
        ws.cell(row=current_row, column=4).alignment = center_alignment
        current_row += 1

        for resumen in resumen_conclusiones:
            ws.cell(row=current_row, column=1, value=resumen['Nivel']).alignment = left_alignment
            ws.cell(row=current_row, column=2, value=resumen['Conclusión']).alignment = left_alignment
            ws.cell(row=current_row, column=4, value=resumen['Recomendación Estadístico']).alignment = left_alignment
            current_row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['K'].width = 30

        try:
            wb.save(archivo_excel)
            print(f"Análisis completado. Resultados guardados en la hoja 'Normalidad' del archivo '{archivo_excel}'.")
        except Exception as e:
            print(f"Error al guardar el archivo '{archivo_excel}': {e}")

# -------------------------
# INTERFAZ GRÁFICA DE USUARIO
# -------------------------
def interfaz_app():
    """Crea y embellece la interfaz gráfica para la aplicación."""
    window = Tk()
    window.title("Análisis de Normalidad y Outliers")
    window.geometry("700x500")
    window.resizable(False, False)

    # Contenedor principal
    main_frame = Frame(window, padx=20, pady=20)
    main_frame.pack(fill="both", expand=True)

    # Título de la aplicación
    title_label = Label(main_frame, text="Análisis de Normalidad y Outliers en Archivos Excel",
                        font=("Arial", 18, "bold"))
    title_label.pack(pady=10)

    # Propósito
    purpose_text = (
        "Esta aplicación permite analizar la normalidad y detectar outliers en archivos Excel.\n"
        "Se evaluarán pruebas estadísticas y se generará un reporte en cada archivo."
    )
    purpose_label = Label(main_frame, text=purpose_text, font=("Arial", 12), justify="center")
    purpose_label.pack(pady=5)

    # Instrucciones de uso
    instructions_text = (
        "Instrucciones de uso:\n"
        "1. Indique la ruta de la carpeta que contiene los archivos Excel o use el botón 'Seleccionar Carpeta'.\n"
        "2. Haga clic en 'Procesar Datos' para iniciar el análisis.\n"
        "3. Los resultados se guardarán en la hoja 'Normalidad' de cada archivo Excel."
    )
    instructions_label = Label(main_frame, text=instructions_text, font=("Arial", 10), justify="left")
    instructions_label.pack(pady=5)

    # Sección para seleccionar la carpeta
    folder_frame = Frame(main_frame)
    folder_frame.pack(pady=10)
    folder_label = Label(folder_frame, text="Carpeta:", font=("Arial", 12))
    folder_label.pack(side="left", padx=5)
    folder_entry = Entry(folder_frame, width=40, font=("Arial", 12))
    folder_entry.pack(side="left", padx=5)

    def browse_folder():
        folder = filedialog.askdirectory(title="Seleccionar carpeta con archivos de Excel")
        if folder:
            folder_entry.delete(0, END)
            folder_entry.insert(0, folder)
    browse_button = Button(folder_frame, text="Seleccionar Carpeta", font=("Arial", 10, "bold"), command=browse_folder)
    browse_button.pack(side="left", padx=5)

    # Etiqueta de estado para informar al usuario
    status_label = Label(main_frame, text="", font=("Arial", 12), fg="green")
    status_label.pack(pady=5)

    # Botón para procesar los datos
    def process_data():
        folder = folder_entry.get()
        if not folder:
            folder = filedialog.askdirectory(title="Seleccionar carpeta con archivos de Excel")
            if not folder:
                status_label.config(text="No se seleccionó ninguna carpeta.", fg="red")
                return
            folder_entry.delete(0, END)
            folder_entry.insert(0, folder)
        status_label.config(text="Procesando...", fg="blue")
        window.update_idletasks()
        main(folder)
        status_label.config(text="Procesamiento completado.", fg="green")
        messagebox.showinfo("Proceso Finalizado", "Procesamiento completado.")

    process_button = Button(main_frame, text="Procesar Datos", font=("Arial", 14, "bold"), command=process_data)
    process_button.pack(pady=20)

    # Botón para salir
    exit_button = Button(main_frame, text="Salir", font=("Arial", 12), command=window.destroy)
    exit_button.pack(pady=10)

    window.mainloop()

if __name__ == "__main__":
    interfaz_app()
