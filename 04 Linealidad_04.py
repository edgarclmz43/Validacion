"""
================================================================================
APLICACIÓN PARA EL ANÁLISIS DE LINEALIDAD Y TEST DE SIGNIFICANCIA
================================================================================

DESCRIPCIÓN:
  Esta aplicación en Python está diseñada para automatizar el análisis estadístico de archivos 
  Excel que contienen datos para evaluar la linealidad de un método analítico. La aplicación 
  se divide en dos partes principales:

  1. Análisis de Regresión Lineal (OLS) por segmentos:
     - Carga los datos de la hoja de Excel (por defecto, "Linealidad - no parametrico") a partir 
       de columnas específicas, omitiendo filas de encabezado. Se detecta automáticamente la fila 
       en la que se encuentra la etiqueta "LP" en la columna B y se saltan las filas hasta la siguiente,
       donde comienzan los datos.
     - Separa los datos en segmentos (por ejemplo, valores negativos y positivos) y, para cada segmento:
         • Agrupa y calcula la media para cada valor único de X.
         • Ajusta un modelo de regresión lineal ordinaria (OLS), pudiendo aplicar transformación 
           logarítmica y errores robustos si alguno de los supuestos estadísticos no se cumple.
         • Realiza pruebas sobre los residuales (Shapiro-Wilk, Durbin-Watson, Breusch-Pagan, White) y 
           calcula el coeficiente de correlación junto con el estadístico t.
         • Genera gráficos compuestos (dispersión con línea de regresión, Q-Q plot y gráfico de escala) y 
           los guarda en archivos temporales, insertándolos en la hoja de Excel.
         • Escribe los resultados (coeficientes, bondad de ajuste, análisis de residuales, correlación y 
           conclusiones) en la misma hoja de Excel utilizando celdas combinadas, tablas y textos multilínea.

  2. Test de Significancia – Linealidad:
     - Lee y limpia los datos de la hoja "Linealidad - no parametrico" (columnas B y C), utilizando la 
       detección automática de la fila de etiquetas y saltando hasta la fila siguiente para que comiencen los datos.
     - Elimina outliers utilizando el método IQR.
     - Calcula el coeficiente de correlación, realiza ANOVA, pruebas t y una prueba de Falta de Ajuste (Lack of Fit).
     - Genera gráficos (por ejemplo, el gráfico de dispersión y regresión) y escribe los resultados en una 
       nueva hoja de Excel denominada "Test Significancia - Linealidad", en la que se incluyen las hipótesis, 
       el análisis estadístico y las conclusiones.

USO:
  - Al ejecutar el script se despliega una interfaz gráfica (usando tkinter) que solicita al usuario la carpeta 
    que contiene los archivos Excel a procesar.
  - Para cada archivo Excel se ejecuta en secuencia:
      • El análisis OLS por segmentos (App 1).
      • El Test de Significancia - Linealidad (App 2).
  - Además, se incluye una función de prueba que permite listar en consola los datos que se usan para el análisis,
    de modo que puedas revisar la tabla de valores (por ejemplo, para verificar el intercepto obtenido).
  - Al finalizar, los resultados se guardan en los archivos Excel originales y se eliminan los archivos de imagen 
    temporales.

REQUISITOS Y DEPENDENCIAS:
  - Python 3.x
  - Librerías: os, glob, numpy, pandas, statsmodels, matplotlib, seaborn, tkinter, scipy y openpyxl.
  - Asegúrese de tener instaladas estas librerías para el correcto funcionamiento de la aplicación.

OBSERVACIONES IMPORTANTES:
  - Algunas características implementadas (como celdas combinadas, inserción de imágenes y formatos de Excel) pueden 
    no ser compatibles con el autoguardado automático de algunos editores o sistemas, lo que podría dañar el archivo 
    al guardar. Se recomienda desactivar el autoguardado o realizar copias de seguridad antes de guardar.
  - Se generan archivos temporales para los gráficos, que se eliminan una vez finalizado el procesamiento de cada archivo.

AUTORES:
  - (Incluir aquí los nombres de los desarrolladores o responsables, según corresponda)

================================================================================
"""

import os
import glob
import numpy as np
import pandas as pd
import statsmodels.api as sm
import statsmodels.formula.api as smf
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
import tkinter as tk
from tkinter import Tk, filedialog, messagebox
from scipy.stats import shapiro, t
from statsmodels.stats.stattools import durbin_watson
from statsmodels.stats.diagnostic import het_breuschpagan, het_white
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as ImgExcel
from openpyxl.drawing.image import Image as Img
from scipy import stats

# =============================================================================
# CONFIGURACIONES PARA LA PRIMERA APP: Análisis de Linealidad (Paramétrico)
# =============================================================================
PATRON_EXCEL = "*.xlsx"                    # Se procesarán todos los archivos .xlsx
NOMBRE_HOJA = "Linealidad - no parametrico"  # Hoja en la que se encuentra la información
COLUMNA_X = "B"      # Columna donde se encuentran los datos de la variable X
COLUMNA_Y = "C"      # Columna donde se encuentran los datos de la variable Y

def buscar_fila_encabezado(ruta_archivo, sheet_name):
    """
    Busca en la columna B la primera celda que contenga 'LP' (sin distinguir mayúsculas)
    y devuelve el número de filas a saltar (índice, donde la primera fila es 0).
    Si no se encuentra, devuelve 6 (lo que equivale a que la cabecera está en la fila 7).
    """
    df = pd.read_excel(ruta_archivo, sheet_name=sheet_name, header=None)
    for i, valor in enumerate(df.iloc[:, 1]):
        if isinstance(valor, str) and "LP" in valor.upper():
            return i
    return 6

FILA_INICIO_EXCEL = 40    # Fila en la hoja Excel donde se empezará a escribir los resultados
COLUMNA_INICIO_EXCEL = 5   # Columna en la hoja Excel donde se empezará a escribir los resultados

# =============================================================================
# Funciones para el Análisis de Linealidad (App 1)
# =============================================================================
def escribir_encabezado_combinado(ws, row, col_start, col_end, texto, fill_color="FFD966"):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    celda = ws.cell(row=row, column=col_start)
    celda.value = texto
    celda.font = Font(name="Calibri", size=12, bold=True)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

def escribir_tabla_dict(ws, start_row, start_col, dict_data, titulo=None):
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    row_actual = start_row
    if titulo:
        celda = ws.cell(row=row_actual, column=start_col, value=titulo)
        celda.font = Font(name="Calibri", size=11, bold=True)
        row_actual += 1
    ws.cell(row=row_actual, column=start_col, value="Parámetro").font = Font(bold=True)
    ws.cell(row=row_actual, column=start_col+1, value="Valor").font = Font(bold=True)
    ws.cell(row=row_actual, column=start_col).border = thin_border
    ws.cell(row=row_actual, column=start_col+1).border = thin_border
    row_actual += 1
    for k, v in dict_data.items():
        ws.cell(row=row_actual, column=start_col, value=str(k)).border = thin_border
        ws.cell(row=row_actual, column=start_col+1, value=str(v)).border = thin_border
        row_actual += 1
    return row_actual + 1

def escribir_texto_multilinea(ws, start_row, col, texto, font_size=10):
    lines = texto.splitlines()
    r = start_row
    for line in lines:
        celda = ws.cell(row=r, column=col, value=line)
        celda.font = Font(name="Calibri", size=font_size)
        celda.alignment = Alignment(wrap_text=True)
        r += 1
    return r + 1

def insertar_imagen_excel(ws, image_path, cell_address):
    if not os.path.exists(image_path):
        print(f"No se encontró la imagen: {image_path}")
        return
    img = ImgExcel(image_path)
    ws.add_image(img, cell_address)

def limpiar_hoja(ws):
    for rango in list(ws.merged_cells.ranges):
        if rango.min_row >= 40:
            ws.unmerge_cells(str(rango))
    max_row = ws.max_row
    max_col = ws.max_column
    for row in range(40, max_row + 1):
        for col in range(5, max_col + 1):
            ws.cell(row=row, column=col).value = None

def cargar_datos_ols(ruta_archivo):
    """
    Carga los datos de la hoja especificada en el archivo Excel y los renombra a ["X", "Y"].
    Se detecta automáticamente la fila que contiene las etiquetas (por ejemplo, "LP" en la columna B)
    y se saltan las filas hasta la siguiente, donde comienzan los datos.
    """
    filas_saltar = buscar_fila_encabezado(ruta_archivo, NOMBRE_HOJA) + 0
    df = pd.read_excel(
        ruta_archivo,
        sheet_name=NOMBRE_HOJA,
        usecols=f"{COLUMNA_X}:{COLUMNA_Y}",
        skiprows=filas_saltar
    )
    df.columns = ["X", "Y"]
    return df

def realizar_regresion_OLS(df, usar_robusto=False, transformar=False):
    Xc = sm.add_constant(df["X"])
    if transformar:
        Y = np.log(df["Y"] + 1)
    else:
        Y = df["Y"]
    if usar_robusto:
        modelo = sm.OLS(Y, Xc).fit(cov_type='HC3')
    else:
        modelo = sm.OLS(Y, Xc).fit()
    return modelo

def analizar_residuales(modelo):
    resid = modelo.resid
    sw = shapiro(resid)
    dw_stat = durbin_watson(resid)
    bp = het_breuschpagan(resid, modelo.model.exog)
    white_ = het_white(resid, modelo.model.exog)
    dict_res = {
        "Shapiro-Wilk": f"{sw.statistic:.4f}",
        "p-valor S-W": f"{sw.pvalue:.4f}",
        "Durbin-Watson": f"{dw_stat:.4f}",
        "Breusch-Pagan (p)": f"{bp[1]:.4f}",
        "White (p)": f"{white_[1]:.4f}"
    }
    return dict_res

def extraer_stats_ols(modelo):
    p = modelo.params
    pv = modelo.pvalues
    indices = list(p.index)
    intercept_name = indices[0]
    slope_name = indices[1]
    dict_coefs = {
        "Intercepto": f"{p[intercept_name]:.4E}",
        "Pendiente": f"{p[slope_name]:.4E}",
        "p-valor Intercepto": f"{pv[intercept_name]:.4E}",
        "p-valor Pendiente": f"{pv[slope_name]:.4E}"
    }
    dict_ajuste = {
        "R²": f"{modelo.rsquared:.4f}",
        "R² ajustado": f"{modelo.rsquared_adj:.4f}",
        "AIC": f"{modelo.aic:.4f}",
        "BIC": f"{modelo.bic:.4f}"
    }
    return dict_coefs, dict_ajuste

def generar_grafico_y_guardar(df, modelo, titulo, filename):
    fig = plt.figure(figsize=(15, 4))
    gs = gridspec.GridSpec(1, 3, width_ratios=[1, 1, 1])
    
    # Subgráfico 1: Scatterplot con línea OLS
    ax0 = plt.subplot(gs[0])
    sns.scatterplot(x=df["X"], y=df["Y"], color="blue", label="Datos", ax=ax0)
    X_sorted = df["X"].sort_values()
    Xc = sm.add_constant(X_sorted)
    y_pred = modelo.predict(Xc)
    ax0.plot(X_sorted, y_pred, color="red", label="OLS")
    ax0.set_title(titulo)
    ax0.set_xlabel("X")
    ax0.set_ylabel("Y")
    ax0.legend()
    ax0.grid(True)
    
    # Subgráfico 2: Q-Q plot de residuales
    ax1 = plt.subplot(gs[1])
    sm.qqplot(modelo.resid, line='45', ax=ax1)
    ax1.set_title("Q-Q Plot de Residuales")
    
    # Subgráfico 3: Gráfico de Escala
    ax2 = plt.subplot(gs[2])
    influence = modelo.get_influence()
    std_resid = influence.resid_studentized_internal
    sqrt_std_resid = np.sqrt(np.abs(std_resid))
    ax2.scatter(modelo.fittedvalues, sqrt_std_resid, color="blue")
    ax2.set_xlabel("Valores ajustados")
    ax2.set_ylabel("√|Residuales estandarizados|")
    ax2.set_title("Gráfico de Escala")
    ax2.grid(True)
    
    plt.tight_layout()
    plt.savefig(filename, dpi=100, bbox_inches='tight')
    plt.close()

def calcular_correlacion_ttest(df):
    m = len(df)
    if m < 3:
        return {
            "R": "Insuficientes datos",
            "R²": "N/A",
            "t_cal": "N/A",
            "t(0.05)": "N/A",
            "Conclusión": "Datos insuficientes para la correlación"
        }
    r_value = np.corrcoef(df["X"], df["Y"])[0, 1]
    R2_value = r_value**2
    if np.isclose(1 - R2_value, 0):
        t_cal = np.inf
    else:
        t_cal = r_value * np.sqrt((m - 2) / (1 - R2_value))
    t_critical = t.ppf(0.975, df=m-2)
    if abs(t_cal) > t_critical:
        conclusion = ("t_cal > t(0.05). Se rechaza H0, lo cual indica una correlación significativa (relación lineal).")
    else:
        conclusion = ("t_cal <= t(0.05). No se rechaza H0, no se evidencia correlación lineal significativa.")
    return {
        "R": f"{r_value:.4f}",
        "R²": f"{R2_value:.4f}",
        "t_cal": f"{t_cal:.4f}",
        "t(0.05)": f"{t_critical:.4f}",
        "Conclusión": conclusion
    }

def analizar_segmento(df, etiqueta_segmento, ws, row_excel):
    if df.empty:
        return row_excel

    modelo_original = realizar_regresion_OLS(df)
    original_coefs, original_adjust = extraer_stats_ols(modelo_original)
    dict_res_inicial = analizar_residuales(modelo_original)
    p_sw = float(dict_res_inicial["p-valor S-W"])
    resid = modelo_original.resid
    bp_test = het_breuschpagan(resid, modelo_original.model.exog)
    white_test = het_white(resid, modelo_original.model.exog)
    p_bp = bp_test[1]
    p_white = white_test[1]

    alt_analysis = False
    if (p_sw <= 0.05) or (p_bp <= 0.05) or (p_white <= 0.05):
        alt_analysis = True
        modelo_alt = realizar_regresion_OLS(df, usar_robusto=True, transformar=True)
        modelo = modelo_alt
        data_used = df.copy()
        data_used["Y_trans"] = np.log(df["Y"] + 1)
    else:
        modelo = modelo_original
        data_used = df

    dict_res = analizar_residuales(modelo)
    resid = modelo.resid
    sw_test = shapiro(resid)
    bp_test = het_breuschpagan(resid, modelo.model.exog)
    white_test = het_white(resid, modelo.model.exog)
    p_sw_new = sw_test.pvalue
    p_bp_new = bp_test[1]
    p_white_new = white_test[1]

    slope_p_value = modelo.pvalues[modelo.params.index[1]]
    if slope_p_value < 0.05:
        conclusion_lineal = "La pendiente es estadísticamente significativa (p < 0.05)."
    else:
        conclusion_lineal = "La pendiente no es estadísticamente significativa (p ≥ 0.05)."

    normalidad_eval = "CUMPLE" if p_sw_new > 0.05 else "NO CUMPLE"
    homoced1_eval = "CUMPLE" if p_bp_new > 0.05 else "NO CUMPLE"
    homoced2_eval = "CUMPLE" if p_white_new > 0.05 else "NO CUMPLE"

    fails = []
    if p_sw_new <= 0.05:
        fails.append("Normalidad")
    if p_bp_new <= 0.05:
        fails.append("Homocedasticidad (Prueba 1)")
    if p_white_new <= 0.05:
        fails.append("Homocedasticidad (Prueba robusta)")
    if len(fails) == 0:
        conclusion_supuestos = "CUMPLE: El modelo cumple con todos los supuestos requeridos."
    else:
        conclusion_supuestos = f"NO CUMPLE: El modelo no cumple con los siguientes supuestos: {', '.join(fails)}."
        if alt_analysis:
            conclusion_supuestos += " Se aplicó transformación logarítmica y errores robustos."

    if alt_analysis:
        transform_info = "Nota: Se aplicó transformación logarítmica a Y (log(Y+1)) y errores robustos."
    else:
        transform_info = "Nota: No se aplicó transformación logarítmica; se utilizó el modelo original."

    corr_dict = calcular_correlacion_ttest(df)
    png_name = f"temp_{etiqueta_segmento}.png"
    generar_grafico_y_guardar(data_used, modelo, f"OLS {etiqueta_segmento}", png_name)

    col_start = COLUMNA_INICIO_EXCEL
    header_row = row_excel
    escribir_encabezado_combinado(ws, header_row, col_start, col_start+4, f"Resultados OLS para {etiqueta_segmento}")
    insertar_imagen_excel(ws, png_name, f"J{header_row}")
    row_excel += 2

    row_excel = escribir_tabla_dict(ws, row_excel, col_start, original_coefs, "Coeficientes del Modelo (Original)")
    row_excel = escribir_tabla_dict(ws, row_excel, col_start, original_adjust, "Bondad de Ajuste (Original)")
    row_excel = escribir_tabla_dict(ws, row_excel, col_start, dict_res, "Análisis de Residuales")
    row_excel = escribir_tabla_dict(ws, row_excel, col_start, corr_dict, "Análisis de Correlación")
    texto_conclusion = (
        "Conclusiones:\n"
        f"{conclusion_lineal}\n\n"
        f"Supuestos del análisis: Normalidad de los residuales\n"
        f"Valor p (Shapiro-Wilk): {p_sw_new:.4f} --> {normalidad_eval}\n"
        "Revisar el gráfico Q-Q de los residuales.\n\n"
        f"Supuestos del análisis: Homocedasticidad (Prueba 1)\n"
        f"Valor p (Breusch-Pagan): {p_bp_new:.4f} --> {homoced1_eval}\n"
        "Revisar el gráfico de residuales y el gráfico de escala.\n\n"
        f"Supuestos del análisis: Homocedasticidad (Prueba robusta)\n"
        f"Valor p (Studentizada Breusch-Pagan): {p_white_new:.4f} --> {homoced2_eval}\n"
        "Revisar el gráfico de residuales y el gráfico de escala.\n\n"
        f"Resumen: {conclusion_supuestos}\n"
        f"{transform_info}\n\n"
        "H0: La correlación poblacional es cero (no hay relación lineal).\n"
        "Ha: La correlación poblacional es distinta de cero (sí hay relación lineal).\n"
        "Si |t_cal| > t(0.05), se rechaza H0 y se confirma la correlación lineal.\n\n"
        "Se recomienda revisar los gráficos (incluyendo el Q-Q de residuales) para confirmar visualmente."
    )
    # Claves para formatear títulos en negrita y con estilo destacado.
    bold_keys = ["Conclusiones", "Supuestos", "Resumen", "H0", "Ha"]
    # Definimos estilos adicionales para títulos y cuerpo de texto.
    title_font = Font(name="Calibri", size=11, bold=True, color="000080")
    body_font = Font(name="Calibri", size=11, bold=False, color="000000")
    title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Escribimos cada línea en celdas combinadas de la columna E a M, aplicando estilos según corresponda.
    for line in texto_conclusion.splitlines():
        if line.strip() == "":
            row_excel += 1
            continue
        ws.merge_cells(start_row=row_excel, start_column=5, end_row=row_excel, end_column=13)
        celda = ws.cell(row=row_excel, column=5, value=line)
        # Aplica estilo de título si la línea inicia con alguna clave, de lo contrario aplica estilo normal.
        if any(line.strip().startswith(key) for key in bold_keys):
            celda.font = title_font
            celda.fill = title_fill
        else:
            celda.font = body_font
            celda.fill = normal_fill
        celda.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
        row_excel += 1
    row_excel += 3
    return row_excel

def analizar_ols_por_segmentos(ruta_excel):
    try:
        df = cargar_datos_ols(ruta_excel)
    except Exception as e:
        print(f"No se pudo leer {ruta_excel}: {e}")
        return

    wb = load_workbook(ruta_excel)
    ws_linealidad = wb["Linealidad - no parametrico"]
    # Combinar siempre las celdas A2:M2 para el título
    ws_linealidad.merge_cells("A2:M2")
    cell = ws_linealidad["A2"]
    cell.value = "Estudio de linealidad"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True, color="FFFFFF", size=16)
    cell.fill = PatternFill(fill_type="solid", fgColor="00008B")
    ws = wb[NOMBRE_HOJA]
    limpiar_hoja(ws)
    row_excel = FILA_INICIO_EXCEL

    # Análisis para el segmento de valores NEGATIVOS
    df_neg = df[df["X"] < 0].copy()
    if not df_neg.empty:
        df_neg = df_neg.groupby("X", as_index=False).mean()
        row_excel = analizar_segmento(df_neg, "Negativos", ws, row_excel)
    else:
        print(f"{os.path.basename(ruta_excel)}: No hay datos negativos.")

    # Análisis para el segmento de valores POSITIVOS
    df_pos = df[df["X"] >= 0].copy()
    if not df_pos.empty:
        df_pos = df_pos.groupby("X", as_index=False).mean()
        row_excel = analizar_segmento(df_pos, "Positivos", ws, row_excel)
    else:
        print(f"{os.path.basename(ruta_excel)}: No hay datos positivos.")

    wb.save(ruta_excel)
    for etiqueta in ["Negativos", "Positivos"]:
        temp_file = f"temp_{etiqueta}.png"
        if os.path.exists(temp_file):
            os.remove(temp_file)

def seleccionar_carpeta_ols():
    root = Tk()
    root.withdraw()
    carpeta = filedialog.askdirectory(title="Seleccione la carpeta para el análisis de Linealidad - primera parte")
    root.destroy()
    return carpeta

# =============================================================================
# Funciones para el Test de Significancia - Linealidad (App 2)
# =============================================================================
def format_p_value(p):
    if not isinstance(p, (float, np.floating)):
        return p
    if p < 1e-16:
        return "<1e-16"
    else:
        return f"{p:.4E}"

def lack_of_fit_test(df, xcol, ycol):
    df_ = df.copy()
    formula = f"{ycol} ~ {xcol}"
    model_ = smf.ols(formula, data=df_).fit()
    df_["_pred_"] = model_.predict(df_[[xcol]])
    
    groups = df_.groupby(xcol)
    SSR_LOF = 0.0
    SSR_PE  = 0.0
    n_levels = groups.ngroups
    N_total  = len(df_)
    
    for _, g in groups:
        y_mean = g[ycol].mean()
        y_pred = g["_pred_"].mean()
        n_i    = len(g)
        SSR_LOF += n_i * (y_mean - y_pred)**2
        SSR_PE  += ((g[ycol] - y_mean)**2).sum()
    
    df_lof = n_levels - 2   # Dos parámetros estimados: intercepto y pendiente
    df_pe  = N_total - n_levels

    if df_lof <= 0 or df_pe <= 0:
        return np.nan
    
    MS_LOF = SSR_LOF / df_lof
    if SSR_PE == 0:
        return "Ajuste perfecto"
    MS_PE  = SSR_PE / df_pe
    F_lof  = MS_LOF / MS_PE if MS_PE != 0 else np.inf
    p_lof  = stats.f.sf(F_lof, df_lof, df_pe)
    
    return p_lof

def process_file(file_path):
    try:
        df = pd.read_excel(file_path,
                           sheet_name="Linealidad - no parametrico",
                           skiprows=buscar_fila_encabezado(file_path, NOMBRE_HOJA),
                           header=0,
                           usecols="B:C")
    except Exception as e:
        print(f"Error al leer la hoja en {file_path}: {e}")
        return False, None

    if "LP" not in df.columns or "LI" not in df.columns:
        print(f"El archivo {file_path} no contiene las columnas 'LP' y 'LI'.")
        return False, None

    df.dropna(subset=["LP", "LI"], inplace=True)
    df["LP"] = pd.to_numeric(df["LP"], errors="coerce")
    df["LI"] = pd.to_numeric(df["LI"], errors="coerce")
    df.dropna(subset=["LP", "LI"], inplace=True)

    if df["LI"].nunique() == 1:
        print(f"En el archivo {file_path}, la variable LI es constante. Se añade ruido mínimo.")
        df["LI"] = df["LI"] + np.random.normal(0, 1e-6, size=len(df["LI"]))

    n_registros_antes = df.shape[0]
    for col in ["LP", "LI"]:
        Q1 = df[col].quantile(0.25)
        Q3 = df[col].quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - 1.5 * IQR
        upper_bound = Q3 + 1.5 * IQR
        df = df[(df[col] >= lower_bound) & (df[col] <= upper_bound)]
    n_registros_despues = df.shape[0]
    registros_eliminados = n_registros_antes - n_registros_despues

    X_data = df["LP"].values
    Y_data = df["LI"].values
    alpha = 0.05
    m = df["LP"].nunique()
    R = np.corrcoef(X_data, Y_data)[0, 1]
    t_cal_R = abs(R) * np.sqrt((m - 2) / (1 - R**2)) if (1 - R**2) > 0 else np.inf
    df_t = m - 2
    p_value_R = stats.t.sf(t_cal_R, df_t) * 2 if (1 - R**2) > 0 else np.nan
    t_tabla_R = stats.t.ppf(1 - alpha/2, df_t) if df_t > 0 else np.nan
    signif_R = t_cal_R > t_tabla_R if not np.isnan(t_cal_R) else False

    model = smf.ols("LI ~ LP", data=df).fit()
    anova_table = sm.stats.anova_lm(model, typ=2)
    p_anova_reg = anova_table.loc["LP", "PR(>F)"] if "LP" in anova_table.index else np.nan
    signif_regresion = (p_anova_reg < alpha) if not np.isnan(p_anova_reg) else False

    p_lof = lack_of_fit_test(df, "LP", "LI")
    falta_ajuste = (not isinstance(p_lof, str) and not np.isnan(p_lof) and p_lof < alpha)

    b = model.params.get("LP", np.nan)
    a = model.params.get("Intercept", np.nan)
    sb = model.bse.get("LP", np.nan)
    sa = model.bse.get("Intercept", np.nan)
    n = len(df)
    gl_t = n - 2
    t_crit = stats.t.ppf(1 - alpha/2, gl_t) if gl_t > 0 else np.nan
    t_b = abs(b) / sb if sb != 0 else np.inf
    t_a = abs(a) / sa if sa != 0 else np.inf
    signif_b = (t_b > t_crit) if not np.isnan(t_crit) else False
    signif_a = (t_a > t_crit) if not np.isnan(t_crit) else False

    plt.figure(figsize=(8, 6))
    plt.scatter(X_data, Y_data, color="blue", label="Datos")
    x_vals = np.linspace(min(X_data), max(X_data), 100)
    y_vals = a + b * x_vals
    plt.plot(x_vals, y_vals, color="red", label="Línea de regresión")
    plt.xlabel("LP")
    plt.ylabel("LI")
    plt.title("Gráfico de dispersión y regresión")
    plt.legend()
    plt.grid(True)
    plot_file = "regression_plot.png"
    plt.savefig(plot_file)
    plt.close()

    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"Error al abrir {file_path} para guardar resultados: {e}")
        return False, None

    sheet_name_results = "Test Significancia - Linealidad"
    if sheet_name_results in wb.sheetnames:
        ws = wb[sheet_name_results]
        try:
            wb.remove(ws)
        except Exception:
            pass
        wb._sheets.insert(4, ws)
    else:
        ws = wb.create_sheet(sheet_name_results, index=4)
    try:
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 20
    except Exception as ex:
        print("Error ajustando dimensiones de columnas:", ex)

    ws.merge_cells("A1:D1")
    ws["A1"] = "Test de Significancia - Linealidad"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal='center', vertical='center')

    ws["A3"] = "Hipótesis para la Relación"
    ws["A3"].font = Font(bold=True, size=14)
    ws["A4"] = "Hipótesis Nula (H0):"
    ws["B4"] = ("No existe una relación lineal significativa entre 'LP' y 'LI'; la asociación observada se debe al azar.")
    ws["A5"] = "Hipótesis Alternativa (H1):"
    ws["B5"] = ("Existe una relación lineal significativa entre 'LP' y 'LI', lo que implica que al menos uno de los parámetros difiere de cero.")

    ws["A7"] = "Hipótesis para la Falta de Ajuste"
    ws["A7"].font = Font(bold=True, size=14)
    ws["A8"] = "Hipótesis Nula (H0):"
    ws["B8"] = ("El modelo lineal es adecuado para representar la relación entre 'LP' y 'LI'.")
    ws["A9"] = "Hipótesis Alternativa (H1):"
    ws["B9"] = ("El modelo lineal no es adecuado (existe falta de ajuste), lo que sugiere que la forma funcional es inadecuada.")

    ws["A11"] = "Análisis de Valores Atípicos"
    ws["A11"].font = Font(bold=True, size=14)
    ws["A12"] = "Registros originales (tras limpieza):"
    ws["B12"] = n_registros_antes
    ws["A13"] = "Registros tras eliminación de outliers:"
    ws["B13"] = n_registros_despues
    ws["A14"] = "Registros eliminados:"
    ws["B14"] = registros_eliminados
    ws["A15"] = "Transformación logarítmica aplicada:"
    ws["B15"] = "No"

    ws["A17"] = "Evaluaciones de Significancia"
    ws["A17"].font = Font(bold=True, size=14)
    ws["A18"] = "Coeficiente de correlación (R):"
    ws["B18"] = round(R, 4)
    ws["A19"] = "p-value correlación:"
    ws["B19"] = format_p_value(p_value_R) if not np.isnan(p_value_R) else "N/A"
    ws["A20"] = "Significancia de la correlación:"
    ws["B20"] = "Significativa" if signif_R else "No significativa"
    ws["A21"] = "p-value ANOVA:"
    ws["B21"] = format_p_value(p_anova_reg) if not np.isnan(p_anova_reg) else "N/A"
    ws["A22"] = "Significancia ANOVA:"
    ws["B22"] = "Significativo" if signif_regresion else "No significativo"
    ws["A23"] = "Pendiente (b):"
    ws["B23"] = round(b, 4)
    ws["A24"] = "Significancia de la pendiente:"
    ws["B24"] = "Significativa" if signif_b else "No significativa"
    ws["A25"] = "Intercepto (a):"
    ws["B25"] = round(a, 4)
    ws["A26"] = "Significancia del intercepto:"
    ws["B26"] = "Significativo" if signif_a else "No significativo"
    ws["A27"] = "t para pendiente:"
    ws["B27"] = round(t_b, 4)
    ws["A28"] = "t para intercepto:"
    ws["B28"] = round(t_a, 4)
    ws["A29"] = "t crítico:"
    ws["B29"] = round(t_crit, 4)

    ws["A31"] = "Evaluación de la Falta de Ajuste"
    ws["A31"].font = Font(bold=True, size=14)
    ws["A32"] = "p-value Falta de Ajuste:"
    ws["B32"] = format_p_value(p_lof) if not isinstance(p_lof, str) and not np.isnan(p_lof) else p_lof if isinstance(p_lof, str) else "N/A"
    ws["A33"] = "Evaluación de la Falta de Ajuste:"
    if isinstance(p_lof, str) and p_lof == "Ajuste perfecto":
        ws["B33"] = "El modelo se ajusta perfectamente a los datos."
    else:
        ws["B33"] = "Presencia de falta de ajuste" if falta_ajuste else "El modelo lineal es adecuado."

    ws["A35"] = "Conclusiones Globales"
    ws["A35"].font = Font(bold=True, size=16)
    conclusion = ""
    if signif_regresion:
        conclusion += ("Los resultados indican que existe una relación estadísticamente significativa entre 'LP' y 'LI' "
                       "(p-value ANOVA: " + format_p_value(p_anova_reg) + ").\n")
    else:
        conclusion += ("Los resultados no muestran evidencia suficiente de una relación significativa entre 'LP' y 'LI' "
                       "(p-value ANOVA: " + format_p_value(p_anova_reg) + ").\n")
    
    if isinstance(p_lof, str) and p_lof == "Ajuste perfecto":
        conclusion += "El modelo presenta un ajuste perfecto a los datos, lo que en algunos casos puede dificultar la interpretación de la prueba de falta de ajuste.\n"
    elif falta_ajuste:
        conclusion += ("Sin embargo, la prueba de falta de ajuste (p-value: " + format_p_value(p_lof) + ") indica que el modelo lineal "
                       "no captura completamente la variabilidad, sugiriendo que la forma funcional podría ser inadecuada.\n")
    else:
        conclusion += ("La prueba de falta de ajuste (p-value: " + format_p_value(p_lof) + ") respalda la adecuación del modelo lineal.\n")
    
    conclusion += ("Se recomienda complementar el análisis con un estudio de residuos y, de ser necesario, explorar modelos alternativos o transformaciones "
                   "para lograr un ajuste óptimo de los datos.\n")
    
    ws["A36"] = conclusion
    ws.merge_cells("A36:D39")
    ws["A36"].alignment = Alignment(vertical="center", horizontal="left", wrapText=True)

    try:
        img = Img(plot_file)
        ws.add_image(img, "E38")
    except Exception as e:
        print(f"Error al insertar el gráfico original en {file_path}: {e}")

    transformation_applied = False
    if falta_ajuste:
        if (df["LP"] > 0).all() and (df["LI"] > 0).all():
            print(f"En {file_path} se detectó falta de ajuste. Aplicando transformación logarítmica para reanálisis.")
            df["LP_adj"] = np.log(df["LP"])
            df["LI_adj"] = np.log(df["LI"])
            transformation_applied = True

            X_data_adj = df["LP_adj"].values
            Y_data_adj = df["LI_adj"].values
            R_adj = np.corrcoef(X_data_adj, Y_data_adj)[0, 1]
            if (1 - R_adj**2) > 0:
                t_cal_R_adj = abs(R_adj) * np.sqrt((m - 2) / (1 - R_adj**2))
                df_t_adj = m - 2
                p_value_R_adj = stats.t.sf(t_cal_R_adj, df_t_adj) * 2
                t_tabla_R_adj = stats.t.ppf(1 - alpha/2, df_t_adj)
                signif_R_adj = t_cal_R_adj > t_tabla_R_adj
            else:
                t_cal_R_adj, t_tabla_R_adj, p_value_R_adj = np.nan, np.nan, np.nan
                signif_R_adj = False

            model_adj = smf.ols("LI_adj ~ LP_adj", data=df).fit()
            anova_table_adj = sm.stats.anova_lm(model_adj, typ=2)
            p_anova_reg_adj = anova_table_adj.loc["LP_adj", "PR(>F)"] if "LP_adj" in anova_table_adj.index else np.nan
            signif_regresion_adj = (p_anova_reg_adj < alpha) if not np.isnan(p_anova_reg_adj) else False

            b_adj = model_adj.params.get("LP_adj", np.nan)
            a_adj = model_adj.params.get("Intercept", np.nan)
            sb_adj = model_adj.bse.get("LP_adj", np.nan)
            sa_adj = model_adj.bse.get("Intercept", np.nan)
            t_crit_adj = stats.t.ppf(1 - alpha/2, len(df)-2) if (len(df)-2) > 0 else np.nan
            t_b_adj = abs(b_adj) / sb_adj if sb_adj != 0 else np.inf
            t_a_adj = abs(a_adj) / sa_adj if sa_adj != 0 else np.inf
            signif_b_adj = (t_b_adj > t_crit_adj) if not np.isnan(t_crit_adj) else False
            signif_a_adj = (t_a_adj > t_crit_adj) if not np.isnan(t_crit_adj) else False

            plt.figure(figsize=(8, 6))
            plt.scatter(X_data_adj, Y_data_adj, color="green", label="Datos ajustados")
            x_vals_adj = np.linspace(min(X_data_adj), max(X_data_adj), 100)
            y_vals_adj = a_adj + b_adj * x_vals_adj
            plt.plot(x_vals_adj, y_vals_adj, color="purple", label="Línea de regresión ajustada")
            plt.xlabel("log(LP)")
            plt.ylabel("log(LI)")
            plt.title("Gráfico de dispersión y regresión (Datos Ajustados)")
            plt.legend()
            plt.grid(True)
            plot_file_adj = "regression_plot_adj.png"
            plt.savefig(plot_file_adj)
            plt.close()

            start_row = 40
            ws[f"A{start_row}"] = "Reanálisis con Transformación Logarítmica"
            ws[f"A{start_row}"].font = Font(bold=True, size=16)
            ws[f"A{start_row+1}"] = "Coeficiente de correlación (R):"
            ws[f"B{start_row+1}"] = round(R_adj, 4)
            ws[f"A{start_row+2}"] = "p-value correlación:"
            ws[f"B{start_row+2}"] = format_p_value(p_value_R_adj) if not np.isnan(p_value_R_adj) else "N/A"
            ws[f"A{start_row+3}"] = "Significancia correlación:"
            ws[f"B{start_row+3}"] = "Significativo" if signif_R_adj else "No significativo"
            ws[f"A{start_row+4}"] = "p-value ANOVA:"
            ws[f"B{start_row+4}"] = format_p_value(p_anova_reg_adj) if not np.isnan(p_anova_reg_adj) else "N/A"
            ws[f"A{start_row+5}"] = "Significancia ANOVA:"
            ws[f"B{start_row+5}"] = "Significativo" if signif_regresion_adj else "No significativo"
            ws[f"A{start_row+6}"] = "Pendiente (b):"
            ws[f"B{start_row+6}"] = round(b_adj, 4)
            ws[f"A{start_row+7}"] = "Significancia pendiente:"
            ws[f"B{start_row+7}"] = "Significativa" if signif_b_adj else "No significativa"
            ws[f"A{start_row+8}"] = "Intercepto (a):"
            ws[f"B{start_row+8}"] = round(a_adj, 4)
            ws[f"A{start_row+9}"] = "Significancia intercepto:"
            ws[f"B{start_row+9}"] = "Significativo" if signif_a_adj else "No significativo"
            ws[f"A{start_row+10}"] = "t para pendiente:"
            ws[f"B{start_row+10}"] = round(t_b_adj, 4)
            ws[f"A{start_row+11}"] = "t para intercepto:"
            ws[f"B{start_row+11}"] = round(t_a_adj, 4)
            ws[f"A{start_row+12}"] = "t crítico:"
            ws[f"B{start_row+12}"] = round(t_crit_adj, 4)
            
            ws[f"A{start_row+13}"] = "Conclusiones (Transformación Logarítmica):"
            ws[f"A{start_row+13}"].font = Font(bold=True, size=14)
            if signif_regresion_adj:
                ws[f"A{start_row+14}"] = (
                    "El reanálisis con transformación logarítmica confirma la existencia de una relación significativa entre log(LP) y log(LI). "
                    "La transformación permitió mejorar el ajuste del modelo, evidenciado por una prueba de falta de ajuste menos problemática."
                )
            else:
                ws[f"A{start_row+14}"] = (
                    "El reanálisis con transformación logarítmica no confirma de forma robusta una relación significativa, "
                    "lo que sugiere que se deben explorar otros modelos o transformaciones."
                )
            ws.merge_cells("A54:C57")
            ws["A54"].alignment = Alignment(vertical="center", horizontal="left", wrapText=True)
            
            try:
                img_adj = Img(plot_file_adj)
                ws.add_image(img_adj, "D58")
            except Exception as e:
                print(f"Error al insertar el gráfico de reanálisis en {file_path}: {e}")
            
            table_start_row = start_row + 18
            ws[f"A{table_start_row}"] = "Tabla de Datos Ajustados"
            ws[f"A{table_start_row}"].font = Font(bold=True, size=14)
            ws[f"A{table_start_row+1}"] = "Índice"
            ws[f"B{table_start_row+1}"] = "LP ajustado"
            ws[f"C{table_start_row+1}"] = "LI ajustado"
            
            current_row = table_start_row + 2
            for idx, row in df.iterrows():
                ws[f"A{current_row}"] = idx
                ws[f"B{current_row}"] = row["LP_adj"] if "LP_adj" in row else ""
                ws[f"C{current_row}"] = row["LI_adj"] if "LI_adj" in row else ""
                current_row += 1
        else:
            print(f"No se puede aplicar la transformación logarítmica en {file_path} por valores no positivos.")

    try:
        wb.save(file_path)
    except Exception as e:
        print(f"Error al guardar {file_path}: {e}")
        return False, None

    return True, "Proceso completado"

# =============================================================================
# Función de prueba para listar los datos usados para el análisis
# =============================================================================
def listar_datos_prueba(ruta_archivo):
    """
    Carga y muestra en consola los datos de la hoja "Linealidad - no parametrico" (columnas B y C)
    utilizando la detección automática de la fila que contiene las etiquetas.
    """
    try:
        df = cargar_datos_ols(ruta_archivo)
        print("Listado de datos para análisis:")
        print(df.to_string(index=False))
        return df
    except Exception as e:
        print(f"Error al cargar datos de prueba: {e}")
        return None

def prueba_listado_datos():
    """
    Solicita al usuario que seleccione el archivo Excel de prueba y lista los datos en consola.
    """
    root = Tk()
    root.withdraw()
    ruta_prueba = filedialog.askopenfilename(
        title="Seleccione el archivo Excel de prueba",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )
    root.destroy()
    if not ruta_prueba:
        print("No se seleccionó ningún archivo.")
        return
    print("Ejecutando la prueba de listado de datos...")
    df = listar_datos_prueba(ruta_prueba)
    if df is not None:
        print("La prueba de listado de datos se ejecutó correctamente.")
    else:
        print("La prueba de listado de datos NO se pudo ejecutar.")

# =============================================================================
# Función principal combinada: Ejecuta ambas aplicaciones secuencialmente
# =============================================================================
def main():
    # Ejecutar la prueba para listar los datos usados en el análisis
    prueba_listado_datos()
    
    root = Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title="Seleccione la carpeta con los archivos Excel para ambos análisis")
    root.destroy()
    if not folder_path:
        print("No se seleccionó ninguna carpeta.")
        return
    archivos = [os.path.join(folder_path, f) for f in os.listdir(folder_path)
                if os.path.isfile(os.path.join(folder_path, f)) and f.lower().endswith('.xlsx')]
    if not archivos:
        print("No se encontraron archivos Excel en la carpeta seleccionada.")
        return

    print("Iniciando Análisis de Linealidad (App 1)...")
    for file_path in archivos:
        print(f"Procesando archivo para Análisis OLS: {os.path.basename(file_path)}")
        analizar_ols_por_segmentos(file_path)
    print("Análisis de Linealidad completado.\n")

    print("Iniciando Test de Significancia - Linealidad (App 2)...")
    procesados = 0
    errores = 0
    for file_path in archivos:
        print(f"Procesando archivo para Test de Significancia: {os.path.basename(file_path)}")
        result, _ = process_file(file_path)
        if result:
            print(f"{os.path.basename(file_path)} procesado correctamente en Test de Significancia.")
            procesados += 1
        else:
            print(f"Error al procesar {os.path.basename(file_path)} en Test de Significancia.")
            errores += 1
    print("\nResumen del Test de Significancia:")
    print(f"Archivos procesados correctamente: {procesados}")
    print(f"Archivos con error: {errores}")
    print("\nProceso completo. Revise los archivos Excel para ver los resultados.")

if __name__ == "__main__":
    main()
