import os
import re
import tkinter as tk
from tkinter import filedialog
import pandas as pd
import ollama
import concurrent.futures
from rich.console import Console
from rich.panel import Panel
from openpyxl import load_workbook
from openpyxl.styles import Alignment

console = Console()

# Función para procesar cada archivo
def process_file(file_tuple):
    numeric_prefix, file_path = file_tuple
    file_name = os.path.basename(file_path)
    try:
        # 1. Extraer datos de la hoja "Linealidad - Parametrico" (celdas E7:M94)
        df = pd.read_excel(file_path, sheet_name="Linealidad - Parametrico", header=None)
        data_range = df.iloc[6:94, 4:13]  # Filas 7 a 94 y columnas E a M (índices 6 y 4)
        data_str = data_range.to_csv(index=False)
        
        # 2. Definir las recomendaciones para la redacción
        recomendaciones = (
            "Recomendaciones para mejorar la redacción: "
            "1. Usa un lenguaje formal y coherente. "
            "2. Sé claro y conciso, sin repeticiones. "
            "3. Organiza la información en párrafos breves y bien estructurados. "
            "4. Revisa la gramática y ortografía. "
            "5. Utiliza términos técnicos adecuados y explica los conceptos si es necesario. "
            "6. Mantén un flujo lógico y evita jergas. "
            "7. Sé objetivo y preciso, respaldando las afirmaciones con datos. "
            "8. Concluye de forma clara resumiendo los puntos principales."
        )
        
        # 3. Primer análisis: Conclusión final
        prompt_conclusion = (
            "Analiza los siguientes datos extraídos de un archivo Excel (hoja 'Linealidad - Parametrico', celdas E7:M94) y "
            "proporciona únicamente la conclusión final sobre los resultados, sin detallar el proceso interno de análisis. "
            "La respuesta debe estar redactada de forma profesional, en un solo párrafo y de forma directa. "
            "No analices los gráficos, solo los datos.\n\n"
            "Analiza todos los supuestos y condiciones de la prueba y proporciona una conclusión final sobre los resultados.\n\n"
            f"{recomendaciones}\n\n"
            "Datos:\n"
            f"{data_str}\n\n"
            "Conclusión:"
        )
        messages_conclusion = [{'role': 'user', 'content': prompt_conclusion}]
        response_conclusion = ollama.chat(model='llama3.2:1b', messages=messages_conclusion)
        if isinstance(response_conclusion, dict):
            conclusion_text = response_conclusion.get("message", {}).get("content", "")
        else:
            conclusion_text = response_conclusion
        conclusion_text = re.sub(r'</?think>', '', conclusion_text).strip()
        
        # 4. Segundo análisis: Resumen general
        prompt_resumen = (
            "Con base en los mismos datos extraídos del archivo Excel (hoja 'Linealidad - Parametrico', celdas E7:M94), "
            "proporciona un resumen general de los resultados. La respuesta debe estar redactada de forma profesional, "
            "en un solo párrafo y de forma directa, sin detallar el proceso interno ni análisis de gráficos.\n\n"
            "Datos:\n"
            f"{data_str}\n\n"
            "Resumen:"
        )
        messages_resumen = [{'role': 'user', 'content': prompt_resumen}]
        response_resumen = ollama.chat(model='llama3.2:3b', messages=messages_resumen)
        if isinstance(response_resumen, dict):
            resumen_text = response_resumen.get("message", {}).get("content", "")
        else:
            resumen_text = response_resumen
        resumen_text = re.sub(r'</?think>', '', resumen_text).strip()
        
        # 5. Escribir ambos resultados en la hoja "Plan de validación"
        wb = load_workbook(file_path)
        if "Plan de validación" in wb.sheetnames:
            ws = wb["Plan de validación"]
        else:
            ws = wb.create_sheet("Plan de validación")
        
        # Escribir la conclusión final en la celda superior izquierda del rango combinado B60:AF67
        ws["B60"] = conclusion_text
        ws["B60"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Escribir el resumen general en la celda superior izquierda del rango combinado B69:AF75
        ws["B70"] = resumen_text
        ws["B70"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        wb.save(file_path)
        return numeric_prefix, file_name, conclusion_text, resumen_text
    except Exception as e:
        return numeric_prefix, file_name, f"Error: {e}", f"Error: {e}"

# Inicializar Tkinter y ocultar la ventana principal
root = tk.Tk()
root.withdraw()

# Seleccionar la carpeta con archivos Excel
folder_path = filedialog.askdirectory(title="Selecciona la carpeta con los archivos Excel")
if not folder_path:
    console.print("[red]No se seleccionó ninguna carpeta.[/red]")
    exit()

# Filtrar y ordenar archivos Excel que comienzan con un sufijo numérico
excel_files = []
for file in os.listdir(folder_path):
    if file.lower().endswith('.xlsx'):
        m = re.match(r'^(\d+)', file)
        if m:
            numeric_prefix = int(m.group(1))
            excel_files.append((numeric_prefix, os.path.join(folder_path, file)))
excel_files.sort(key=lambda x: x[0])
if not excel_files:
    console.print("[red]No se encontraron archivos Excel con sufijo numérico al inicio del nombre.[/red]")
    exit()

# Procesar archivos concurrentemente con un spinner para animar la espera
results = []
with console.status("[bold green]Procesando archivos...", spinner="dots") as status:
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_file, file_tuple) for file_tuple in excel_files]
        for future in concurrent.futures.as_completed(futures):
            results.append(future.result())

# Ordenar resultados según el sufijo numérico
results.sort(key=lambda x: x[0])

# Mostrar los resultados embellecidos en la consola (se muestran ambos análisis)
for numeric_prefix, file_name, conclusion_text, resumen_text in results:
    contenido = f"[bold]Conclusión:[/bold]\n{conclusion_text}\n\n[bold]Resumen:[/bold]\n{resumen_text}"
    panel = Panel(contenido, title=f"[bold blue]{file_name}[/bold blue]", expand=False)
    console.print(panel)
