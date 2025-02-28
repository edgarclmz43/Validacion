#!/usr/bin/env python3
import os
import re
import tkinter as tk
from tkinter import filedialog
import pandas as pd
import ollama
import concurrent.futures
from rich.console import Console
from rich.panel import Panel
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

console = Console()

def process_file(file_tuple):
    numeric_prefix, file_path = file_tuple
    file_name = os.path.basename(file_path)
    try:
        # Recomendaciones comunes para la redacción
        recomendaciones = (
            "Recomendaciones para mejorar la redacción: "
            "1. Usa un lenguaje formal y coherente. "
            "2. Sé claro y conciso, sin repeticiones. "
            "3. Organiza la información en párrafos breves y bien estructurados. "
            "4. Revisa la gramática y ortografía. "
            "5. Utiliza términos técnicos adecuados y explica los conceptos si es necesario. "
            "6. Mantén un flujo lógico y evita jergas. "
            "7. Sé objetivo y preciso, respaldando las afirmaciones con datos. "
            "8. Concluye de forma clara resumiendo los puntos principales."
        )

        # ==================== Análisis 1: Linealidad ====================
        # Se extrae la información desde la celda E7 (rango E7:M94) de la hoja "Linealidad - Parametrico"
        df_linealidad = pd.read_excel(file_path, sheet_name="Linealidad - Parametrico", header=None)
        data_range_linealidad = df_linealidad.iloc[6:94, 4:13]  # filas 7 a 94, columnas E a M
        data_str_linealidad = data_range_linealidad.to_csv(index=False)
        prompt_linealidad = (
            "Analiza los siguientes datos extraídos del archivo Excel (hoja 'Linealidad - Parametrico', desde la celda E7) "
            "y proporciona un mensaje final sobre los resultados. La respuesta debe estar redactada de forma profesional, "
            "en un solo párrafo y de forma directa, sin detallar el proceso interno ni analizar gráficos.\n\n"
            f"{recomendaciones}\n\n"
            "Datos:\n" + data_str_linealidad + "\n\n"
            "Mensaje:"
        )
        response_linealidad = ollama.chat(model='llama3.2:3b', messages=[{'role': 'user', 'content': prompt_linealidad}])
        if isinstance(response_linealidad, dict):
            mensaje_linealidad = response_linealidad.get("message", {}).get("content", "")
        elif hasattr(response_linealidad, "message"):
            mensaje_linealidad = response_linealidad.message.content
        elif hasattr(response_linealidad, "content"):
            mensaje_linealidad = response_linealidad.content
        else:
            mensaje_linealidad = str(response_linealidad)
        mensaje_linealidad = re.sub(r'</?think>', '', mensaje_linealidad).strip()

        # ==================== Análisis 2: Significante ====================
        # Se extrae la información desde la celda A1 de la hoja "Test Significancia - Linealidad"
        df_significante = pd.read_excel(file_path, sheet_name="Test Significancia - Linealidad", header=None)
        data_range_significante= df_linealidad.iloc[0:57, 0:3]  # filas 0 a 205 columnas A a D
        #data_range_significante = df_significante  # se extrae toda la info desde A1
        data_str_significante = data_range_significante.to_csv(index=False)
        prompt_significante = (
            "Analiza los siguientes datos extraídos del archivo Excel (hoja 'Test Significancia - Linealidad', desde la celda A1) "
            "y proporciona un mensaje final sobre los resultados. La respuesta debe estar redactada de forma profesional, "
            "en un solo párrafo y de forma directa, sin detallar el proceso interno ni analizar gráficos.\n\n"
            f"{recomendaciones}\n\n"
            "Datos:\n" + data_str_significante + "\n\n"
            "Mensaje:"
        )
        response_significante = ollama.chat(model='llama3.2:3b', messages=[{'role': 'user', 'content': prompt_significante}])
        if isinstance(response_significante, dict):
            mensaje_significante = response_significante.get("message", {}).get("content", "")
        elif hasattr(response_significante, "message"):
            mensaje_significante = response_significante.message.content
        elif hasattr(response_significante, "content"):
            mensaje_significante = response_significante.content
        else:
            mensaje_significante = str(response_significante)
        mensaje_significante = re.sub(r'</?think>', '', mensaje_significante).strip()

        # ==================== Análisis 3: Repetibilidad ====================
        # Se extrae la información desde la celda B114 a L172 de la hoja "sr ysR(Método) ISO 5725"
        df_repetibilidad = pd.read_excel(file_path, sheet_name="sr ysR(Método) ISO 5725", header=None)
        data_range_repetibilidad = df_repetibilidad.iloc[113:172, 1:12]  # filas 114 a 172, columnas B a L
        data_str_repetibilidad = data_range_repetibilidad.to_csv(index=False)
        prompt_repetibilidad = (
            "Analiza los siguientes datos extraídos del archivo Excel (hoja 'sr ysR(Método) ISO 5725', desde la celda B114 a L172) "
            "y proporciona un mensaje final sobre los resultados. La respuesta debe estar redactada de forma profesional, "
            "en un solo párrafo y de forma directa, sin detallar el proceso interno ni analizar gráficos.\n\n"
            f"{recomendaciones}\n\n"
            "Datos:\n" + data_str_repetibilidad + "\n\n"
            "Mensaje:"
        )
        response_repetibilidad = ollama.chat(model='llama3.2:3b', messages=[{'role': 'user', 'content': prompt_repetibilidad}])
        if isinstance(response_repetibilidad, dict):
            mensaje_repetibilidad = response_repetibilidad.get("message", {}).get("content", "")
        elif hasattr(response_repetibilidad, "message"):
            mensaje_repetibilidad = response_repetibilidad.message.content
        elif hasattr(response_repetibilidad, "content"):
            mensaje_repetibilidad = response_repetibilidad.content
        else:
            mensaje_repetibilidad = str(response_repetibilidad)
        mensaje_repetibilidad = re.sub(r'</?think>', '', mensaje_repetibilidad).strip()

        # ==================== Análisis 4: Precisión ====================
        # Se extrae la información desde la celda A1 a E37 de la hoja "Precisión Intermedia"
        df_precision = pd.read_excel(file_path, sheet_name="Precisión Intermedia", header=None)
        data_range_precision = df_precision.iloc[0:37, 0:5]  # filas 1 a 37, columnas A a E
        data_str_precision = data_range_precision.to_csv(index=False)
        prompt_precision = (
            "Analiza los siguientes datos extraídos del archivo Excel (hoja 'Precisión Intermedia', desde la celda A1 a E37) "
            "y proporciona un mensaje final sobre los resultados. La respuesta debe estar redactada de forma profesional, "
            "en un solo párrafo y de forma directa, sin detallar el proceso interno ni analizar gráficos.\n\n"
            f"{recomendaciones}\n\n"
            "Datos:\n" + data_str_precision + "\n\n"
            "Mensaje:"
        )
        response_precision = ollama.chat(model='llama3.2:3b', messages=[{'role': 'user', 'content': prompt_precision}])
        if isinstance(response_precision, dict):
            mensaje_precision = response_precision.get("message", {}).get("content", "")
        elif hasattr(response_precision, "message"):
            mensaje_precision = response_precision.message.content
        elif hasattr(response_precision, "content"):
            mensaje_precision = response_precision.content
        else:
            mensaje_precision = str(response_precision)
        mensaje_precision = re.sub(r'</?think>', '', mensaje_precision).strip()

        # ==================== Análisis 5: Veracidad ====================
        # Se extrae la información desde la celda B20 a M46 de la hoja "Estudio Veracidad"
        df_veracidad = pd.read_excel(file_path, sheet_name="Estudio Veracidad", header=None)
        data_range_veracidad = df_veracidad.iloc[19:46, 1:13]  # filas 20 a 46, columnas B a M
        data_str_veracidad = data_range_veracidad.to_csv(index=False)
        prompt_veracidad = (
            "Analiza los siguientes datos extraídos del archivo Excel (hoja 'Estudio Veracidad', desde la celda B20 a M46) "
            "y proporciona un mensaje final sobre los resultados. La respuesta debe estar redactada de forma profesional, "
            "en un solo párrafo y de forma directa, sin detallar el proceso interno ni analizar gráficos.\n\n"
            f"{recomendaciones}\n\n"
            "Datos:\n" + data_str_veracidad + "\n\n"
            "Mensaje:"
        )
        response_veracidad = ollama.chat(model='llama3.2:3b', messages=[{'role': 'user', 'content': prompt_veracidad}])
        if isinstance(response_veracidad, dict):
            mensaje_veracidad = response_veracidad.get("message", {}).get("content", "")
        elif hasattr(response_veracidad, "message"):
            mensaje_veracidad = response_veracidad.message.content
        elif hasattr(response_veracidad, "content"):
            mensaje_veracidad = response_veracidad.content
        else:
            mensaje_veracidad = str(response_veracidad)
        mensaje_veracidad = re.sub(r'</?think>', '', mensaje_veracidad).strip()

        # ==================== Análisis 6: Robustez ====================
        # Se extrae la información desde la celda B2 a K50 de la hoja "Robustez"
        df_robustez = pd.read_excel(file_path, sheet_name="Robustez", header=None)
        data_range_robustez = df_robustez.iloc[1:50, 1:11]  # filas 2 a 50, columnas B a K
        data_str_robustez = data_range_robustez.to_csv(index=False)
        prompt_robustez = (
            "Analiza los siguientes datos extraídos del archivo Excel (hoja 'Robustez', desde la celda B2 a K50) "
            "y proporciona un mensaje final sobre los resultados. La respuesta debe estar redactada de forma profesional, "
            "en un solo párrafo y de forma directa, sin detallar el proceso interno ni analizar gráficos.\n\n"
            f"{recomendaciones}\n\n"
            "Datos:\n" + data_str_robustez + "\n\n"
            "Mensaje:"
        )
        response_robustez = ollama.chat(model='llama3.2:3b', messages=[{'role': 'user', 'content': prompt_robustez}])
        if isinstance(response_robustez, dict):
            mensaje_robustez = response_robustez.get("message", {}).get("content", "")
        elif hasattr(response_robustez, "message"):
            mensaje_robustez = response_robustez.message.content
        elif hasattr(response_robustez, "content"):
            mensaje_robustez = response_robustez.content
        else:
            mensaje_robustez = str(response_robustez)
        mensaje_robustez = re.sub(r'</?think>', '', mensaje_robustez).strip()

        # ==================== Análisis 7: Incertidumbre ====================
        # Se extrae la información desde la celda A1 a I52 de la hoja "Estudio Incertidumbre"
        df_incertidumbre = pd.read_excel(file_path, sheet_name="Estudio Incertidumbre", header=None)
        data_range_incertidumbre = df_incertidumbre.iloc[0:52, 0:9]  # filas 1 a 52, columnas A a I
        data_str_incertidumbre = data_range_incertidumbre.to_csv(index=False)
        prompt_incertidumbre = (
            "Analiza los siguientes datos extraídos del archivo Excel (hoja 'Estudio Incertidumbre', desde la celda A1 a I52) "
            "y proporciona un mensaje final sobre los resultados. La respuesta debe estar redactada de forma profesional, "
            "en un solo párrafo y de forma directa, sin detallar el proceso interno ni analizar gráficos.\n\n"
            f"{recomendaciones}\n\n"
            "Datos:\n" + data_str_incertidumbre + "\n\n"
            "Mensaje:"
        )
        response_incertidumbre = ollama.chat(model='llama3.2:3b', messages=[{'role': 'user', 'content': prompt_incertidumbre}])
        if isinstance(response_incertidumbre, dict):
            mensaje_incertidumbre = response_incertidumbre.get("message", {}).get("content", "")
        elif hasattr(response_incertidumbre, "message"):
            mensaje_incertidumbre = response_incertidumbre.message.content
        elif hasattr(response_incertidumbre, "content"):
            mensaje_incertidumbre = response_incertidumbre.content
        else:
            mensaje_incertidumbre = str(response_incertidumbre)
        mensaje_incertidumbre = re.sub(r'</?think>', '', mensaje_incertidumbre).strip()

        # ==================== Escritura de Resultados en un solo bloque ====================
        # Se combinará el contenido en un solo bloque con la estructura similar a la salida en consola
        contenido = (
            f"Estudio de Linealidad:\n{mensaje_linealidad}\n\n"
            f"Estudio de Significante Confirmación de Linealidad:\n{mensaje_significante}\n\n"
            f"Estudio de Repetibilidad:\n{mensaje_repetibilidad}\n\n"
            f"Estudio de Precisión Intermedia:\n{mensaje_precision}\n\n"
            f"Estudio de Veracidad:\n{mensaje_veracidad}\n\n"
            f"Estudio de Robustez:\n{mensaje_robustez}\n\n"
            f"Estudio de Incertidumbre:\n{mensaje_incertidumbre}"
        )

        wb = load_workbook(file_path)
        if "Plan de validación" in wb.sheetnames:
            ws = wb["Plan de validación"]
        else:
            ws = wb.create_sheet("Plan de validación")

        # Fusionar las celdas desde B59 a AF180 para colocar la respuesta en un solo bloque
        ws.merge_cells("B59:AF210")
        celda_merged = ws["B59"]
        celda_merged.value = contenido

        # Formato de la celda fusionada para "embellecer" la salida
        celda_merged.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        celda_merged.font = Font(name="Calibri", size=10)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        celda_merged.border = thin_border
        celda_merged.fill = PatternFill("solid", fgColor="F2F2F2")  # fondo gris muy claro

        wb.save(file_path)
        return (numeric_prefix, file_name, mensaje_linealidad, mensaje_significante, mensaje_repetibilidad,
                mensaje_precision, mensaje_veracidad, mensaje_robustez, mensaje_incertidumbre)
    except Exception as e:
        return numeric_prefix, file_name, f"Error: {e}"

# ==================== Selección de Carpeta y Procesamiento de Archivos ====================
root = tk.Tk()
root.withdraw()
folder_path = filedialog.askdirectory(title="Selecciona la carpeta con los archivos Excel")
if not folder_path:
    console.print("[red]No se seleccionó ninguna carpeta.[/red]")
    exit()

excel_files = []
for file in os.listdir(folder_path):
    if file.lower().endswith('.xlsx'):
        m = re.match(r'^(\d+)', file)
        if m:
            numeric_prefix = int(m.group(1))
            excel_files.append((numeric_prefix, os.path.join(folder_path, file)))
excel_files.sort(key=lambda x: x[0])
if not excel_files:
    console.print("[red]No se encontraron archivos Excel con sufijo numérico al inicio del nombre.[/red]")
    exit()

results = []
with console.status("[bold green]Procesando archivos...", spinner="dots"):
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_file, file_tuple) for file_tuple in excel_files]
        for future in concurrent.futures.as_completed(futures):
            results.append(future.result())

results.sort(key=lambda x: x[0])

# Mostrar resultados en consola (se mantiene la misma estructura)
for result in results:
    if len(result) >= 9:
        (numeric_prefix, file_name, mensaje_linealidad, mensaje_significante, mensaje_repetibilidad,
         mensaje_precision, mensaje_veracidad, mensaje_robustez, mensaje_incertidumbre) = result
        contenido = (
            f"Estudio de Linealidad:\n{mensaje_linealidad}\n\n"
            f"Estudio de Significante Confirmación de Linealidad:\n{mensaje_significante}\n\n"
            f"Estudio de Repetibilidad:\n{mensaje_repetibilidad}\n\n"
            f"Estudio de Precisión Intermedia:\n{mensaje_precision}\n\n"
            f"Estudio de Veracidad:\n{mensaje_veracidad}\n\n"
            f"Estudio de Robustez:\n{mensaje_robustez}\n\n"
            f"Estudio de Incertidumbre:\n{mensaje_incertidumbre}"
        )
        panel = Panel(contenido, title=f"{file_name}", expand=False)
        console.print(panel)
    else:
        console.print(result)
