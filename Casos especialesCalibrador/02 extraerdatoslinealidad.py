import pandas as pd
import os
import sqlite3
from datetime import datetime
from tkinter import (Tk, Label, Entry, Button, Toplevel, Listbox, Scrollbar, END, 
                     filedialog, messagebox, Frame, Canvas, VERTICAL, LEFT, RIGHT, Y, BOTH)
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ====================== Base de Datos para Configuración ======================
DB_NAME = "skiprows_config.db"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    # Tabla para configuraciones de skip rows
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS configuraciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            skiprows TEXT NOT NULL,
            comentario TEXT,
            fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # Tabla para guardar selecciones de archivos de entrada y salida
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS selecciones_archivos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            archivo1 TEXT,
            archivo2 TEXT,
            archivo3 TEXT,
            archivo4 TEXT,
            archivo5 TEXT,
            archivo6 TEXT,
            archivo7 TEXT,
            archivo8 TEXT,
            archivo9 TEXT,
            archivo10 TEXT,
            archivo11 TEXT,
            archivo12 TEXT,
            output_directory TEXT,
            fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def save_skiprows_config(skiprows, comentario):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO configuraciones (skiprows, comentario)
        VALUES (?, ?)
    ''', (skiprows, comentario))
    conn.commit()
    conn.close()
    messagebox.showinfo("Configuración Guardada", "La configuración de skip rows se ha guardado.")

def load_latest_skiprows_config():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT skiprows, comentario, fecha FROM configuraciones ORDER BY id DESC LIMIT 1
    ''')
    row = cursor.fetchone()
    conn.close()
    return row

def search_skiprows_config(keyword):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, skiprows, comentario, fecha FROM configuraciones 
        WHERE skiprows LIKE ? OR comentario LIKE ? ORDER BY fecha DESC
    ''', (f'%{keyword}%', f'%{keyword}%'))
    rows = cursor.fetchall()
    conn.close()
    return rows

def save_selecciones(archivos, output_directory):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO selecciones_archivos 
        (archivo1, archivo2, archivo3, archivo4, archivo5, archivo6, archivo7, archivo8, archivo9, archivo10, archivo11, archivo12, output_directory)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (archivos[0], archivos[1], archivos[2], archivos[3], archivos[4], archivos[5],
          archivos[6], archivos[7], archivos[8], archivos[9], archivos[10], archivos[11],
          output_directory))
    conn.commit()
    conn.close()
    messagebox.showinfo("Selección Guardada", "La selección de archivos se ha guardado.")

def load_latest_selecciones():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT archivo1, archivo2, archivo3, archivo4, archivo5, archivo6, archivo7, archivo8, archivo9, archivo10, archivo11, archivo12, output_directory
        FROM selecciones_archivos ORDER BY id DESC LIMIT 1
    ''')
    row = cursor.fetchone()
    conn.close()
    return row

init_db()

# ===================== Funciones de la Interfaz =====================
def seleccionar_archivo(entry):
    filepath = filedialog.askopenfilename(
        title='Seleccionar archivo de Excel',
        filetypes=[('Archivos Excel', '*.xlsm *.xlsx *.xls')]
    )
    if filepath:
        entry.delete(0, END)
        entry.insert(0, filepath)

def seleccionar_directorio(entry):
    folder_path = filedialog.askdirectory(title="Seleccionar Carpeta de Destino")
    if folder_path:
        entry.delete(0, END)
        entry.insert(0, folder_path)

def guardar_config():
    skiprows_val = entry_skiprows.get()
    comentario_val = entry_comentario.get()
    if not skiprows_val:
        messagebox.showerror("Error", "El campo de Skip rows no puede estar vacío.")
        return
    save_skiprows_config(skiprows_val, comentario_val)

def cargar_ultima_config():
    config = load_latest_skiprows_config()
    if config:
        skiprows_val, comentario_val, fecha = config
        entry_skiprows.delete(0, END)
        entry_skiprows.insert(0, skiprows_val)
        entry_comentario.delete(0, END)
        entry_comentario.insert(0, comentario_val)
        messagebox.showinfo("Configuración Cargada", f"Última configuración guardada el {fecha}.")
    else:
        messagebox.showinfo("Sin Configuración", "No se encontraron configuraciones previas.")

def buscar_config():
    def realizar_busqueda():
        keyword = entry_busqueda.get()
        resultados = search_skiprows_config(keyword)
        listbox_resultados.delete(0, END)
        if resultados:
            for reg in resultados:
                reg_id, skiprows, comentario, fecha = reg
                listbox_resultados.insert(END, f"ID:{reg_id} | Skip rows: {skiprows} | Comentario: {comentario} | Fecha: {fecha}")
        else:
            listbox_resultados.insert(END, "No se encontraron resultados.")
    ventana_busqueda = Toplevel(root)
    ventana_busqueda.title("Buscar Configuración de Skip rows")
    Label(ventana_busqueda, text="Palabra clave:").grid(row=0, column=0, padx=10, pady=10)
    entry_busqueda = Entry(ventana_busqueda, width=40)
    entry_busqueda.grid(row=0, column=1, padx=10, pady=10)
    Button(ventana_busqueda, text="Buscar", command=realizar_busqueda).grid(row=0, column=2, padx=10, pady=10)
    listbox_resultados = Listbox(ventana_busqueda, width=100)
    listbox_resultados.grid(row=1, column=0, columnspan=3, padx=10, pady=10)
    scrollbar = Scrollbar(ventana_busqueda, orient=VERTICAL)
    scrollbar.grid(row=1, column=3, sticky='ns')
    listbox_resultados.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox_resultados.yview)

def guardar_seleccion_archivos():
    archivos = [entry.get() for entry in archivo_entries]
    out_dir = entry_output_dir.get()
    save_selecciones(archivos, out_dir)

def cargar_seleccion_archivos():
    datos = load_latest_selecciones()
    if datos:
        for i in range(12):
            archivo_entries[i].delete(0, END)
            archivo_entries[i].insert(0, datos[i] if datos[i] else "")
        entry_output_dir.delete(0, END)
        entry_output_dir.insert(0, datos[12] if datos[12] else "")
        messagebox.showinfo("Selección Cargada", "La última selección se ha cargado correctamente.")
    else:
        messagebox.showinfo("Sin Selección", "No se encontraron selecciones previas.")

# ===================== Función para limpiar y convertir "Promedio" =====================
def clean_promedio(val):
    if isinstance(val, str):
        val = val.strip().replace(',', '.')
        try:
            return float(val)
        except:
            return None
    return val

# ===================== Función para procesar archivos =====================
def procesar_archivos(skiprows_values, output_directory, archivo_entries):
    """
    Para cada valor en skiprows_values se procesa un grupo de datos:
      - Se extraen los datos (de las columnas 'Promedio' y 'LP') de cada archivo de entrada.
      - Se limpian los espacios en blanco de ambas columnas.
      - Se convierte "Promedio" a numérico (con clean_promedio) y se renombra a "LI".
      - Se eliminan las filas sin un valor numérico en "LI" y se ordenan de menor a mayor.
      - Se crea una hoja temporal en el archivo de salida para pegar y ordenar los datos.
      - Luego se copian los datos de la hoja temporal (sin encabezado) a las columnas B y C de la hoja
        "Linealidad - no parametrico", borrando los valores previos en ese rango, e insertando un encabezado.
    """
    archivos = [entry.get() for entry in archivo_entries]
    if any(not archivo for archivo in archivos):
        messagebox.showerror("Error", "Por favor, selecciona los 12 archivos de Excel.")
        return
    if not output_directory:
        messagebox.showerror("Error", "Por favor, selecciona una carpeta de destino.")
        return

    try:
        skiprows_list = list(map(int, [s.strip() for s in skiprows_values.split(',')]))
    except ValueError:
        messagebox.showerror("Error", "Los valores de 'Skip rows' deben ser enteros separados por comas.")
        return

    total_grupos = len(skiprows_list)
    progress_window = Toplevel(root)
    progress_window.title("Procesando...")
    progress_window.geometry("300x100")
    progress_label = Label(progress_window, text="Procesando... |")
    progress_label.pack(pady=10)
    progress_bar = ttk.Progressbar(progress_window, orient='horizontal', length=250, mode='determinate')
    progress_bar.pack(pady=10)
    progress_bar['maximum'] = total_grupos
    anim_frames = ["|", "/", "-", "\\"]

    for index, skiprows in enumerate(skiprows_list, start=1):
        datos_combinados = pd.DataFrame()
        for archivo in archivos:
            try:
                xls = pd.ExcelFile(archivo)
                if 'Registro' not in xls.sheet_names:
                    messagebox.showerror("Error", f"La hoja 'Registro' no se encuentra en:\n{archivo}")
                    progress_window.destroy()
                    return
                datos = pd.read_excel(archivo, sheet_name='Registro', skiprows=skiprows, nrows=10, usecols=['Promedio', 'LP'])
                # Limpiar espacios en ambas columnas
                datos["LP"] = datos["LP"].astype(str).str.strip()
                datos["Promedio"] = datos["Promedio"].astype(str).str.strip()
                # Convertir "Promedio" a numérico y renombrar a "LI"
                datos["LI"] = datos["Promedio"].apply(clean_promedio)
                # Conservar sólo las columnas "LP" y "LI"
                datos = datos[["LP", "LI"]]
                datos_combinados = pd.concat([datos_combinados, datos], ignore_index=True)
            except Exception as e:
                messagebox.showerror("Error al leer archivos", f"Ocurrió un error al leer {archivo}:\n{str(e)}")
                progress_window.destroy()
                return

        # Eliminar filas sin valor numérico en "LI" y ordenar de menor a mayor
        datos_combinados = datos_combinados.dropna(subset=["LI"])
        datos_combinados = datos_combinados.sort_values(by="LI", ascending=True)

        # Buscar el archivo de salida con el prefijo correspondiente (ej. "01", "02", ...)
        prefix = str(index).zfill(2)
        matching_files = [f for f in os.listdir(output_directory) if f.startswith(prefix)]
        if not matching_files:
            messagebox.showerror("Error", f"No se encontró un archivo con prefijo {prefix} en la carpeta de destino.")
            progress_window.destroy()
            return
        archivo_destino = os.path.join(output_directory, matching_files[0])
        try:
            wb = load_workbook(archivo_destino)
        except Exception as e:
            messagebox.showerror("Error al abrir archivo destino", f"No se pudo abrir {archivo_destino}:\n{str(e)}")
            progress_window.destroy()
            return

        # Verificar que la hoja "Linealidad - no parametrico" exista
        hoja_objetivo = "Linealidad - no parametrico"
        if hoja_objetivo not in wb.sheetnames:
            messagebox.showerror("Error", f"La hoja '{hoja_objetivo}' no se encontró en {archivo_destino}.")
            progress_window.destroy()
            return
        ws_target = wb[hoja_objetivo]

        # Crear una hoja temporal y pegar los datos ordenados
        temp_sheet_name = "TempSheet"
        if temp_sheet_name in wb.sheetnames:
            wb.remove(wb[temp_sheet_name])
        ws_temp = wb.create_sheet(temp_sheet_name)
        # Pegar los datos del DataFrame en la hoja temporal (incluyendo encabezados)
        rows = list(dataframe_to_rows(datos_combinados, index=False, header=True))
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws_temp.cell(row=r_idx, column=c_idx, value=value)

        # Borrar los datos previos en las columnas B y C de la hoja objetivo desde la fila 7 hasta la 1000
        for r in range(7, 1001):
            ws_target.cell(row=r, column=2).value = None
            ws_target.cell(row=r, column=3).value = None

        # Insertar encabezado con el propósito e instrucciones en la hoja objetivo (fila 7)
        ws_target.cell(row=7, column=2, value="LP")
        ws_target.cell(row=7, column=3, value="LI")
        # Puedes agregar más información de instrucciones en celdas adicionales si lo deseas.

        # Copiar los datos ordenados (sin encabezado) desde la hoja temporal a la hoja objetivo,
        # pegándolos en las columnas B y C a partir de la fila 8.
        target_row = 8
        # Asumimos que la fila 1 de ws_temp es el encabezado, así que comenzamos en la fila 2.
        for r in range(2, len(rows) + 1):
            lp_val = ws_temp.cell(row=r, column=1).value  # Columna 1: LP
            li_val = ws_temp.cell(row=r, column=2).value  # Columna 2: LI
            # Convertir LP a número si es posible
            try:
                lp_val = float(lp_val)
            except (ValueError, TypeError):
                lp_val = None
            ws_target.cell(row=target_row, column=2, value=lp_val)
            ws_target.cell(row=target_row, column=3, value=li_val)
            target_row += 1

        # Eliminar la hoja temporal
        wb.remove(ws_temp)

        try:
            wb.save(archivo_destino)
        except Exception as e:
            messagebox.showerror("Error al guardar archivo destino", f"No se pudo guardar {archivo_destino}:\n{str(e)}")
            progress_window.destroy()
            return

        progress_bar['value'] = index
        progress_label.config(text=f"Procesando... {anim_frames[index % len(anim_frames)]} Grupo {index}/{total_grupos}")
        progress_window.update_idletasks()

    progress_window.destroy()
    messagebox.showinfo("Éxito", f"Todos los datos fueron actualizados correctamente en:\n{output_directory}")

# ===================== Configuración de la Ventana Principal con Scroll =====================
root = Tk()
root.title("Extractor de Datos de Excel - Linealidad Paramétrico")
root.geometry("900x700")

canvas = Canvas(root)
canvas.pack(side=LEFT, fill=BOTH, expand=1)
scrollbar = Scrollbar(root, orient=VERTICAL, command=canvas.yview)
scrollbar.pack(side=RIGHT, fill=Y)
canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
main_frame = Frame(canvas)
canvas.create_window((0,0), window=main_frame, anchor="nw")

# -------------------- Encabezado --------------------
frame_encabezado = Frame(main_frame)
frame_encabezado.grid(row=0, column=0, columnspan=3, padx=10, pady=10)
titulo = Label(frame_encabezado, text="Extractor de Datos de Excel - Linealidad No Paramétrico", font=("Helvetica", 18, "bold"))
titulo.pack()
instrucciones = Label(frame_encabezado, text=(
    "Instrucciones:\n"
    "1. Selecciona 12 archivos de Excel que contengan la hoja 'Registro'.\n"
    "2. Ingresa los valores de 'Skip rows' (filas a omitir) separados por comas.\n"
    "3. Selecciona la carpeta de destino con archivos que tengan prefijos numéricos (01, 02, ...).\n"
    "   En cada archivo se actualizará la hoja 'Linealidad - no parametrico' con los datos de cada grupo,\n"
    "   ordenados de menor a mayor (por 'LI') y pegados en las columnas B y C a partir de la fila 8.\n"
    "4. Puedes guardar y gestionar la configuración de skip rows y la selección de archivos de entrada y salida."
), justify="center")
instrucciones.pack()

# -------------------- Selección de Archivos --------------------
archivo_entries = []
for i in range(12):
    Label(main_frame, text=f"Archivo {i+1}:").grid(row=i+1, column=0, padx=10, pady=5, sticky="e")
    entry = Entry(main_frame, width=70)
    entry.grid(row=i+1, column=1, padx=10, pady=5)
    Button(main_frame, text="Seleccionar", command=lambda entry=entry: seleccionar_archivo(entry)).grid(row=i+1, column=2, padx=10, pady=5)
    archivo_entries.append(entry)

# -------------------- Configuración de Skip Rows --------------------
Label(main_frame, text="Skip rows (separados por coma):").grid(row=13, column=0, padx=10, pady=10, sticky="e")
entry_skiprows = Entry(main_frame, width=70)
entry_skiprows.grid(row=13, column=1, padx=10, pady=10)
entry_skiprows.insert(0, "48, 62, 76, 90, 104, 119, 133, 147, 161")
Label(main_frame, text="Comentario configuración:").grid(row=14, column=0, padx=10, pady=10, sticky="e")
entry_comentario = Entry(main_frame, width=70)
entry_comentario.grid(row=14, column=1, padx=10, pady=10)
Button(main_frame, text="Guardar Configuración", command=guardar_config).grid(row=14, column=2, padx=10, pady=10)
Button(main_frame, text="Cargar Última Configuración", command=cargar_ultima_config).grid(row=15, column=2, padx=10, pady=10)
Button(main_frame, text="Buscar Configuración", command=buscar_config).grid(row=15, column=1, padx=10, pady=10, sticky="w")

# -------------------- Selección de Carpeta de Destino --------------------
Label(main_frame, text="Carpeta de destino:").grid(row=16, column=0, padx=10, pady=10, sticky="e")
entry_output_dir = Entry(main_frame, width=70)
entry_output_dir.grid(row=16, column=1, padx=10, pady=10)
Button(main_frame, text="Seleccionar", command=lambda: seleccionar_directorio(entry_output_dir)).grid(row=16, column=2, padx=10, pady=10)

# -------------------- Guardar/Cargar Selección de Archivos --------------------
Button(main_frame, text="Guardar Selección", command=guardar_seleccion_archivos).grid(row=17, column=0, padx=10, pady=10)
Button(main_frame, text="Cargar Selección", command=cargar_seleccion_archivos).grid(row=17, column=1, padx=10, pady=10)

# -------------------- Botón para Procesar Archivos --------------------
Button(main_frame, text="Procesar Archivos", font=("Helvetica", 12, "bold"),
       command=lambda: procesar_archivos(entry_skiprows.get(), entry_output_dir.get(), archivo_entries)
).grid(row=18, column=0, columnspan=3, pady=20)

root.mainloop()
