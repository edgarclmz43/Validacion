#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
===========================================================================
Aplicación para Evaluación de Precisión Intermedia según ISO 5725
===========================================================================
Este script procesa archivos de Excel para evaluar la precisión intermedia de un
método de medición de acuerdo con la norma ISO 5725. Para cada archivo se
realiza lo siguiente:

1. Se extraen los datos de medición de la hoja "sr ysR(Método) ISO 5725" en el
   rango G11:I50, que corresponden a tres niveles de medición: Bajo, Medio y Alto.
2. Se calcula la desviación estándar (s_I) de precisión intermedia para cada nivel.
3. Se extraen los valores de repetibilidad (s_r) y reproducibilidad (s_R) de celdas
   específicas (F64, G64, H64 para s_r y F69, G69, H69 para s_R) utilizando la carga
   del libro en modo 'data_only' para obtener los valores calculados.
4. Se evalúa la precisión intermedia comparando s_I con s_r y s_R, determinando si el
   método presenta baja variabilidad o si es necesario optimizar el proceso.
5. Se agrega una nueva hoja "Precisión Intermedia" en el mismo archivo Excel, en la
   que se documenta la evaluación, se presentan los resultados en una tabla y se
   incluye una conclusión detallada.
6. Se preservan las fórmulas y el formato original de las demás hojas, ya que el
   archivo se carga nuevamente sin el parámetro 'data_only' para la modificación.

Uso:
    - Seleccione la carpeta que contiene los archivos Excel a procesar.
    - El script procesará cada archivo con extensiones válidas (.xlsx, .xlsm, .xlsb, .xls)
      y agregará la hoja de resultados sin afectar el contenido original.

Autor: Ing. Edgar Colmenarez
Fecha: 2025-02-18

"""

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from tkinter import Tk
from tkinter.filedialog import askdirectory

def process_file(file_path):
    """
    Procesa un archivo Excel para evaluar la precisión intermedia según ISO 5725.

    Se extraen los datos de la hoja "sr ysR(Método) ISO 5725", se calculan las desviaciones
    estándar de precisión intermedia para tres niveles de medición, se obtienen los valores
    de repetibilidad y reproducibilidad y se evalúa la precisión intermedia. Finalmente, se crea
    una nueva hoja en el archivo Excel para documentar los resultados, preservando las fórmulas
    y el formato original de las demás hojas.

    Parámetros:
        file_path (str): Ruta completa al archivo Excel a procesar.
    """
    print(f"Procesando archivo: {file_path}")
    
    # 1. Extracción de datos de medición usando pandas
    try:
        # Se leen los datos del rango G11:I50 de la hoja "sr ysR(Método) ISO 5725"
        df_data = pd.read_excel(
            file_path,
            sheet_name='sr ysR(Método) ISO 5725',
            usecols="G:I",   # Columnas G, H e I
            skiprows=10,     # Se omiten las primeras 10 filas (la fila 11 es la primera con datos)
            nrows=40         # Se leen 40 filas (filas 11 a 50)
        )
    except Exception as e:
        print("Error al leer los datos de medición:", e)
        return

    # Renombrar las columnas para identificar los niveles
    df_data.columns = ['Bajo', 'Medio', 'Alto']

    # 2. Cálculo de la desviación estándar (s_I) para cada nivel (precisión intermedia)
    s_I_bajo  = np.std(df_data['Bajo'], ddof=1)
    s_I_medio = np.std(df_data['Medio'], ddof=1)
    s_I_alto  = np.std(df_data['Alto'], ddof=1)

    # 3. Cargar el workbook en modo 'data_only' para extraer los valores calculados (sin fórmulas)
    try:
        wb_data = load_workbook(file_path, data_only=True)
        ws_data = wb_data["sr ysR(Método) ISO 5725"]
    except Exception as e:
        print("Error al cargar el libro para extraer datos:", e)
        return

    # Extraer valores de repetibilidad (s_r) y reproducibilidad (s_R) de celdas específicas
    s_r_bajo  = ws_data['F64'].value
    s_r_medio = ws_data['G64'].value
    s_r_alto  = ws_data['H64'].value

    s_R_bajo  = ws_data['F69'].value
    s_R_medio = ws_data['G69'].value
    s_R_alto  = ws_data['H69'].value

    # 4. Función interna para evaluar la precisión intermedia
    def evaluate_precision(s_I, s_r, s_R):
        """
        Evalúa la precisión intermedia comparando s_I con s_r y s_R.

        Parámetros:
            s_I (float): Desviación estándar de precisión intermedia.
            s_r (float): Valor de repetibilidad.
            s_R (float): Valor de reproducibilidad.

        Retorna:
            str: Evaluación descriptiva de la precisión intermedia.
        """
        if s_r is None or s_R is None:
            return "Datos incompletos"
        if s_r == 0:
            if s_I == 0:
                return "Mediciones idénticas. Precisión óptima."
            else:
                return "Error: s_r es cero pero s_I es diferente de cero."
        diff = abs(s_I - s_r) / s_r 
        if diff < 0.1:
            return "La precisión intermedia es similar a la repetibilidad. Variabilidad baja y método estable."
        elif s_I <= s_R:
            return "La precisión intermedia se encuentra entre la repetibilidad y la reproducibilidad. Variabilidad leve; método aceptable."
        else:
            return "La precisión intermedia excede la reproducibilidad. Alta variabilidad; se recomienda revisar y optimizar el método."

    # Evaluar la precisión intermedia para cada nivel
    eval_bajo  = evaluate_precision(s_I_bajo,  s_r_bajo,  s_R_bajo)
    eval_medio = evaluate_precision(s_I_medio, s_r_medio, s_R_medio)
    eval_alto  = evaluate_precision(s_I_alto,  s_r_alto, s_R_alto)

    # 5. Crear un DataFrame con los resultados para la tabla
    results = pd.DataFrame({
        'Nivel': ['Bajo', 'Medio', 'Alto'],
        's_I (Precisión Intermedia)': [s_I_bajo, s_I_medio, s_I_alto],
        's_r (Repetibilidad)':         [s_r_bajo, s_r_medio, s_r_alto],
        's_R (Reproducibilidad)':      [s_R_bajo, s_R_medio, s_R_alto],
        'Evaluación':                  [eval_bajo, eval_medio, eval_alto]
    })

    # 6. Cargar nuevamente el workbook SIN data_only para preservar fórmulas y formato original
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print("Error al cargar el libro para modificar:", e)
        return

    # Crear o reemplazar la hoja "Precisión Intermedia" en la posición 07 del libro
    sheet_name = "Precisión Intermedia"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_result = wb.create_sheet(sheet_name, index=6)

    # 7. Escribir el contenido en la nueva hoja de resultados

    # Título de la hoja
    title = "Evaluación de Precisión Intermedia según ISO 5725"
    ws_result.merge_cells('A1:E1')
    cell = ws_result['A1']
    cell.value = title
    cell.font = Font(size=16, bold=True)
    cell.alignment = Alignment(horizontal="center")

    # Texto introductorio
    intro_text = (
        "Este informe presenta la evaluación de la precisión intermedia de un método de medición "
        "realizada en condiciones de laboratorio. Los datos corresponden a tres niveles de medición: "
        "Bajo, Medio y Alto. Para cada nivel se ha calculado la desviación estándar de precisión intermedia (s_I) "
        "y se han comparado con los valores de repetibilidad (s_r) y reproducibilidad (s_R), con el objetivo de "
        "determinar la estabilidad y confiabilidad del método de medida."
    )
    ws_result.merge_cells('A3:E6')
    cell_intro = ws_result['A3']
    cell_intro.value = intro_text
    cell_intro.alignment = Alignment(wrapText=True, horizontal="justify")
    cell_intro.font = Font(size=11)

    # Encabezado de la tabla de resultados
    header = list(results.columns)
    header_row = 8
    # Definir borde delgado para las celdas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Escribir los encabezados de la tabla
    for col_num, header_text in enumerate(header, start=1):
        cell = ws_result.cell(row=header_row, column=col_num, value=header_text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.border = thin_border
        ws_result.column_dimensions[get_column_letter(col_num)].width = 22

    # Escribir los datos de la tabla
    for i, row in results.iterrows():
        for j, value in enumerate(row, start=1):
            cell = ws_result.cell(row=header_row + 1 + i, column=j, value=value)
            cell.alignment = Alignment(horizontal="center", wrapText=True)
            cell.border = thin_border
            cell.font = Font(size=10)

    # 8. Sección de conclusiones (versión mejorada)
    conclusion_row = header_row + 1 + len(results) + 2
    ws_result.merge_cells(start_row=conclusion_row, start_column=1, end_row=conclusion_row, end_column=5)
    cell_concl_title = ws_result.cell(row=conclusion_row, column=1)
    cell_concl_title.value = "Conclusiones"
    cell_concl_title.font = Font(size=14, bold=True)
    cell_concl_title.alignment = Alignment(horizontal="center")

    conclusions_text = (
        "La evaluación de la precisión intermedia se ha realizado comparando la desviación estándar calculada "
        "a partir de los datos de medición (s_I) con los valores obtenidos de repetibilidad (s_r) y reproducibilidad (s_R) "
        "para cada nivel (Bajo, Medio y Alto).\n\n"
        "Criterios de evaluación aplicados:\n"
        "1. Si s_r es nulo o no se dispone de s_R, se indica que los datos son incompletos.\n"
        "2. Si s_r es 0 y s_I también es 0, se concluye que las mediciones son idénticas, reflejando una precisión óptima.\n"
        "3. Si s_r es 0 pero s_I es diferente de 0, se identifica una inconsistencia en los datos.\n"
        "4. Si la diferencia relativa entre s_I y s_r es inferior al 10% (|s_I - s_r|/s_r < 0.1), se determina que la precisión "
        "intermedia es muy similar a la repetibilidad, lo que indica baja variabilidad y estabilidad en el método.\n"
        "5. Si s_I se sitúa entre s_r y s_R, se considera que la precisión intermedia presenta una variabilidad leve, siendo el método aceptable.\n"
        "6. Si s_I excede a s_R, se evidencia una alta variabilidad, lo que sugiere que el método de medición debe ser revisado y optimizado.\n\n"
        "Resumen de resultados:\n"
        f"• Nivel Bajo: {eval_bajo}\n"
        f"• Nivel Medio: {eval_medio}\n"
        f"• Nivel Alto: {eval_alto}\n\n"
        "Se recomienda analizar detalladamente cada caso para implementar las mejoras necesarias en aquellos niveles donde "
        "la variabilidad resulte significativa."
    )
    start_row_concl = conclusion_row + 1
    ws_result.merge_cells(start_row=start_row_concl, start_column=1, end_row=start_row_concl+20, end_column=5)
    cell_concl = ws_result.cell(row=start_row_concl, column=1)
    cell_concl.value = conclusions_text
    cell_concl.alignment = Alignment(wrapText=True, horizontal="justify")
    cell_concl.font = Font(size=11)

    # 9. Guardar el archivo Excel preservando el contenido original de las demás hojas
    try:
        wb.save(file_path)
        print(f"Análisis completado en: {file_path}\n")
    except Exception as e:
        print("Error al guardar los resultados en Excel:", e)

def main():
    """
    Función principal que:
    - Solicita al usuario seleccionar una carpeta mediante un diálogo.
    - Procesa todos los archivos Excel en la carpeta que tengan las extensiones válidas.
    """
    # Ocultar la ventana principal de Tkinter
    Tk().withdraw()
    folder_path = askdirectory(title="Seleccione la carpeta con los archivos Excel")
    if not folder_path:
        print("No se seleccionó ninguna carpeta, saliendo del programa.")
        return

    # Extensiones válidas para los archivos Excel
    valid_extensions = ('.xlsx', '.xlsm', '.xlsb', '.xls')
    # Procesar cada archivo en la carpeta seleccionada
    for file in os.listdir(folder_path):
        if file.lower().endswith(valid_extensions):
            file_path = os.path.join(folder_path, file)
            process_file(file_path)

if __name__ == "__main__":
    main()
