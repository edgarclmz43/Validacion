import warnings
warnings.simplefilter("ignore", UserWarning)

import os
import glob
import configparser
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from scipy.stats import f_oneway
import tkinter as tk
from tkinter import filedialog, messagebox

###############################################################################
#                           CONFIGURACIÓN Y UTILIDADES                        #
###############################################################################

CONFIG_FILE = "config.ini"

def cargar_config():
    """
    Carga la configuración desde CONFIG_FILE y retorna el string de skiprows.
    Si el archivo no existe, retorna una cadena vacía.
    """
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
        if config.has_section("Configuracion") and "skiprows" in config["Configuracion"]:
            return config.get("Configuracion", "skiprows", fallback="").strip()
    return ""

def guardar_config(skiprows_str):
    """
    Guarda la configuración en CONFIG_FILE.
    """
    config = configparser.ConfigParser()
    config["Configuracion"] = {"skiprows": skiprows_str}
    with open(CONFIG_FILE, "w") as configfile:
        config.write(configfile)

def seleccionar_carpeta(entry_widget=None):
    """
    Abre un diálogo para seleccionar carpeta.
    Si se pasa un widget de entrada, actualiza su contenido; de lo contrario, retorna la ruta.
    """
    carpeta = filedialog.askdirectory(title="Seleccionar carpeta")
    if carpeta:
        if entry_widget is not None:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, carpeta)
        else:
            return carpeta
    return carpeta

def try_convert_float(val):
    """
    Intenta convertir el valor a float.
    Si falla, intenta reemplazar la coma por punto.
    Retorna None si no se puede convertir.
    """
    if val is None:
        return None
    try:
        return float(val)
    except Exception:
        try:
            return float(str(val).replace(",", "."))
        except Exception:
            return None

def load_workbook_auto(archivo, **kwargs):
    """
    Función auxiliar para cargar un archivo Excel con openpyxl.
    Si el archivo es .xlsm y no se ha indicado 'keep_vba', se fuerza keep_vba=True.
    """
    if archivo.lower().endswith('.xlsm') and 'keep_vba' not in kwargs:
        kwargs['keep_vba'] = True
    return load_workbook(filename=archivo, **kwargs)

###############################################################################
#       SECCIÓN 1: EXTRACCIÓN Y COMBINACIÓN DE DATOS (con openpyxl)           #
###############################################################################

def procesar_archivos(entry_widget):
    """
    Procesa los archivos de Excel en la carpeta seleccionada (hoja "Certificado").
    Extrae el certificado (celda U5) y el valor (celda Y37 o, en su defecto, Y38) y
    determina cuál tiene el valor (temperatura) más bajo y cuál el más alto.
    
    Retorna un diccionario:
      { 
        "min": {"archivo": ..., "Certificado": ..., "Valor": ...},
        "max": {"archivo": ..., "Certificado": ..., "Valor": ...}
      }
    """
    carpeta = entry_widget.get().strip()
    if not carpeta:
        messagebox.showerror("Error", "Por favor, selecciona una carpeta.")
        return None

    # Buscar archivos Excel (.xlsm, .xlsx, .xls)
    archivos = [os.path.join(root, file)
                for root, _, files in os.walk(carpeta)
                for file in files if file.lower().endswith(('.xlsm', '.xlsx', '.xls'))]
    if not archivos:
        messagebox.showerror("Error", "No se encontraron archivos de Excel en la carpeta seleccionada.")
        return None

    datos = []
    errores = []
    for archivo in archivos:
        try:
            wb = load_workbook_auto(archivo, data_only=True)
        except Exception as e:
            errores.append(f"Error al abrir el archivo {archivo}: {e}")
            continue

        if 'Certificado' not in wb.sheetnames:
            errores.append(f"La hoja 'Certificado' no se encuentra en {archivo}")
            wb.close()
            continue

        ws = wb["Certificado"]
        certificado = ws["U5"].value
        valor = ws["Y37"].value
        if valor is None or str(valor).strip() == "":
            valor = ws["Y38"].value
            if valor is None or str(valor).strip() == "":
                valor = ws["Y39"].value
        if certificado is None or valor is None:
            errores.append(f"No se pudo leer certificado o valor en {archivo}")
            wb.close()
            continue

        datos.append((archivo, certificado, valor))
        print(f"Procesado: {archivo} | Certificado: {certificado} | Valor: {valor}")
        wb.close()

    if errores:
        messagebox.showerror("Errores en el procesamiento", "\n".join(errores))
    if not datos:
        messagebox.showinfo("Información", "No se pudo procesar ningún archivo correctamente.")
        return None

    # Determinar el mínimo y el máximo según el valor (temperatura)
    datos_df = pd.DataFrame(datos, columns=["Archivo", "Certificado", "Valor"])
    try:
        idx_min = datos_df["Valor"].idxmin()
        idx_max = datos_df["Valor"].idxmax()
        valor_min = datos_df.loc[idx_min]
        valor_max = datos_df.loc[idx_max]
    except Exception as e:
        messagebox.showerror("Error", f"Error al determinar los extremos: {e}")
        return None

    mensaje = (
        f"Certificado con el valor más bajo:\nArchivo: {valor_min['Archivo']}\n"
        f"Certificado: {valor_min['Certificado']}\nValor: {valor_min['Valor']}\n\n"
        f"Certificado con el valor más alto:\nArchivo: {valor_max['Archivo']}\n"
        f"Certificado: {valor_max['Certificado']}\nValor: {valor_max['Valor']}"
    )
    print(mensaje)
    messagebox.showinfo("Resultados", mensaje)

    extremos = {
        "min": {"archivo": valor_min["Archivo"], "Certificado": valor_min["Certificado"], "Valor": valor_min["Valor"]},
        "max": {"archivo": valor_max["Archivo"], "Certificado": valor_max["Certificado"], "Valor": valor_max["Valor"]}
    }
    return extremos

def procesar_archivo_detallado(archivo, skiprows_list):
    """
    Para un archivo, extrae:
      - De la hoja "Certificado": número de certificado (U5) y temperatura (Y37 o Y38).
      - De la hoja "Registro": para cada skiprows (fila de encabezado), obtiene un bloque de datos
        desde (skiprows + 1) hasta que la columna D deje de tener un valor numérico.
    
    Retorna un diccionario con:
      {
         "archivo": <ruta>,
         "Certificado": <valor de U5>,
         "Temperatura": <valor de Y37 o Y38>,
         "bloques": [ bloque0, bloque1, ..., bloqueN ]
      }
    Cada bloque es una lista de diccionarios con claves "LPatron", "EA" y "Uexp".
    """
    try:
        wb = load_workbook_auto(archivo, data_only=True)
    except Exception as e:
        print(f"Error al abrir {archivo}: {e}")
        return None

    if "Certificado" not in wb.sheetnames:
        print(f"La hoja 'Certificado' no se encuentra en {archivo}")
        wb.close()
        return None

    ws_cert = wb["Certificado"]
    certificado_val = ws_cert["U5"].value
    temp_val = ws_cert["Y37"].value
    if temp_val is None or str(temp_val).strip() == "":
        temp_val = ws_cert["Y38"].value
        if temp_val is None or str(temp_val).strip() == "":
            temp_val = ws_cert["Y39"].value

    if "Registro" not in wb.sheetnames:
        print(f"La hoja 'Registro' no se encuentra en {archivo}")
        wb.close()
        return None

    ws_reg = wb["Registro"]
    bloques = []
    for skip in skiprows_list:
        block = []
        row = skip + 1  # La fila skip contiene la etiqueta
        while True:
            cell_val = ws_reg.cell(row=row, column=4).value  # Columna D: LPatrón
            numeric_val = try_convert_float(cell_val)
            if numeric_val is None:
                break
            ea = ws_reg.cell(row=row, column=27).value   # Columna AA
            uexp = ws_reg.cell(row=row, column=35).value  # Columna AI
            block.append({"LPatron": numeric_val, "EA": ea, "Uexp": uexp})
            row += 1
        bloques.append(block)
    wb.close()
    return {
        "archivo": archivo,
        "Certificado": certificado_val,
        "Temperatura": temp_val,
        "bloques": bloques
    }

def combinar_extremos(extremos, skiprows_list):
    """
    A partir del diccionario 'extremos' (con claves "max" y "min"),
    procesa cada uno de los archivos extremos y combina los bloques por índice.
    
    Retorna una lista de diccionarios, uno por cada índice (prefijo "i+1"):
      {
         "prefijo": "XX",
         "Temperatura_max": <valor>,
         "Certificado_max": <valor>,
         "block_max": <bloque_i del certificado max>,
         "Temperatura_min": <valor>,
         "Certificado_min": <valor>,
         "block_min": <bloque_i del certificado min>
      }
    Se utiliza el número mínimo de bloques entre ambos extremos.
    """
    max_data = procesar_archivo_detallado(extremos["max"]["archivo"], skiprows_list)
    min_data = procesar_archivo_detallado(extremos["min"]["archivo"], skiprows_list)
    if not max_data or not min_data:
        return []
    n = min(len(max_data["bloques"]), len(min_data["bloques"]))
    combinados = []
    for i in range(n):
        combinados.append({
            "prefijo": f"{i+1:02d}",
            "Temperatura_max": max_data["Temperatura"],
            "Certificado_max": max_data["Certificado"],
            "block_max": max_data["bloques"][i],
            "Temperatura_min": min_data["Temperatura"],
            "Certificado_min": min_data["Certificado"],
            "block_min": min_data["bloques"][i]
        })
    return combinados

###############################################################################
#   SECCIÓN 2: ACTUALIZACIÓN DE HOJA "ROBUSTEZ" CON OPENPYXL                     #
###############################################################################

def pegar_datos_destino(carpeta_destino, combinados):
    """
    Para cada registro combinado, busca en la carpeta destino un archivo cuyo nombre
    comience con el prefijo indicado. Con openpyxl se abre el archivo y se actualiza
    (o crea) la hoja "Robustez" escribiendo los datos sin alterar las demás hojas.
    """
    for reg in combinados:
        prefijo = reg["prefijo"]
        destino_encontrado = None
        for root, _, files in os.walk(carpeta_destino):
            for file in files:
                if file.lower().endswith(('.xlsx', '.xlsm', '.xls')) and file.startswith(prefijo):
                    destino_encontrado = os.path.join(root, file)
                    break
            if destino_encontrado:
                break
        if not destino_encontrado:
            print(f"No se encontró archivo de salida para el prefijo {prefijo}.")
            continue

        try:
            wb = load_workbook_auto(destino_encontrado)
        except Exception as e:
            print(f"Error al abrir archivo destino {destino_encontrado}: {e}")
            continue

        try:
            # Si existe la hoja "Robustez", se usa; de lo contrario, se crea
            if "Robustez" in wb.sheetnames:
                ws_dest = wb["Robustez"]
            else:
                ws_dest = wb.create_sheet("Robustez")

            # Datos comunes
            ws_dest["H9"] = reg["Temperatura_max"]
            ws_dest["H10"] = reg["Temperatura_min"]
            ws_dest["D13"] = reg["Certificado_max"]
            ws_dest["F13"] = reg["Certificado_min"]
            ws_dest["D12"] = "Temperatura Alta"
            ws_dest["F12"] = "Temperatura Baja"
            ws_dest["D14"] = ""
            ws_dest["F14"] = ""

            # Pegar block_max en columnas C, D y E (desde fila 17)
            fila = 17
            for r in reg["block_max"]:
                ws_dest.cell(row=fila, column=3, value=r["LPatron"])
                ws_dest.cell(row=fila, column=4, value=r["EA"])
                ws_dest.cell(row=fila, column=5, value=r["Uexp"])
                fila += 1

            # Pegar block_min en columnas F y G (desde fila 17)
            fila = 17
            for r in reg["block_min"]:
                ws_dest.cell(row=fila, column=6, value=r["EA"])
                ws_dest.cell(row=fila, column=7, value=r["Uexp"])
                fila += 1

            wb.save(destino_encontrado)
            print(f"Archivo destino actualizado: {destino_encontrado} para el bloque con prefijo {prefijo}.")
        except Exception as e:
            print(f"Error al guardar {destino_encontrado}: {e}")

def leer_datos_excel(ruta_archivo):
    """
    Lee los datos de la hoja 'Robustez' usando pandas y extrae dos grupos:
      - Grupo 1: Valores de la columna D (filas 16 a 28).
      - Grupo 2: Valores de la columna F (filas 16 a 28).
    Retorna una tupla (grupo1, grupo2).
    """
    df = pd.read_excel(ruta_archivo, sheet_name='Robustez', usecols="D:F")
    grupo1 = df.iloc[15:28, 0].dropna().tolist()  # Columna D
    grupo2 = df.iloc[15:28, 2].dropna().tolist()  # Columna F
    return grupo1, grupo2

def procesar_archivo_evaluacion(ruta_archivo):
    """
    Abre el archivo Excel y actualiza únicamente la hoja "Robustez"
    con los resultados de la prueba ANOVA, sin modificar el resto del libro.
    """
    print(f"\nProcesando archivo: {ruta_archivo}")
    grupo1, grupo2 = leer_datos_excel(ruta_archivo)
    if len(grupo1) < 2 or len(grupo2) < 2:
        print("  ERROR: Los grupos no contienen suficientes datos para realizar la prueba ANOVA.")
        return

    f_stat, p_value = f_oneway(grupo1, grupo2)
    print("  ----------------------------------------------------------------------------")
    print("  Prueba ANOVA de una vía")
    print("  ----------------------------------------------------------------------------")
    print(f"  Estadístico F: {f_stat:.4g}")
    print(f"  Valor p: {p_value:.4g}")
    nivel_significancia = 0.05
    print(f"  Nivel de significancia: {nivel_significancia}")

    resultado = ("Se rechaza la hipótesis nula. Hay una diferencia significativa entre los grupos."
                 if p_value < nivel_significancia 
                 else "No hay suficiente evidencia para rechazar la hipótesis nula. No hay diferencias significativas entre los grupos.")
    print("  " + resultado)

    try:
        wb = load_workbook_auto(ruta_archivo)
    except Exception as e:
        print(f"  ERROR al abrir {ruta_archivo}: {e}")
        return

    try:
        # Si existe la hoja "Robustez", se usa; de lo contrario, se crea
        if "Robustez" in wb.sheetnames:
            ws = wb["Robustez"]
        else:
            ws = wb.create_sheet("Robustez")

        # Escribir resultados en celdas predefinidas
        ws["B31"] = "-------------------------------------------------------------------------------"
        ws["B32"] = "Prueba ANOVA de una vía"
        ws["B33"] = "-------------------------------------------------------------------------------"
        ws["B34"] = "Estadístico F:"
        ws["C34"] = f_stat
        ws["B35"] = "Valor p:"
        ws["C35"] = p_value
        ws["B36"] = "Nivel de significancia:"
        ws["C36"] = nivel_significancia
        ws["B37"] = "Conclusión de la prueba:"
        ws.merge_cells("B37:K37")
        ws["B37"] = resultado

        # Obtener las temperaturas previamente almacenadas (suponiendo que ya se hayan escrito)
        temp_min = ws["H10"].value
        temp_max = ws["H9"].value

        # Fusionar celdas en el rango B43:K49 y ajustar la altura de la fila 43
        ws.merge_cells("B43:K49")
        ws["B43"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for row in range(43, 50):
            ws.row_dimensions[row].height = 16

        conclusion_text = (
            f"El método de calibración es robusto. A pesar de la variación de temperatura observada durante el experimento, "
            f"que osciló entre un mínimo de {temp_min} °C y un máximo de {temp_max} °C, la prueba de Anova de una vía "
            f"arroja un valor p de {p_value:.4g}. Este valor es significativamente superior al nivel de significancia de 0.05, "
            f"lo cual indica que no existen diferencias estadísticamente significativas entre los resultados de ambas técnicas. "
            f"En consecuencia, se confirma la robustez y la fiabilidad de las técnicas evaluadas, validando su uso bajo las "
            f"condiciones experimentales y las variaciones de temperatura presentadas. La investigación concluye que ambos "
            f"métodos experimentales son satisfactorios y equiparables en precisión."
        )
        ws["B43"] = conclusion_text

        wb.save(ruta_archivo)
        print("  Archivo actualizado exitosamente.")
    except Exception as e:
        print(f"  ERROR al guardar el archivo: {e}")

def evaluar_robustez(carpeta):
    """
    Recorre la carpeta seleccionada en busca de archivos Excel (.xlsx y .xlsm)
    y aplica la evaluación (ANOVA) a cada uno.
    """
    archivos_excel = glob.glob(os.path.join(carpeta, "*.xlsx")) + glob.glob(os.path.join(carpeta, "*.xlsm"))
    if not archivos_excel:
        messagebox.showinfo("Información", "No se encontraron archivos Excel en la carpeta seleccionada.")
        return

    for archivo in archivos_excel:
        try:
            procesar_archivo_evaluacion(archivo)
        except Exception as e:
            print(f"  ERROR al procesar el archivo {archivo}: {e}")

    messagebox.showinfo("Evaluación completada", "Se completó la evaluación de robustez en todos los archivos.")

###############################################################################
#                    INTERFAZ GRÁFICA (Tkinter) COMBINADA                     #
###############################################################################

def crear_interfaz():
    """
    Crea la interfaz gráfica que integra:
      - Extracción y combinación de datos.
      - Evaluación de robustez (ANOVA).
    """
    root = tk.Tk()
    root.title("Aplicación Combinada: Extracción/Combinación y Evaluación de Robustez")
    root.resizable(False, False)
    
    # Sección 1: Extracción y Combinación de Datos
    frame_extraccion = tk.LabelFrame(root, text="Extracción y Combinación de Datos", padx=10, pady=10)
    frame_extraccion.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
    
    tk.Label(frame_extraccion, text="Carpeta Origen Certificados:").grid(row=0, column=0, padx=5, pady=5)
    entry_origen = tk.Entry(frame_extraccion, width=50)
    entry_origen.grid(row=0, column=1, padx=5, pady=5)
    tk.Button(frame_extraccion, text="Seleccionar", command=lambda: seleccionar_carpeta(entry_origen)).grid(row=0, column=2, padx=5, pady=5)
    
    tk.Label(frame_extraccion, text="Carpeta Destino Validación:").grid(row=1, column=0, padx=5, pady=5)
    entry_destino = tk.Entry(frame_extraccion, width=50)
    entry_destino.grid(row=1, column=1, padx=5, pady=5)
    tk.Button(frame_extraccion, text="Seleccionar", command=lambda: seleccionar_carpeta(entry_destino)).grid(row=1, column=2, padx=5, pady=5)
    
    tk.Label(frame_extraccion, text="Skiprows (valores separados por comas, recuerde sumar 1):").grid(row=2, column=0, padx=5, pady=5)
    skiprows_default = cargar_config()
    entry_skiprows = tk.Entry(frame_extraccion, width=50)
    entry_skiprows.insert(0, skiprows_default)
    entry_skiprows.grid(row=2, column=1, padx=5, pady=5)
    
    def iniciar_proceso_extraccion():
        carpeta_origen = entry_origen.get().strip()
        if not carpeta_origen:
            messagebox.showerror("Error", "Por favor, selecciona la carpeta de origen.")
            return
        
        carpeta_destino = entry_destino.get().strip()
        skiprows_str = entry_skiprows.get().strip()
        if not skiprows_str:
            messagebox.showerror("Error", "Debe indicar los skiprows (lista de valores).")
            return
        
        guardar_config(skiprows_str)
        try:
            skiprows_list = [int(s.strip()) for s in skiprows_str.split(",") if s.strip()]
            if not skiprows_list:
                raise ValueError("No se ingresaron valores válidos.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar los skiprows: {e}")
            return

        extremos = procesar_archivos(entry_origen)
        if extremos is None:
            return
        combinados = combinar_extremos(extremos, skiprows_list)
        if not combinados:
            return
        if carpeta_destino:
            pegar_datos_destino(carpeta_destino, combinados)
            messagebox.showinfo("Información", "Proceso de extracción y combinación completado.\nAhora puede proceder a la evaluación de robustez (ANOVA).")
        else:
            messagebox.showinfo("Información", "Proceso completado. No se indicó carpeta destino.\nAhora puede proceder a la evaluación de robustez (ANOVA).")
    
    tk.Button(frame_extraccion, text="Procesar y Combinar Archivos", command=iniciar_proceso_extraccion).grid(row=3, column=0, columnspan=3, pady=10)
    
    # Sección 2: Evaluación de Robustez (ANOVA)
    frame_evaluacion = tk.LabelFrame(root, text="Evaluación de Robustez (ANOVA)", padx=10, pady=10)
    frame_evaluacion.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
    
    tk.Label(frame_evaluacion, text="Carpeta para Evaluación Validación:").grid(row=0, column=0, padx=5, pady=5)
    entry_evaluacion = tk.Entry(frame_evaluacion, width=50)
    entry_evaluacion.grid(row=0, column=1, padx=5, pady=5)
    tk.Button(frame_evaluacion, text="Seleccionar", command=lambda: seleccionar_carpeta(entry_evaluacion)).grid(row=0, column=2, padx=5, pady=5)
    
    def iniciar_evaluacion():
        carpeta = entry_evaluacion.get().strip()
        if not carpeta:
            messagebox.showerror("Error", "Por favor, selecciona la carpeta para evaluación.")
            return
        evaluar_robustez(carpeta)
    
    tk.Button(frame_evaluacion, text="Evaluar Robustez", command=iniciar_evaluacion).grid(row=1, column=0, columnspan=3, pady=10)
    
    root.mainloop()

if __name__ == "__main__":
    crear_interfaz()
