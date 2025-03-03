"""
================================================================================
Actualización de Archivos Excel
================================================================================

Propósito:
  Esta aplicación automatiza la actualización de archivos Excel que cumplen con un
  patrón específico (archivos que comienzan con dos dígitos y extensión .xlsx).
  Para cada archivo, se busca la hoja "sr ysR(Método) ISO 5725" y se extraen los
  valores de las celdas G10, H10 e I10 (representando medias). Con estos valores se
  generan 40 datos distribuidos normalmente (utilizando una desviación configurable)
  y se insertan en el rango G11:I50 de la misma hoja. Al finalizar, se abre la
  carpeta seleccionada para que el usuario pueda revisar los archivos actualizados.

Instrucciones de Uso:
  1. Selecciona o ingresa la carpeta donde se encuentran los archivos Excel.
  2. Ingresa el valor de desviación deseado para la generación de datos.
  3. Presiona "Procesar Archivos" para iniciar el proceso.
  4. Se buscarán archivos que comiencen con dos dígitos y tengan extensión .xlsx.
  5. Los valores de carpeta y desviación se guardarán en una base de datos para usos futuros.
  6. Al finalizar se abrirá la carpeta para revisar los archivos actualizados.
"""

import numpy as np
import openpyxl
import glob
import locale
import os
import sqlite3
import tkinter as tk
from tkinter import filedialog, messagebox

# ------------------------------------------------------------------------------
# Configuración regional: Se intenta usar "es_ES.UTF-8" para el manejo de formatos.
# ------------------------------------------------------------------------------
try:
    locale.setlocale(locale.LC_NUMERIC, 'es_ES.UTF-8')
except locale.Error:
    print("La configuración regional 'es_ES.UTF-8' no está disponible en este sistema.")

# ------------------------------------------------------------------------------
# Funciones para la Base de Datos (Configuración)
# ------------------------------------------------------------------------------
DB_PATH = "settings.db"

def crear_conexion():
    """Crea y retorna una conexión a la base de datos SQLite."""
    return sqlite3.connect(DB_PATH)

def inicializar_db():
    """Crea la tabla de configuración si no existe."""
    conn = crear_conexion()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS config (
            id INTEGER PRIMARY KEY,
            folder TEXT,
            desviacion TEXT
        )
    """)
    conn.commit()
    conn.close()

def cargar_config():
    """Carga la configuración (carpeta y desviación) de la base de datos."""
    conn = crear_conexion()
    cursor = conn.cursor()
    cursor.execute("SELECT folder, desviacion FROM config ORDER BY id DESC LIMIT 1")
    row = cursor.fetchone()
    conn.close()
    if row:
        return row[0], row[1]
    else:
        return "", "0.0001"

def guardar_config(folder, desviacion):
    """Guarda la configuración en la base de datos."""
    conn = crear_conexion()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO config (folder, desviacion) VALUES (?, ?)", (folder, desviacion))
    conn.commit()
    conn.close()

# Inicializar la base de datos
inicializar_db()

# ------------------------------------------------------------------------------
# Funciones de la Aplicación
# ------------------------------------------------------------------------------
def seleccionar_carpeta():
    """Abre el diálogo para seleccionar una carpeta y actualiza el Entry correspondiente."""
    folder_path = filedialog.askdirectory(title="Selecciona la carpeta con los archivos Excel")
    if folder_path:
        entry_carpeta.delete(0, tk.END)
        entry_carpeta.insert(0, folder_path)

def procesar_archivos():
    """Procesa los archivos Excel de la carpeta seleccionada según el patrón definido."""
    folder_path = entry_carpeta.get().strip()
    desviacion_texto = entry_desviacion.get().strip()
    
    if not folder_path:
        messagebox.showerror("Error", "Por favor, selecciona una carpeta.")
        return
    try:
        desviacion_val = float(desviacion_texto)
    except ValueError:
        messagebox.showerror("Error", "El valor de la desviación debe ser numérico.")
        return

    # Guardar configuración actual en la base de datos
    guardar_config(folder_path, desviacion_texto)
    
    # Actualizar el estado de procesamiento
    status_label.config(text="Procesando archivos...", fg="blue")
    ventana.update_idletasks()
    
    # Buscar archivos Excel que comiencen con dos dígitos
    pattern = os.path.join(folder_path, "[0-9][0-9]*.xlsx")
    archivos = glob.glob(pattern)
    
    if not archivos:
        messagebox.showinfo("Información", "No se encontraron archivos que coincidan con el patrón en la carpeta especificada.")
        status_label.config(text="No se encontraron archivos.", fg="red")
        return
    
    for archivo in archivos:
        try:
            wb = openpyxl.load_workbook(archivo)
        except Exception as e:
            print(f"Error al cargar el archivo {archivo}: {e}")
            continue
        
        # Verificar existencia de la hoja requerida
        if "sr ysR(Método) ISO 5725" in wb.sheetnames:
            ws = wb["sr ysR(Método) ISO 5725"]
            # Leer valores de celdas G10, H10 e I10
            media1 = ws["G10"].value
            media2 = ws["H10"].value
            media3 = ws["I10"].value
            
            if media1 is None or media2 is None or media3 is None:
                print(f"En el archivo {archivo} faltan valores en G10, H10 o I10.")
                continue
            
            # Generar 40 datos distribuidos normalmente para cada media
            datos1 = np.random.normal(loc=media1, scale=desviacion_val, size=40)
            datos2 = np.random.normal(loc=media2, scale=desviacion_val, size=40)
            datos3 = np.random.normal(loc=media3, scale=desviacion_val, size=40)
            
            # Insertar los datos generados en el rango G11:I50 con formato de 7 decimales
            for i in range(40):
                cell1 = ws.cell(row=11 + i, column=7, value=float(datos1[i]))
                cell1.number_format = "0.0000000"
                
                cell2 = ws.cell(row=11 + i, column=8, value=float(datos2[i]))
                cell2.number_format = "0.0000000"
                
                cell3 = ws.cell(row=11 + i, column=9, value=float(datos3[i]))
                cell3.number_format = "0.0000000"
            
            wb.save(archivo)
            print(f"Archivo {archivo} actualizado correctamente.")
        else:
            print(f"El archivo {archivo} no tiene la hoja 'sr ysR(Método) ISO 5725'.")
    
    # Actualizar estado y abrir carpeta para revisión
    status_label.config(text="Procesamiento completado.", fg="green")
    os.startfile(folder_path)
    messagebox.showinfo("Proceso Finalizado", f"Se han procesado los archivos.\nSe ha abierto la carpeta:\n{folder_path}")

# ------------------------------------------------------------------------------
# Configuración de la Interfaz Gráfica
# ------------------------------------------------------------------------------
ventana = tk.Tk()
ventana.title("Actualización de Archivos Excel")
ventana.geometry("700x600")
ventana.resizable(True, True)

# Fuente base para la interfaz
fuente = ("Helvetica", 11)

# Frame principal
frame = tk.Frame(ventana, padx=20, pady=20)
frame.pack(expand=True, fill="both")

# Título de la aplicación
etiqueta_titulo = tk.Label(frame, text="Actualización de Archivos Excel", font=("Helvetica", 18, "bold"))
etiqueta_titulo.pack(pady=(0, 20))

# Frame de Instrucciones
frame_instrucciones = tk.LabelFrame(frame, text="Instrucciones", font=("Helvetica", 12, "bold"), padx=15, pady=15)
frame_instrucciones.pack(fill="both", expand=True, padx=10, pady=10)

instrucciones_texto = (
    "• Ingresa o selecciona la carpeta donde se encuentran los archivos Excel a procesar.\n\n"
    "• Ingresa el valor de desviación deseado para la generación de datos.\n\n"
    "• Presiona 'Procesar Archivos' para iniciar el proceso.\n\n"
    "• Se buscarán archivos que tengan un prefijo de dos dígitos y extensión .xlsx. Para cada uno, "
    "se actualizará la hoja 'sr ysR(Método) ISO 5725' con datos generados a partir de las celdas G10, H10 e I10.\n\n"
    "• Los valores ingresados se guardarán para futuros usos.\n\n"
    "• Al finalizar, se abrirá la carpeta seleccionada para que puedas revisar los archivos actualizados."
)
etiqueta_instrucciones = tk.Label(frame_instrucciones, text=instrucciones_texto, font=fuente, justify="left", wraplength=630)
etiqueta_instrucciones.pack()

# Frame para la selección de carpeta
frame_carpeta = tk.Frame(frame)
frame_carpeta.pack(fill="x", pady=(10, 10))
etiqueta_carpeta = tk.Label(frame_carpeta, text="Carpeta con archivos:", font=fuente)
etiqueta_carpeta.pack(side="left", padx=(0, 10))
entry_carpeta = tk.Entry(frame_carpeta, font=fuente, width=50)
entry_carpeta.pack(side="left", fill="x", expand=True)
boton_buscar = tk.Button(frame_carpeta, text="Buscar", font=fuente, command=seleccionar_carpeta)
boton_buscar.pack(side="left", padx=(10, 0))

# Frame para la entrada de la desviación
frame_desviacion = tk.Frame(frame)
frame_desviacion.pack(fill="x", pady=(10, 20))
etiqueta_desviacion = tk.Label(frame_desviacion, text="Valor de Desviación:", font=fuente)
etiqueta_desviacion.pack(side="left", padx=(0, 10))
entry_desviacion = tk.Entry(frame_desviacion, font=fuente, width=20)
entry_desviacion.pack(side="left", fill="x", expand=True)

# Botón para procesar archivos
boton_procesar = tk.Button(frame_desviacion, text="Procesar Archivos", font=("Helvetica", 12, "bold"),
                           bg="#4CAF50", fg="white", padx=15, pady=8, command=procesar_archivos)
boton_procesar.pack(side="left", padx=(10, 0))

# Etiqueta de estado para mostrar el progreso
status_label = tk.Label(frame, text="", font=fuente, fg="blue")
status_label.pack(pady=10)

# Cargar la configuración guardada y actualizar los campos correspondientes
config_folder, config_desviacion = cargar_config()
entry_carpeta.insert(0, config_folder)
entry_desviacion.insert(0, config_desviacion)

ventana.mainloop()
