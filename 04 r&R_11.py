#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
===========================================================================
Aplicación para Evaluación de Repetibilidad y Reproducibilidad según ISO 5725
===========================================================================

Descripción:
    Este script procesa archivos de Excel para evaluar la repetibilidad (s_r) y
    la reproducibilidad (s_R) de un método de medición conforme a la norma ISO 5725.
    Para cada archivo se realiza lo siguiente:

    1. Se extraen los valores de repetibilidad (s_r) y reproducibilidad (s_R) de la
       hoja "sr ysR(Método) ISO 5725" usando un workbook cargado en modo 'data_only'
       para obtener los resultados de las fórmulas (sin modificar el contenido original).
       - s_r: celdas F64, G64 y H64.
       - s_R: celdas F69, G69 y H69.
    2. Se evalúa cada valor comparándolo con un criterio establecido (<= 10.0) y se
       determina si es "Aceptable" o "No Aceptable".
    3. Se genera un informe detallado y profesional, en el que cada dato (valor y
       evaluación) se presenta en líneas separadas.
    4. Se vuelve a cargar el archivo en modo normal (sin 'data_only') para preservar
       las fórmulas y el formato original, y se inserta el informe en la misma hoja
       "sr ysR(Método) ISO 5725" a partir de la celda B115, colocando cada línea en
       una celda individual.
    5. Se guarda el archivo sin afectar el contenido original de las demás celdas.

Requisitos:
    - Python 3.x
    - openpyxl
    - tkinter

Uso:
    Ejecutar el script y utilizar la interfaz gráfica para seleccionar la carpeta
    que contiene los archivos Excel a evaluar.

Autor: Ing. Edgar Colmenarez
Fecha: 2025-02-18
=========================================================================== 
"""

import os
import glob
import tkinter as tk
from tkinter.filedialog import askdirectory
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# Definición del criterio de evaluación y del nombre de la hoja de datos
CRITERIO = 10.0
HOJA_DATOS = "sr ysR(Método) ISO 5725"

def sanitize_text(text):
    """
    Elimina caracteres no compatibles (no ASCII) del texto.
    
    Si se detectan caracteres incompatibles, se remueven y se informa en consola.
    
    Parámetros:
        text (str): Texto a sanitizar.
        
    Retorna:
        str: Texto sanitizado.
    """
    try:
        sanitized = text.encode('ascii', 'ignore').decode('ascii')
        if sanitized != text:
            print("Se han removido caracteres incompatibles.")
        return sanitized
    except Exception as e:
        print(f"Error en sanitización del texto: {e}")
        return text

def to_float(valor):
    """
    Convierte el valor a tipo float.
    
    Si el valor es una cadena que utiliza coma decimal, la reemplaza por punto.
    Retorna None si el valor es None o no se puede convertir.
    """
    try:
        if valor is None:
            return None
        if isinstance(valor, str):
            valor = valor.replace(',', '.')
        return float(valor)
    except Exception:
        return None

def evaluar_valor(valor):
    """
    Evalúa el valor comparándolo con el CRITERIO.
    
    Retorna:
        "Aceptable" si el valor es menor o igual a CRITERIO; de lo contrario, "No Aceptable".
        Si el valor es None, retorna "No evaluado".
    """
    if valor is None:
        return "No evaluado"
    return "Aceptable" if valor <= CRITERIO else "No Aceptable"

def formato_valor(valor):
    """
    Formatea el valor para presentación, mostrando en notación científica.
    """
    return f"{valor:.2e}" if valor is not None else "N/D"

def generar_informe(file_path):
    """
    Genera el informe de evaluación de repetibilidad (s_r) y reproducibilidad (s_R).

    Se extraen los valores de la hoja de datos usando un workbook en modo 'data_only'
    y se evalúan. Luego se vuelve a cargar el archivo en modo normal para insertar
    el informe en la misma hoja "sr ysR(Método) ISO 5725" a partir de la celda B115,
    colocando cada línea en celdas individuales.

    Parámetros:
        file_path (str): Ruta completa del archivo Excel.

    Retorna:
        bool: True si la evaluación se realizó correctamente; False en caso de error.
    """
    print(f"Procesando archivo: {file_path}")

    # --- Paso 1: Extraer valores en modo data_only ---
    try:
        wb_data = load_workbook(file_path, data_only=True)
        ws_data = wb_data[HOJA_DATOS]
    except Exception as e:
        print(f"Error al cargar el libro en modo data_only: {e}")
        return False

    # Extraer valores de repetibilidad (s_r) y reproducibilidad (s_R)
    s_r_bajo  = to_float(ws_data['F64'].value)
    print(s_r_bajo)
    s_r_medio = to_float(ws_data['G64'].value)
    print(s_r_medio)
    s_r_alto  = to_float(ws_data['H64'].value)
    print(s_r_alto)
    s_R_bajo  = to_float(ws_data['F69'].value)
    print(s_R_bajo)
    s_R_medio = to_float(ws_data['G69'].value)
    print(s_R_medio)
    s_R_alto  = to_float(ws_data['H69'].value)
    print(s_R_alto)

    # Preparar listas para evaluación y presentación
    s_r_values = [s_r_bajo, s_r_medio, s_r_alto]
    s_R_values = [s_R_bajo, s_R_medio, s_R_alto]

    s_r_textos = [formato_valor(v) for v in s_r_values]
    s_R_textos = [formato_valor(v) for v in s_R_values]

    s_r_eval = [evaluar_valor(v) for v in s_r_values]
    s_R_eval = [evaluar_valor(v) for v in s_R_values]

    # --- Generar conclusiones específicas ---
    non_aceptable_sr = []
    non_aceptable_SR = []
    niveles = ["Bajo", "Medio", "Alto"]

    if s_r_eval[0] != "Aceptable":
        non_aceptable_sr.append("Bajo")
    if s_r_eval[1] != "Aceptable":
        non_aceptable_sr.append("Medio")
    if s_r_eval[2] != "Aceptable":
        non_aceptable_sr.append("Alto")

    if s_R_eval[0] != "Aceptable":
        non_aceptable_SR.append("Bajo")
    if s_R_eval[1] != "Aceptable":
        non_aceptable_SR.append("Medio")
    if s_R_eval[2] != "Aceptable":
        non_aceptable_SR.append("Alto")

    conclusion = "CONCLUSIONES:\n"
    if non_aceptable_sr:
        conclusion += f" - La repetibilidad (s_r) no es aceptable en el/los nivel(es): {', '.join(non_aceptable_sr)}.\n"
    else:
        conclusion += " - La repetibilidad (s_r) es aceptable en todos los niveles.\n"

    if non_aceptable_SR:
        conclusion += f" - La reproducibilidad (s_R) no es aceptable en el/los nivel(es): {', '.join(non_aceptable_SR)}.\n"
    else:
        conclusion += " - La reproducibilidad (s_R) es aceptable en todos los niveles.\n"

    if not non_aceptable_sr and not non_aceptable_SR:
        conclusion += "\nEn general, la repetibilidad y reproducibilidad del metodo son aceptables.\n"
    else:
        conclusion += "\nEn general, el metodo presenta deficiencias en los niveles indicados y se recomienda revisar el proceso de medicion.\n"

    # --- Paso 2: Construir la narrativa del informe ---
    informe = (
        "************************************************************\n"
        "        INFORME DE EVALUACION DEL METODO ISO 5725\n"
        "************************************************************\n\n"
        "1. Evaluacion de Repetibilidad (s_r):\n"
        "------------------------------------------------------------\n"
        "   - Valor Bajo de Sr:\n"
        f"         - Valor: {s_r_textos[0]}\n"
        f"         - Evaluacion: {s_r_textos[0]} Menor que 10.0\n"
        f"         - Decision: {s_r_eval[0]}\n\n"
        "   - Valor Medio de Sr:\n"
        f"         - Valor: {s_r_textos[1]}\n"
        f"         - Evaluacion: {s_r_textos[1]} Menor que 10.0\n"
        f"         - Decision: {s_r_eval[1]}\n\n"
        "   - Valor Alto de Sr:\n"
        f"         - Valor: {s_r_textos[2]}\n"
        f"         - Evaluacion: {s_r_textos[2]} Menor que 10.0\n"
        f"         - Desición: {s_r_eval[2]}\n"
        "------------------------------------------------------------\n\n"
        "2. Evaluacion de Reproducibilidad (s_R):\n"
        "------------------------------------------------------------\n"
        "   - Valor Bajo de SR:\n"
        f"         - Valor: {s_R_textos[0]}\n"
        f"         - Evaluacion: {s_R_textos[0]} Menor que 10.0\n"
        f"         - Decision: {s_R_eval[0]}\n\n"
        "   - Valor Medio de SR:\n"
        f"         - Valor: {s_R_textos[1]}\n"
        f"         - Evaluacion: {s_R_textos[1]} Menor que 10.0\n"
        f"         - Decision: {s_R_eval[1]}\n\n"
        "   - Valor Alto de SR:\n"
        f"         - Valor: {s_R_textos[2]}\n"
        f"         - Evaluacion: {s_R_textos[2]} Menor que 10.0\n"
        f"         - Decision: {s_R_eval[2]}\n\n"
        "------------------------------------------------------------\n\n"
        "Recomendaciones:\n"
        "   - Revisar y ajustar los parametros en caso de que algun valor no cumpla con el criterio (<= 10.0).\n"
        "   - Mantener un riguroso control de las condiciones de medicion.\n"
        "   - Realizar mediciones adicionales para verificar la precision del metodo.\n"
        "   - Los valores de s_r y s_R deben ser menores o iguales a 10.0 (en porcentaje) para ser aceptables.\n\n"
        "************************************************************\n"
        "Evaluacion de la Repetibilidad y Reproducibilidad.\n"
        "************************************************************\n\n"
    )

    # Agregar las conclusiones al informe
    informe += conclusion

    # --- Sanitizar el informe para eliminar caracteres incompatibles ---
    informe = sanitize_text(informe)

    # --- Paso 3: Insertar el informe en la hoja de datos sin dañar las fórmulas ---
    try:
        wb = load_workbook(file_path)  # Cargar en modo normal para preservar fórmulas
        ws = wb[HOJA_DATOS]
    except Exception as e:
        print(f"Error al cargar el libro en modo normal: {e}")
        return False

    # Desfusionar el rango B115:L116 (si está fusionado) para evitar errores de escritura
    try:
        ws.unmerge_cells("B115:L116")
    except Exception as e:
        print(f"No se pudo desfusionar el rango B115:L116: {e}")

    # Dividir el informe en líneas y escribirlas desde la celda B115
    lines = informe.split('\n')
    start_row = 118  # Comenzar en la fila 118, columna B (columna 2)
    
    for i, line in enumerate(lines):
        cell = ws.cell(row=start_row + i, column=2)  # Columna B es la número 2
        cell.value = line
        cell.alignment = Alignment(vertical="top")
        cell.font = Font(name="Calibri", size=11)
        # Ajustar la altura de la fila (puedes modificar la altura según sea necesario)
        ws.row_dimensions[start_row + i].height = 15

    try:
        wb.save(file_path)
        print(f"Evaluación completada en: {file_path}")
    except Exception as e:
        print(f"Error al guardar el archivo {file_path}: {e}")
        wb.close()
        return False

    wb.close()
    return True

def seleccionar_carpeta():
    """
    Abre un diálogo para seleccionar una carpeta y retorna su ruta.

    Retorna:
        str: Ruta de la carpeta seleccionada.
    """
    root = tk.Tk()
    root.withdraw()
    folder = askdirectory(title="Seleccione la carpeta con los archivos Excel para el analisis de Repetibilidad y Reproducibilidad")
    root.destroy()
    return folder

def procesar_archivos_en_carpeta(carpeta):
    """
    Procesa todos los archivos Excel en la carpeta especificada.

    Parámetros:
        carpeta (str): Ruta de la carpeta que contiene los archivos Excel.
    """
    extensiones = ('*.xlsx', '*.xlsm', '*.xls')
    archivos = []
    for ext in extensiones:
        archivos.extend(glob.glob(os.path.join(carpeta, ext)))
    
    if not archivos:
        messagebox.showinfo("Información", "No se encontraron archivos de Excel en la carpeta seleccionada.")
        return

    errores = []
    for file in archivos:
        print(f"Procesando archivo: {file}")
        if not generar_informe(file):
            errores.append(file)
    
    if errores:
        messagebox.showwarning("Advertencia", "No se procesaron correctamente los siguientes archivos:\n" + "\n".join(errores))
    else:
        messagebox.showinfo("Completado", "La evaluación se completó en todos los archivos.")

def crear_interfaz():
    """
    Crea la interfaz gráfica para seleccionar la carpeta que contiene los archivos Excel
    y ejecutar la evaluación de repetibilidad y reproducibilidad.
    """
    root = tk.Tk()
    root.title("Evaluación ISO 5725: Repetibilidad y Reproducibilidad")
    root.resizable(False, False)

    frame = tk.LabelFrame(root, text="Evaluación de Archivos de Excel", padx=10, pady=10)
    frame.grid(row=0, column=0, padx=10, pady=10)

    tk.Label(frame, text="Carpeta de Excel:").grid(row=0, column=0, padx=5, pady=5)
    entry_carpeta = tk.Entry(frame, width=50)
    entry_carpeta.grid(row=0, column=1, padx=5, pady=5)
    tk.Button(frame, text="Seleccionar", command=lambda: entry_carpeta.insert(0, seleccionar_carpeta())).grid(row=0, column=2, padx=5, pady=5)

    def iniciar_evaluacion():
        carpeta = entry_carpeta.get().strip()
        if not carpeta:
            messagebox.showerror("Error", "Por favor, seleccione una carpeta.")
            return
        procesar_archivos_en_carpeta(carpeta)

    tk.Button(frame, text="Evaluar Archivos", command=iniciar_evaluacion).grid(row=1, column=0, columnspan=3, pady=10)
    root.mainloop()

def main():
    """
    Función principal que solicita al usuario seleccionar una carpeta y procesa
    todos los archivos Excel válidos en ella.
    """
    folder = seleccionar_carpeta()
    if not folder:
        print("No se seleccionó ninguna carpeta. Saliendo del programa.")
        return
    procesar_archivos_en_carpeta(folder)

if __name__ == "__main__":
    main()
