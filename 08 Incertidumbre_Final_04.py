"""
Aplicación para el Estudio de Incertidumbre en Archivos Excel
--------------------------------------------------------------

Descripción:
    Esta aplicación procesa archivos Excel que contengan hojas específicas para generar
    un estudio de incertidumbre conforme a la norma ISO 21748. Se buscan las siguientes hojas:
        - "sr ysR(Método) ISO 5725": Contiene datos y fórmulas relacionadas con las mediciones.
        - "Estudio Veracidad": Contiene valores de distribuciones y otros parámetros.

    Si se encuentran las hojas requeridas, se elimina de forma segura la hoja
    "Estudio Incertidumbre" (si ya existe) y se crea una nueva que incorpora el análisis,
    dividido en las siguientes secciones:
        A) Incertidumbre y veracidad por técnico.
        B) Estimación de la incertidumbre.
        C) Procedimiento de comparación.
        D) Descripción, hipótesis, evaluación y conclusión.

    El resultado se guarda en el mismo archivo, actualizando el mismo con la nueva hoja.
    
Uso:
    1. Ejecutar la aplicación.
    2. Seleccionar la carpeta que contiene los archivos Excel a procesar.
    3. Presionar el botón "Procesar Archivos".
    4. El archivo se actualizará con la hoja "Estudio Incertidumbre".
    
Requisitos:
    - Python 3.x
    - Módulos: os, tkinter, openpyxl, math, scipy.stats, tkinter.ttk
    - Los archivos Excel deben incluir las hojas "sr ysR(Método) ISO 5725" y "Estudio Veracidad".
    
Autor:
    Ing. Edgar Colmenarez
    
Fecha:
    2025-02-18
"""

import os
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import math
import scipy.stats as stats
from tkinter import ttk

def set_column_widths(ws):
    """
    Ajusta los anchos de columnas para mejorar la presentación en la hoja.
    
    :param ws: Objeto de la hoja de Excel.
    """
    widths = {
        'A': 3, 'B': 12, 'C': 14, 'D': 14, 'E': 14,
        'F': 14, 'G': 12, 'H': 12, 'I': 12
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def style_header(cell):
    """
    Aplica estilos de encabezado a una celda (fuente, relleno y borde).
    
    :param cell: Objeto de la celda a formatear.
    """
    cell.font = Font(bold=True, color="000000")
    cell.fill = PatternFill("solid", fgColor="B7DEE8")
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_data_cell(cell):
    """
    Aplica un borde fino y alineación centrada a una celda de datos.
    
    :param cell: Objeto de la celda a formatear.
    """
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def crear_tabla(ws, ref, table_name):
    """
    Crea y añade una tabla con estilo al rango indicado en la hoja.
    Si ya existe una tabla con ese nombre, se elimina antes de crear la nueva.
    
    :param ws: Objeto de la hoja de Excel.
    :param ref: Rango de celdas en formato Excel (ej., "B5:I8").
    :param table_name: Nombre que se asignará a la tabla.
    """
    # Filtrar solo objetos que sean instancias de Table
    existing_tables = [t for t in ws._tables if isinstance(t, Table) and t.name == table_name]
    for t in existing_tables:
        ws._tables.remove(t)
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

def procesar_excel(archivo, hoja_origen="sr ysR(Método) ISO 5725", hoja_destino="Estudio Incertidumbre"):
    """
    Procesa un archivo Excel: extrae datos, realiza cálculos y genera la hoja 'Estudio Incertidumbre'
    sin alterar las fórmulas en las hojas originales. El resultado se guarda en el mismo archivo.
    
    :param archivo: Ruta completa del archivo Excel.
    :param hoja_origen: Nombre de la hoja con datos y fórmulas.
    :param hoja_destino: Nombre de la hoja de salida.
    :return: Mensaje indicando el éxito del proceso o el error encontrado.
    """
    try:
        # Cargar el libro original (con fórmulas intactas) y la versión data_only para obtener los valores calculados.
        wb = load_workbook(archivo)
        wb_data = load_workbook(archivo, data_only=True)
    except Exception as e:
        return f"Error al cargar '{archivo}': {e}"

    # Si ya existe la hoja de salida, se elimina
    if hoja_destino in wb.sheetnames:
        # Cambiar la hoja activa si es la de salida para evitar conflictos
        if wb.active.title == hoja_destino:
            for nombre in wb.sheetnames:
                if nombre != hoja_destino:
                    wb.active = wb[nombre]
                    break
        del wb[hoja_destino]
    
    # Recargar en modo data_only para actualizar valores (si fuera necesario)
    try:
        wb_data = load_workbook(archivo, data_only=True)
    except Exception as e:
        return f"Error al recargar '{archivo}' en modo data_only: {e}"
    
    # Verificar que existan las hojas requeridas
    if hoja_origen not in wb.sheetnames:
        return f"La hoja '{hoja_origen}' no existe en '{archivo}'."
    if "Estudio Veracidad" not in wb.sheetnames:
        return f"La hoja 'Estudio Veracidad' no existe en '{archivo}'."
    
    # Obtener las hojas necesarias
    hoja_sr = wb[hoja_origen]           # Con fórmulas intactas
    hoja_sr_data = wb_data[hoja_origen]   # Con valores calculados
    hoja_veracidad = wb["Estudio Veracidad"]
    
    # Extraer tres niveles (celdas G10, H10, I10) de la hoja de origen
    niveles = [
        hoja_sr["G10"].value,
        hoja_sr["H10"].value,
        hoja_sr["I10"].value
    ]
    
    # Crear la nueva hoja de salida
    ws = wb.create_sheet(hoja_destino)
    set_column_widths(ws)
    
    # --- TÍTULO PRINCIPAL ---
    ws.merge_cells("A1:I1")
    ws["A1"] = "Estudio de la Incertidumbre según ISO 21748"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, size=16, color="000080")
    
    # ========================
    # SECCIÓN A: INCERTIDUMBRE – VERACIDAD POR TÉCNICO
    # ========================
    headers_A = [
        ("Niveles", "B5"), ("Técnico 1", "C5"), ("Técnico 2", "D5"),
        ("Técnico 3", "E5"), ("Técnico 4", "F5"),
        ("Umax", "G5"), ("Umin", "H5"), ("%U≤ 5 %", "I5")
    ]
    for texto, celda in headers_A:
        ws[celda] = str(texto)
        style_header(ws[celda])
    
    for i, nivel in enumerate(niveles):
        cell = ws[f"B{6+i}"]
        cell.value = nivel
        style_data_cell(cell)
    
    # Mapear datos desde "Estudio Veracidad"
    distribuciones = {
        "C6": "M4", "C7": "M5", "C8": "M6",
        "D6": "M7", "D7": "M8", "D8": "M9",
        "E6": "M10", "E7": "M11", "E8": "M12",
        "F6": "M13", "F7": "M14", "F8": "M15"
    }
    for celda_dest, celda_origen in distribuciones.items():
        ws[celda_dest] = hoja_veracidad[celda_origen].value
        style_data_cell(ws[celda_dest])
    
    # Para cada fila (6 a 8), calcular Umax, Umin y establecer la condición
    for fila in range(6, 9):
        valores = []
        for col in ["C", "D", "E"]:
            val = ws[f"{col}{fila}"].value
            if val is not None:
                valores.append(val)
        Umax = max(valores) if valores else None
        cell_Umax = ws[f"G{fila}"]
        cell_Umax.value = Umax
        if Umax is not None:
            cell_Umax.number_format = "0.00E+0"
        style_data_cell(cell_Umax)
    
        Umin = min(valores) if valores else None
        cell_Umin = ws[f"H{fila}"]
        cell_Umin.value = Umin
        if Umin is not None:
            cell_Umin.number_format = "0.00E+0"
        style_data_cell(cell_Umin)
    
        cell_cond = ws[f"I{fila}"]
        cell_cond.value = "Cumple" if (Umax is not None and Umax <= 5) else "No Cumple"
        style_data_cell(cell_cond)
    
    crear_tabla(ws, "B5:I8", "Tabla_Incertidumbre_Veracidad")
    
    # ========================
    # SECCIÓN B: ESTIMACIÓN DE INCERTIDUMBRE
    # ========================
    ws.merge_cells("B10:I11")
    ws["B10"] = ("Estimación de incertidumbre usando estimados de reproducibilidad y veracidad "
                 "(apartado 6, GTC 142 (ISO 21748))")
    ws["B10"].alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
    ws["B10"].font = Font(bold=True, color="000080")
    
    headers_B = [("Valor evaluado", "C13"), ("Uref", "D13"),
                 ("Sr", "E13"), ("Umed", "F13"), ("U(y)", "G13")]
    for texto, celda in headers_B:
        ws[celda] = str(texto)
        style_header(ws[celda])
    
    # Extraer valores numéricos de la hoja en modo data_only
    sr_values = [
        hoja_sr_data["F69"].value,
        hoja_sr_data["G69"].value,
        hoja_sr_data["H69"].value
    ]
    
    for i in range(3):
        fila_tabla = 14 + i
        cell = ws[f"C{fila_tabla}"]
        cell.value = ws[f"B{6+i}"].value
        style_data_cell(cell)
    
        sr = sr_values[i]
        cell_sr = ws[f"E{fila_tabla}"]
        cell_sr.value = sr
        if sr is not None:
            cell_sr.number_format = "0.00E+0"
        style_data_cell(cell_sr)
    
        try:
            sr_value = float(sr)
        except (ValueError, TypeError):
            sr_value = None
        uref = sr_value / math.sqrt(10) if sr_value is not None else None
        cell_uref = ws[f"D{fila_tabla}"]
        cell_uref.value = uref
        if uref is not None:
            cell_uref.number_format = "0.00E+0"
        style_data_cell(cell_uref)
    
        cell_u_med = ws[f"F{fila_tabla}"]
        actual_umax = ws[f"G{6+i}"].value
        cell_u_med.value = actual_umax if actual_umax is not None else None
        if actual_umax is not None:
            cell_u_med.number_format = "0.00E+0"
        style_data_cell(cell_u_med)
    
        if uref is not None and sr is not None:
            u_y = math.sqrt(uref**2 + sr_value**2)
        else:
            u_y = None
        cell_u_y = ws[f"G{fila_tabla}"]
        cell_u_y.value = u_y
        if u_y is not None:
            cell_u_y.number_format = "0.00E+0"
        style_data_cell(cell_u_y)
    
    crear_tabla(ws, "C13:G16", "Tabla_Estimacion_Incertidumbre")
    
    # ========================
    # SECCIÓN C: PROCEDIMIENTO DE COMPARACIÓN
    # ========================
    ws.merge_cells("B18:I19")
    ws["B18"] = ("Procedimiento de comparación (apartado 14,2, GTC 142 (ISO 21748))")
    ws["B18"].alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
    ws["B18"].font = Font(bold=True, color="000080")
    
    headers_C = [("U1", "C22"), ("U2", "D22"), ("F", "F22"), ("Fcrítico", "G22")]
    for texto, celda in headers_C:
        ws[celda] = str(texto)
        style_header(ws[celda])
    
    f_critico = stats.f.ppf(0.95, 10, 10)
    
    for i in range(3):
        fila_origen = 14 + i
        fila_comp = 23 + i
        u1 = ws[f"G{fila_origen}"].value  # U(y) calculado
        u2 = ws[f"D{fila_origen}"].value  # Uref
        F_val = (u2 / u1) ** 2 if u1 not in (None, 0) and u2 is not None else None
    
        cell_u1 = ws[f"C{fila_comp}"]
        cell_u1.value = u1
        if u1 is not None:
            cell_u1.number_format = "0.00E+0"
        style_data_cell(cell_u1)
    
        cell_u2 = ws[f"D{fila_comp}"]
        cell_u2.value = u2
        if u2 is not None:
            cell_u2.number_format = "0.00E+0"
        style_data_cell(cell_u2)
    
        cell_F = ws[f"F{fila_comp}"]
        cell_F.value = F_val
        if F_val is not None:
            cell_F.number_format = "0.0000E+0"
        style_data_cell(cell_F)
    
        cell_fcrit = ws[f"G{fila_comp}"]
        cell_fcrit.value = f_critico
        if f_critico is not None:
            cell_fcrit.number_format = "0.0000"
        style_data_cell(cell_fcrit)
    
    crear_tabla(ws, "C22:D25", "Tabla_Comparacion1")
    crear_tabla(ws, "F22:G25", "Tabla_Comparacion2")
    
    # ========================
    # SECCIÓN D: DESCRIPCIÓN, HIPÓTESIS, EVALUACIÓN Y CONCLUSIÓN
    # ========================
    method_description = (
        "Descripción del método:\n"
        "Se realizó el estudio de la incertidumbre conforme a ISO 21748, combinando estimados de reproducibilidad "
        "y veracidad a partir de mediciones en distintos niveles. Los parámetros Umax, Umin, Uref y U(y) se calcularon "
        "para evaluar la consistencia y el desempeño del método."
    )
    hypotheses = (
        "Hipótesis:\n"
        "H0: La incertidumbre combinada no es significativamente mayor (F ≤ Fcrítico), lo que implica que el método cumple con los criterios establecidos.\n"
        "H1: La incertidumbre combinada es significativamente mayor (F > Fcrítico), lo que indica la necesidad de revisar el proceso de medición."
    )
    evaluation = (
        "Evaluación:\n"
        "Se calcularon Uref y U(y) para cada nivel y se comparó la razón F = (Uref/U(y))² con el valor crítico obtenido "
        "(n=10, 95% de confianza) para determinar diferencias significativas."
    )
    evaluacion_detallada = ""
    for i in range(3):
        nivel_num = i + 1
        fila_comp = 23 + i
        fcal = ws[f"F{fila_comp}"].value
        fcrit = ws[f"G{fila_comp}"].value
        if fcal is not None and fcrit is not None:
            resultado = "Cumple" if fcal <= fcrit else "No Cumple"
            evaluacion_detallada += f"Nivel {nivel_num}: Fcal ({fcal:.4g}) vs Fcrítico ({fcrit:.4g}) => {resultado}\n"
        else:
            evaluacion_detallada += f"Nivel {nivel_num}: Datos insuficientes para evaluación.\n"
    conclusion = (
        "Conclusión:\n"
        "Si Fcal es mayor que Fcrítico, la incertidumbre combinada se considera significativamente mayor, lo que indica la necesidad "
        "de revisar el proceso de medición. De lo contrario, el método cumple con los criterios establecidos.\n\n"
        "Evaluación detallada:\n" + evaluacion_detallada
    )
    resumen_text = "\n\n".join([method_description, hypotheses, evaluation, conclusion])
    ws.merge_cells("B27:I51")
    ws["B27"] = resumen_text
    ws["B27"].alignment = Alignment(wrapText=True, vertical="top")
    ws["B27"].font = Font(italic=True, size=11)
    
    try:
        wb.save(archivo)
        return f"Procesado: {archivo} (actualizado con la hoja '{hoja_destino}')"
    except Exception as e:
        return f"Error al guardar '{archivo}': {e}"

def seleccionar_carpeta():
    """
    Abre una ventana de diálogo para seleccionar la carpeta que contiene los archivos Excel.
    """
    carpeta = filedialog.askdirectory(title="Seleccione la carpeta con los archivos Excel de Validaciones")
    if carpeta:
        entry_ruta.delete(0, tk.END)
        entry_ruta.insert(0, carpeta)

def ejecutar_proceso():
    """
    Ejecuta el procesamiento de todos los archivos Excel en la carpeta seleccionada.
    """
    carpeta = entry_ruta.get()
    if not os.path.isdir(carpeta):
        messagebox.showerror("Error", "Seleccione una carpeta válida.")
        return

    archivos = [
        os.path.join(carpeta, f)
        for f in os.listdir(carpeta)
        if f.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm"))
    ]

    if not archivos:
        messagebox.showinfo("Información", "No hay archivos Excel válidos en la carpeta seleccionada.")
        return

    progreso['maximum'] = len(archivos)

    for i, archivo in enumerate(archivos, 1):
        resultado = procesar_excel(archivo)
        listbox.insert(tk.END, resultado)
        progreso['value'] = i
        ventana.update_idletasks()

    messagebox.showinfo("Completado", "Procesamiento finalizado.")

# --- Interfaz Gráfica ---
ventana = tk.Tk()
ventana.title("Estudio de Incertidumbre (ISO 21748)")
ventana.geometry("700x500")
ventana.configure(bg="#f4f4f4")

label_instruccion = tk.Label(
    ventana,
    text="Seleccione la carpeta con los archivos Excel:",
    bg="#f4f4f4",
    font=("Arial", 10, "bold")
)
label_instruccion.pack(pady=5)

frame = tk.Frame(ventana, bg="#f4f4f4")
frame.pack()

entry_ruta = tk.Entry(frame, width=60)
entry_ruta.pack(side=tk.LEFT, padx=5)

boton_examinar = tk.Button(
    frame,
    text="Examinar",
    command=seleccionar_carpeta,
    bg="#2196F3",
    fg="white",
    relief="raised"
)
boton_examinar.pack(side=tk.LEFT)

boton_procesar = tk.Button(
    ventana,
    text="Procesar Archivos",
    command=ejecutar_proceso,
    bg="#4CAF50",
    fg="white",
    font=("Arial", 10, "bold"),
    relief="raised"
)
boton_procesar.pack(pady=10)

progreso = ttk.Progressbar(ventana, length=600, mode='determinate')
progreso.pack(pady=5)

listbox = tk.Listbox(ventana, width=90, height=10)
listbox.pack(pady=5)

ventana.mainloop()
