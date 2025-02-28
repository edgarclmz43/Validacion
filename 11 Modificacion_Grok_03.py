import os
from openpyxl import load_workbook
from tkinter import Tk, filedialog
from unicodedata import normalize

def normalizar(texto):
    """Normaliza una cadena a NFC y elimina espacios extra al inicio y final."""
    return normalize("NFC", texto).strip()

def get_sheet_by_name(wb, target_name):
    """
    Busca una hoja en el libro wb cuyo nombre normalizado coincida con target_name.
    Retorna el objeto hoja o None si no se encuentra.
    """
    normalized_target = normalizar(target_name)
    for sheet in wb.worksheets:
        if normalizar(sheet.title) == normalized_target:
            return sheet
    return None

def procesar_archivo(ruta_archivo, wb_plantilla):
    wb = load_workbook(ruta_archivo)
    # Buscar las hojas existentes en el archivo destino
    hoja_plan_origen = get_sheet_by_name(wb, "Plan de validación")
    hoja_informe_origen = get_sheet_by_name(wb, "Informe de Validación")
    
    if hoja_plan_origen and hoja_informe_origen:
        # Copiar datos de las celdas J8:AE14 de la hoja Plan de validacion
        datos_extraidos = {}
        for fila in range(8, 15):      # Filas 8 a 14
            for col in range(10, 32):  # Columnas J (10) a AE (31)
                datos_extraidos[(fila, col)] = hoja_plan_origen.cell(row=fila, column=col).value
        
        # Eliminar las hojas existentes en destino
        wb.remove(hoja_plan_origen)
        wb.remove(hoja_informe_origen)
        
        # Obtener las hojas de la plantilla (se asume que están correctamente nombradas)
        hoja_plantilla_plan = get_sheet_by_name(wb_plantilla, "Plan de validación")
        hoja_plantilla_informe = get_sheet_by_name(wb_plantilla, "Informe de Validación")
        
        if not hoja_plantilla_plan or not hoja_plantilla_informe:
            print("Error: No se encontraron las hojas requeridas en la plantilla.")
            return

        # --- AGREGAR HOJAS DE LA PLANTILLA AL DESTINO USANDO copy_worksheet ---
        # copy_worksheet solo funciona con hojas del mismo libro, por lo que se agrega temporalmente
        # la hoja de la plantilla al workbook destino. Se modifica el atributo _parent (no documentado)
        # para que la hoja pertenezca a wb.
        hoja_plantilla_plan._parent = wb
        hoja_plantilla_informe._parent = wb
        wb._sheets.append(hoja_plantilla_plan)
        wb._sheets.append(hoja_plantilla_informe)
        
        # Usar copy_worksheet para copiar las hojas ya incorporadas
        nueva_plan = wb.copy_worksheet(hoja_plantilla_plan)
        nueva_informe = wb.copy_worksheet(hoja_plantilla_informe)
        
        # Renombrar las hojas copiadas
        nueva_plan.title = "Plan de validación"
        nueva_informe.title = "Informe de Validación"
        
        # Eliminar las hojas temporales que se agregaron (las originales de la plantilla)
        wb.remove(hoja_plantilla_plan)
        wb.remove(hoja_plantilla_informe)
        # --- FIN DE LA COPIA CON copy_worksheet ---

        # Pegar los datos extraidos en la nueva hoja "Plan de validacion" en las celdas J8:AE14
        for (fila, col), valor in datos_extraidos.items():
            nueva_plan.cell(row=fila, column=col, value=valor)
        
        # Reubicar las hojas para que "Plan de validacion" y "Informe de validacion" queden en la posición 1 y 2
        hojas_restantes = [sheet for sheet in wb._sheets if normalizar(sheet.title) not in 
                           [normalizar("Plan de validación"), normalizar("Informe de Validación")]]
        wb._sheets = [nueva_plan, nueva_informe] + hojas_restantes

        wb.save(ruta_archivo)
        print(f"Procesado: {ruta_archivo}")
    else:
        print(f"Se omite {ruta_archivo} - no se encontraron las hojas requeridas.")
        print("Hojas encontradas:", wb.sheetnames)

def main():
    # Ocultar la ventana principal de Tkinter
    root = Tk()
    root.withdraw()

    # Solicitar al usuario la carpeta con los archivos a modificar
    ruta_carpeta = filedialog.askdirectory(title="Selecciona la carpeta de los archivos a modificar")
    if not ruta_carpeta or not os.path.exists(ruta_carpeta):
        print("La carpeta seleccionada no existe o no se proporcionó una ruta válida. Finalizando.")
        return

    # Solicitar al usuario el archivo de plantilla
    ruta_plantilla = filedialog.askopenfilename(title="Selecciona el archivo de plantilla",
                                                  filetypes=[("Archivos Excel", "*.xlsx")])
    if not ruta_plantilla or not os.path.exists(ruta_plantilla):
        print("El archivo de plantilla no existe o no se proporcionó una ruta válida. Finalizando.")
        return

    # Cargar la plantilla
    wb_plantilla = load_workbook(ruta_plantilla)

    # Listar y ordenar los archivos que inician con un prefijo numérico (ej. "01", "02", ...)
    archivos = [archivo for archivo in os.listdir(ruta_carpeta)
                if archivo.endswith(".xlsx") and archivo[:2].isdigit()]
    archivos.sort()

    # Procesar cada archivo encontrado
    for archivo in archivos:
        ruta_archivo = os.path.join(ruta_carpeta, archivo)
        procesar_archivo(ruta_archivo, wb_plantilla)

if __name__ == "__main__":
    main()
