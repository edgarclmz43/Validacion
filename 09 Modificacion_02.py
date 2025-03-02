import os
from openpyxl import load_workbook
from tkinter import Tk, filedialog
from unicodedata import normalize
from copy import copy

def normalizar(texto):
    """Normaliza una cadena a NFC y elimina espacios extra al inicio y final."""
    return normalize("NFC", texto).strip()

def get_sheet_by_name(wb, target_name):
    """
    Busca una hoja en el libro wb cuyo nombre normalizado coincida con target_name.
    Retorna el objeto hoja o None si no se encuentra.
    """
    normalized_target = normalizar(target_name)
    for sheet in wb.worksheets:
        if normalizar(sheet.title) == normalized_target:
            return sheet
    return None

def copiar_hoja(hoja_origen, wb_destino, nuevo_nombre):
    """
    Copia el contenido de hoja_origen a una nueva hoja en wb_destino con el nombre nuevo_nombre,
    manteniendo valores, fórmulas y el formato completo (estilos, dimensiones, celdas fusionadas,
    congelación de paneles y área de impresión).
    """
    hoja_dest = wb_destino.create_sheet(title=nuevo_nombre)
    
    # Copiar cada celda: valores y estilos
    for row in hoja_origen.iter_rows():
        for celda in row:
            nueva_celda = hoja_dest.cell(row=celda.row, column=celda.column)
            nueva_celda.value = celda.value
            if celda.has_style:
                nueva_celda.font = copy(celda.font)
                nueva_celda.border = copy(celda.border)
                nueva_celda.fill = copy(celda.fill)
                nueva_celda.number_format = celda.number_format
                nueva_celda.protection = copy(celda.protection)
                nueva_celda.alignment = copy(celda.alignment)
    
    # Copiar dimensiones de columnas y filas
    for col, dim in hoja_origen.column_dimensions.items():
        hoja_dest.column_dimensions[col] = copy(dim)
    for row, dim in hoja_origen.row_dimensions.items():
        hoja_dest.row_dimensions[row] = copy(dim)
    
    # Copiar celdas fusionadas
    if hoja_origen.merged_cells.ranges:
        for merged_range in hoja_origen.merged_cells.ranges:
            hoja_dest.merge_cells(str(merged_range))
    
    # Copiar congelación de paneles y área de impresión
    hoja_dest.freeze_panes = hoja_origen.freeze_panes
    hoja_dest.print_area = hoja_origen.print_area

    return hoja_dest

def procesar_archivo(ruta_archivo, wb_plantilla):
    wb = load_workbook(ruta_archivo)
    # Buscar las hojas de origen usando la función de búsqueda normalizada
    hoja_plan_origen = get_sheet_by_name(wb, "Plan de validación")
    hoja_informe_origen = get_sheet_by_name(wb, "Informe de Validación")
    
    if hoja_plan_origen and hoja_informe_origen:
        # Extraer datos de la hoja "Plan de validación" en el rango J8:AE14
        datos_extraidos = {}
        for fila in range(8, 15):  # Filas 8 a 14
            for col in range(10, 32):  # Columnas J (10) a AE (31)
                datos_extraidos[(fila, col)] = hoja_plan_origen.cell(row=fila, column=col).value

        # Eliminar las hojas originales
        wb.remove(hoja_plan_origen)
        wb.remove(hoja_informe_origen)

        # Obtener las hojas de la plantilla de forma normalizada
        hoja_plantilla_plan = get_sheet_by_name(wb_plantilla, "Plan de validación")
        hoja_plantilla_informe = get_sheet_by_name(wb_plantilla, "Informe de Validación")
        
        if not hoja_plantilla_plan or not hoja_plantilla_informe:
            print("Error: No se encontraron las hojas requeridas en la plantilla.")
            return

        # Copiar las hojas de la plantilla al libro actual
        nueva_plan = copiar_hoja(hoja_plantilla_plan, wb, "Plan de validación")
        nueva_informe = copiar_hoja(hoja_plantilla_informe, wb, "Informe de Validación")

        # Pegar los datos extraídos en la nueva hoja "Plan de validación"
        for (fila, col), valor in datos_extraidos.items():
            # Solo se sobreescribe el valor, conservando el formato de la plantilla
            nueva_plan.cell(row=fila, column=col, value=valor)

        # Reubicar las hojas para que "Plan de validación" y "Informe de Validación" queden en la posición 1 y 2
        hojas_restantes = [sheet for sheet in wb._sheets if normalizar(sheet.title) not in 
                           [normalizar("Plan de validación"), normalizar("Informe de Validación")]]
        wb._sheets = [nueva_plan, nueva_informe] + hojas_restantes

        wb.save(ruta_archivo)
        print(f"Procesado: {ruta_archivo}")
    else:
        print(f"Se omite {ruta_archivo} - no se encontraron las hojas requeridas.")
        print("Hojas encontradas:", wb.sheetnames)

def main():
    # Ocultar la ventana principal de Tkinter
    root = Tk()
    root.withdraw()

    # Solicitar al usuario la carpeta con los archivos a modificar
    ruta_carpeta = filedialog.askdirectory(title="Selecciona la carpeta de los archivos a modificar")
    if not ruta_carpeta or not os.path.exists(ruta_carpeta):
        print("La carpeta seleccionada no existe o no se proporcionó una ruta válida. Finalizando.")
        return

    # Solicitar al usuario el archivo de plantilla
    ruta_plantilla = filedialog.askopenfilename(title="Selecciona el archivo de plantilla",
                                                  filetypes=[("Archivos Excel", "*.xlsx")])
    if not ruta_plantilla or not os.path.exists(ruta_plantilla):
        print("El archivo de plantilla no existe o no se proporcionó una ruta válida. Finalizando.")
        return

    # Cargar la plantilla
    wb_plantilla = load_workbook(ruta_plantilla)

    # Listar y ordenar los archivos que inician con un prefijo numérico (ej. "01", "02", ...)
    archivos = [archivo for archivo in os.listdir(ruta_carpeta)
                if archivo.endswith(".xlsx") and archivo[:2].isdigit()]
    archivos.sort()

    # Procesar cada archivo encontrado
    for archivo in archivos:
        ruta_archivo = os.path.join(ruta_carpeta, archivo)
        procesar_archivo(ruta_archivo, wb_plantilla)

if __name__ == "__main__":
    main()
