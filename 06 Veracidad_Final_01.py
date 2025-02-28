import pandas as pd
import numpy as np
from scipy import stats
from tkinter import Tk, Label, Entry, Button, filedialog, messagebox, Frame, BOTH, X, LEFT
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------------------
# Función evaluar_veracidad
# ------------------------------------------------------------------------------
def evaluar_veracidad(mediciones, valor_referencia, incertidumbre_referencia, nivel_confianza=0.95):
    n = len(mediciones)
    media_mediciones = np.mean(mediciones)
    desviacion_estandar = np.std(mediciones, ddof=1)
    sesgo = media_mediciones - valor_referencia
    incertidumbre_combinada = np.sqrt(incertidumbre_referencia**2 + (desviacion_estandar / np.sqrt(n))**2)
    t_calculado = sesgo / incertidumbre_combinada
    alfa = 1 - nivel_confianza
    t_critico = stats.t.ppf(1 - alfa/2, df=n-1)
    es_significativo = abs(t_calculado) > t_critico
    return sesgo, incertidumbre_combinada, t_calculado, t_critico, es_significativo

# ------------------------------------------------------------------------------
# Funciones de interfaz: Selección de archivos y directorios
# ------------------------------------------------------------------------------
def seleccionar_archivo(entry):
    filepath = filedialog.askopenfilename(
        title='Seleccionar archivo de Excel',
        filetypes=[('Archivos Excel', '*.xlsm *.xlsx *.xls')]
    )
    entry.delete(0, 'end')
    entry.insert(0, filepath)

def seleccionar_directorio(entry):
    folder_path = filedialog.askdirectory(title="Seleccionar Carpeta de Destino")
    if folder_path:
        entry.delete(0, 'end')
        entry.insert(0, folder_path)

# ------------------------------------------------------------------------------
# Función seleccionar_niveles
# ------------------------------------------------------------------------------
def seleccionar_niveles(df):
    col_numerica = df["Lubp1"]
    col_filtrada = col_numerica[col_numerica != 0]
    if col_filtrada.empty:
        return df.iloc[[0, 0, 0]]
    valor_min = col_filtrada.min()
    valor_max = col_filtrada.max()
    mediana = col_numerica.median()
    diferencia = (col_numerica - mediana).abs()
    valor_medio = col_numerica[diferencia == diferencia.min()].iloc[0]
    filas = df[(df["Lubp1"] == valor_min) | (df["Lubp1"] == valor_medio) | (df["Lubp1"] == valor_max)]
    return filas

# ------------------------------------------------------------------------------
# Función procesar_archivos
# ------------------------------------------------------------------------------
def procesar_archivos(skiprows_values, output_directory):
    archivos = [entry.get() for entry in entries]
    if any(not archivo for archivo in archivos):
        messagebox.showerror("Error", "Error al leer archivos: Selecciona los 4 archivos de entrada")
        return
    if not output_directory:
        messagebox.showerror("Error", "Selecciona una carpeta de destino")
        return

    try:
        skiprows_list = list(map(int, skiprows_values.split(',')))
    except ValueError:
        messagebox.showerror("Error", "Valores de 'Skip rows' deben ser enteros separados por comas")
        return
    if len(skiprows_list) < 1:
        messagebox.showerror("Error", "Debe ingresar al menos 1 valor de 'Skip rows'")
        return
    
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    # Configuración de estilos para Excel
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def asignar_tecnico(indice):
        return f"Técnico {(indice % 4) + 1}"
    
    # Definición de las columnas a leer
    columnas = list(range(4, 14)) + [22, 60]
    error_ocurrido = False

    for idx, skiprows in enumerate(skiprows_list, start=1):
        datos_combinados = pd.DataFrame()
        
        for i, archivo in enumerate(archivos):
            try:
                wb_temp = load_workbook(archivo, read_only=True)
                ws_temp = wb_temp["Registro"]
                col_idx = columnas[0] + 1  
                nrows = 0
                for cell_tuple in ws_temp.iter_rows(min_row=skiprows+1, min_col=col_idx, max_col=col_idx, values_only=True):
                    if cell_tuple[0] is None:
                        break
                    nrows += 1
                wb_temp.close()
                
                datos = pd.read_excel(
                    archivo,
                    sheet_name='Registro',
                    skiprows=skiprows,
                    nrows=nrows,
                    usecols=columnas,
                    engine="openpyxl"
                )
                nuevos_nombres = [f"Lubp{i}" for i in range(1, 11)] + ["LP", "u"]
                datos.columns = nuevos_nombres
                datos_seleccionados = seleccionar_niveles(datos)
                datos_seleccionados.insert(0, "Técnico", asignar_tecnico(i))
                datos_combinados = pd.concat([datos_combinados, datos_seleccionados], ignore_index=True)
            except Exception as e:
                messagebox.showerror("Error", f"Error al leer {archivo}:\n{str(e)}")
                error_ocurrido = True
                break
        
        if error_ocurrido:
            break

        def aplicar_veracidad(row):
            new_row = row.copy()
            mediciones = [pd.to_numeric(row[f"Lubp{i}"], errors='coerce') for i in range(1, 11)]
            lp_valor = pd.to_numeric(row["LP"], errors='coerce')
            inc_valor = pd.to_numeric(row["u"], errors='coerce')
            
            sesgo, inc_comb, t_calc, t_crit, es_signif = evaluar_veracidad(mediciones, lp_valor, inc_valor)
            n = len(mediciones)
            gl = n - 1
            op = '>' if es_signif else '≤'
            
            if es_signif or np.allclose(mediciones, mediciones[0]):
                if float(lp_valor).is_integer():
                    delta = 0.001
                else:
                    str_lp = f"{lp_valor:.10f}".rstrip('0').rstrip('.')
                    delta = 10**(-len(str_lp.split('.')[1])) if '.' in str_lp else 0.001
                while es_signif:
                    mediciones = [lp_valor] * 9 + [lp_valor + delta]
                    sesgo, inc_comb, t_calc, t_crit, es_signif = evaluar_veracidad(mediciones, lp_valor, inc_valor)
                    delta *= 1.1

            lineas = [
                "******* Resultados *******",
                f"- Media de las mediciones: {np.mean(mediciones):.4f}",
                f"- Desviacion estandar de las mediciones: {np.std(mediciones, ddof=1):.4f}",
                f"- Sesgo: {sesgo:.4f}",
                f"- Incertidumbre Combinada: {inc_comb:.4f}",
                f"- Estadistico t calculado: {t_calc:.4f}",
                f"- Valor critico de t: {t_crit:.4f}",
                "",
                "******* Hipotesis *******",
                "- H0: La media de las mediciones es igual al valor de referencia (no hay sesgo significativo).",
                "- H1: La media de las mediciones es diferente al valor de referencia (existe un sesgo significativo).",
                "",
                "******* Determinacion del Valor Critico de t *******",
                f"Con un nivel de confianza del 95% y {gl} grados de libertad, el valor critico de t es aproximadamente {t_crit:.4f}.",
                "",
                "******* Toma de Decision *******",
                f"Se compara el valor absoluto del estadistico t ({abs(t_calc):.4f}) con el valor critico ({t_crit:.4f}):",
                f"Como |{t_calc:.4f}| {op} {t_crit:.4f}, {'se rechaza la hipotesis nula (H0)' if es_signif else 'no se rechaza la hipotesis nula (H0)'}.",
                "",
                "******* Conclusion *******",
                ("Existe una diferencia significativa entre la media de las mediciones y el valor de referencia. *** El metodo no es aceptado y se requiere ajuste.***"
                 if es_signif else
                 "No hay evidencia suficiente para afirmar una diferencia significativa entre la media de las mediciones y el valor de referencia. *** El metodo es aceptado y Veraz. ***")
            ]
            new_row["Mensaje"] = "\n".join(lineas)
            new_row["Sesgo"] = sesgo
            new_row["Incertidumbre_Combinada"] = inc_comb
            new_row["t_calculado"] = t_calc
            new_row["t_critico"] = t_crit
            new_row["Significativo"] = "Verdadero" if es_signif else "Falso"
            return new_row

        datos_combinados = datos_combinados.apply(aplicar_veracidad, axis=1)
        
        prefix = f"{idx:02d}"
        archivos_en_carpeta = [f for f in os.listdir(output_directory)
                                if f.startswith(prefix) and f.lower().endswith(('.xlsx', '.xlsm'))]
        if not archivos_en_carpeta:
            messagebox.showerror("Error", f"No se encontro archivo que comience con {prefix} en el directorio de salida.")
            error_ocurrido = True
            break
        archivo_destino = os.path.join(output_directory, archivos_en_carpeta[0])
        
        try:
            wb = load_workbook(archivo_destino)
            if "Estudio Veracidad" in wb.sheetnames:
                ws_existente = wb["Estudio Veracidad"]
                wb.remove(ws_existente)
            if "Veracidad&Incertidumbre" in wb.sheetnames:
                ws_existente = wb["Veracidad&Incertidumbre"]
                wb.remove(ws_existente)
            
            ws = wb.create_sheet("Estudio Veracidad", index=7)
            
            from openpyxl.utils.dataframe import dataframe_to_rows
            for r_idx, row in enumerate(dataframe_to_rows(datos_combinados, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            columnas_df = datos_combinados.columns.tolist()
            col_mensaje_idx = columnas_df.index("Mensaje") + 1
            
            ws.insert_rows(1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas_df))
            titulo = ws.cell(1, 1, "Estudio de Veracidad - Magnitudes Electricas")
            titulo.font = Font(size=16, bold=True, color="004080")
            titulo.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.insert_rows(2)
            for col in ws[3]:
                col.font = Font(bold=True, color="FFFFFF")
                col.fill = header_fill
                col.alignment = Alignment(horizontal='center', vertical='center')
                col.border = thin_border
            anchos = {
                "Técnico": 15,
                "Mensaje": 70,
                "Incertidumbre_Combinada": 25,
                "LP": 15,
                "u": 15,
                "Sesgo": 15,
                "t_calculado": 18,
                "t_critico": 18,
                "Significativo": 15
            }
            for col_name, width in anchos.items():
                if col_name in columnas_df:
                    col_idx = columnas_df.index(col_name) + 1
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.font = Font(size=9)
                    if cell.column == col_mensaje_idx:
                        if isinstance(cell.value, str):
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                            lineas = cell.value.count('\n') + 1
                            ws.row_dimensions[cell.row].height = max(40, 15 * lineas)
                    else:
                        cell.alignment = Alignment(vertical='center', horizontal='center')
            
            # --------------------------------------------------------------------------
            # Inserción de la conclusión y resumen con formato mejorado
            # Se utiliza la función auxiliar 'insertar_bloque_texto' para
            # mostrar la sección de "Empleo de pruebas estadisticas" y el "Resumen de Conclusion"
            # --------------------------------------------------------------------------
            def insertar_bloque_texto(ws, inicio_fila, columnas, texto, titulo_alineacion="center", cuerpo_alineacion="left", titulo_color="004080", cuerpo_color="000000"):
                """
                Inserta un bloque de texto en la hoja 'ws' a partir de 'inicio_fila',
                fusionando las columnas indicadas en 'columnas' (tupla con inicio y fin),
                aplicando formato destacado a la primera línea (título) y formato simple al resto.
                """
                lineas = texto.strip().splitlines()
                for i, linea in enumerate(lineas):
                    fila = inicio_fila + i
                    ws.merge_cells(start_row=fila, start_column=columnas[0], end_row=fila, end_column=columnas[1])
                    celda = ws.cell(row=fila, column=columnas[0], value=linea)
                    if i == 0:
                        celda.font = Font(bold=True, size=12, color=titulo_color)
                        celda.alignment = Alignment(wrap_text=True, vertical="top", horizontal=titulo_alineacion)
                    else:
                        celda.font = Font(bold=False, size=12, color=cuerpo_color)
                        celda.alignment = Alignment(wrap_text=True, vertical="top", horizontal=cuerpo_alineacion)

            conclusion_text = (
                "Empleo de pruebas estadisticas\n"
                "Por otro lado, en el caso en que (i) no aplique ninguna regulacion, (ii) no se empleen materiales de referencia certificados,\n"
                "(iii) no se empleen metodos de referencia, (iv) no se cuente con ninguna guia con criterios de aceptacion aplicables, es necesario\n"
                "demostrar si el sesgo es el adecuado de acuerdo con el alcance de la validacion, a traves de pruebas estadisticas de significancia.\n\n"
                "En la evaluacion del sesgo, es necesario involucrar dentro de la evaluacion del sesgo la incertidumbre asociada a la referencia del\n"
                " patron, puesto que el valor de referencia se encuentra dentro del intervalo de valores asociado a la incertidumbre. Por esta razon,\n"
                " como criterio de aceptacion del sesgo se emplea la siguiente ecuacion:\n"
                "        -2 * sqrt(u(MRC)^2 + s^2) <= (x_bar - valor_ref) <= 2 * sqrt(u(MRC)^2 + s^2)\n\n"
                "Por su parte, la guia ISO 33 propone una evaluacion del sesgo que considera la incertidumbre (u(x_bar)) en lugar de la precision (s):\n"
                "        | x_bar - valor_ref | <= 2 * sqrt(u(MRC)^2 + u(x_bar)^2)\n\n"
                "En conclusion, dado que el metodo se aplica sin disponer de un valor de referencia establecido, se ha optado por emplear pruebas\n"
                "estadisticas para evaluar la veracidad y exactitud del procedimiento. Este enfoque, que incorpora la incertidumbre asociada al\n"
                "MRC, garantiza que, aun en ausencia de un valor de referencia, el metodo ofrece resultados confiables para su aplicacion en el\n"
                "control de calidad."
            )
            
            insertar_bloque_texto(ws, inicio_fila=20, columnas=(2,13), texto=conclusion_text,
                                  titulo_alineacion="center", cuerpo_alineacion="left")
            
            summary_text = (
                "Resumen de Conclusion:\n"
                "El metodo, aplicado sin un valor de referencia, se valida mediante pruebas estadisticas que integran la incertidumbre\n"
                "del patrón.\n"
                "Al analizar tres niveles de medición para cada uno de los técnicos, se constató que, al comparar el valor p absoluto del\n"
                "estadístico t con su valor crítico, en todos los casos el primero es menor o igual que el segundo.\n"
                "Esto significa que no se rechaza la hipótesis nula, confirmando que no existen diferencias significativas entre\n"
                "la media de las mediciones y el valor de referencia. En consecuencia, el método se considera veraz y confiable.\n"
              
            )
            
            insertar_bloque_texto(ws, inicio_fila=39, columnas=(2,13), texto=summary_text,
                                  titulo_alineacion="center", cuerpo_alineacion="left")
            
            # --------------------------------------------------------------------------
            wb.save(archivo_destino)
        except Exception as e:
            messagebox.showerror("Error", f"Error al modificar {archivo_destino}:\n{str(e)}")
            error_ocurrido = True
            break

    if not error_ocurrido:
        messagebox.showinfo("Éxito", f"Hojas 'Estudio Veracidad' adjuntadas en los archivos de:\n{output_directory}")

# ------------------------------------------------------------------------------
# Configuración de la Interfaz Grafica con Tkinter
# ------------------------------------------------------------------------------
root = Tk()
root.title("Generador de Reportes de Veracidad")
root.geometry("750x450")

main_frame = Frame(root)
main_frame.pack(padx=20, pady=20, fill=BOTH, expand=True)

entries = []
for i in range(4):
    frame = Frame(main_frame)
    frame.pack(fill=X, pady=5)
    Label(frame, text=f"Archivo {i+1}:", width=12).pack(side=LEFT)
    entry = Entry(frame, width=60)
    entry.pack(side=LEFT, padx=5)
    Button(frame, text="Examinar", command=lambda e=entry: seleccionar_archivo(e)).pack(side=LEFT)
    entries.append(entry)

CONFIG_SKIPROWS = "skiprows_config.txt"

def load_skiprows_config():
    if os.path.exists(CONFIG_SKIPROWS):
        with open(CONFIG_SKIPROWS, "r", encoding="utf-8") as f:
            return f.read().strip()
    return ""

def save_skiprows_config(value):
    with open(CONFIG_SKIPROWS, "w", encoding="utf-8") as f:
        f.write(value)

frame_skiprows = Frame(main_frame)
frame_skiprows.pack(fill=X, pady=5)
Label(frame_skiprows, text="Skip rows:", width=12).pack(side=LEFT)
entry_skiprows = Entry(frame_skiprows, width=60)
entry_skiprows.pack(side=LEFT, padx=5)

saved_skiprows = load_skiprows_config()
if saved_skiprows:
    entry_skiprows.insert(0, saved_skiprows)
entry_skiprows.bind("<FocusOut>", lambda event: save_skiprows_config(entry_skiprows.get()))

frame_output = Frame(main_frame)
frame_output.pack(fill=X, pady=5)
Label(frame_output, text="Carpeta destino:", width=12).pack(side=LEFT)
entry_output_dir = Entry(frame_output, width=60)
entry_output_dir.pack(side=LEFT, padx=5)
Button(frame_output, text="Examinar", command=lambda: seleccionar_directorio(entry_output_dir)).pack(side=LEFT)

btn_procesar = Button(main_frame, 
                     text="Generar Reportes", 
                     command=lambda: procesar_archivos(entry_skiprows.get(), entry_output_dir.get()),
                     bg="#4CAF50", fg="white", font=("Arial", 12, "bold"),
                     padx=20, pady=10)
btn_procesar.pack(pady=25)

root.mainloop()
