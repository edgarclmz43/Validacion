import os
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl

def select_folder():
    """Abre una ventana para seleccionar una carpeta y la muestra en el entry."""
    folder = filedialog.askdirectory()
    if folder:
        folder_entry.delete(0, tk.END)
        folder_entry.insert(0, folder)

def process_files():
    """Recorre los archivos Excel de la carpeta seleccionada y elimina la hoja indicada si existe.
    
    Se cargan los archivos usando data_only=False para preservar fórmulas y, en caso de archivos .xlsm,
    se utiliza keep_vba=True para conservar macros y otros elementos.
    """
    folder = folder_entry.get()
    sheet_name = sheet_entry.get().strip()
    if not folder:
        messagebox.showerror("Error", "Por favor, seleccione una carpeta.")
        return
    if not sheet_name:
        messagebox.showerror("Error", "Por favor, ingrese el nombre de la hoja a borrar.")
        return
    
    log_text.delete("1.0", tk.END)  # Limpia el área de log
    files = os.listdir(folder)
    # Filtra archivos de Excel (incluye .xlsx y .xlsm)
    excel_files = [f for f in files if f.endswith(".xlsx") or f.endswith(".xlsm")]
    
    if not excel_files:
        messagebox.showinfo("Información", "No se encontraron archivos de Excel en la carpeta.")
        return
    
    for file in excel_files:
        file_path = os.path.join(folder, file)
        try:
            # Cargar el libro conservando fórmulas y formatos.
            # Para archivos .xlsm se usa keep_vba=True.
            if file.endswith(".xlsm"):
                wb = openpyxl.load_workbook(file_path, keep_vba=True, data_only=False)
            else:
                wb = openpyxl.load_workbook(file_path, data_only=False)
                
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                wb.remove(ws)
                wb.save(file_path)
                log_text.insert(tk.END, f"Se eliminó la hoja '{sheet_name}' en {file}\n")
            else:
                log_text.insert(tk.END, f"La hoja '{sheet_name}' no se encontró en {file}\n")
        except Exception as e:
            log_text.insert(tk.END, f"Error al procesar {file}: {e}\n")
    
    log_text.insert(tk.END, "Proceso culminado.\n")
    messagebox.showinfo("Proceso finalizado", "Proceso culminado.")

# Configuración de la interfaz gráfica
root = tk.Tk()
root.title("Eliminar Hoja de Excel")

# Fila 0: Selección de carpeta
tk.Label(root, text="Carpeta:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
folder_entry = tk.Entry(root, width=50)
folder_entry.grid(row=0, column=1, padx=5, pady=5)
tk.Button(root, text="Seleccionar carpeta", command=select_folder).grid(row=0, column=2, padx=5, pady=5)

# Fila 1: Ingreso del nombre de la hoja a borrar
tk.Label(root, text="Nombre de la hoja a borrar:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
sheet_entry = tk.Entry(root, width=50)
sheet_entry.grid(row=1, column=1, padx=5, pady=5)

# Fila 2: Botón para iniciar el proceso
tk.Button(root, text="Procesar", command=process_files).grid(row=2, column=1, padx=5, pady=10)

# Fila 3: Área de log para mostrar mensajes del proceso
log_text = tk.Text(root, height=10, width=70)
log_text.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

root.mainloop()
